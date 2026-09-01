from django.shortcuts import render, redirect, get_object_or_404
from .forms import *
import os
from django.conf import settings
from django.http import FileResponse
from django.db.models import Sum
from .pdf_tc_service_ue import generate_bulletin_lmd_pdf
from django.core.paginator import Paginator
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .forms import EtudiantDroitForm ,EtudiantGestionForm,UEForm,ECUEForm,MasterEtudiantForm ,QHSEEtudiantForm ,QHSEECUEForm
from .pdf_gestion_service import generer_bulletin_gestion_pdf
from .services import generer_bulletin_lmd_pdf
from .pdf_tronc_commun_service import generer_bulletin_tronc_commun_pdf
from .models import MasterUE,EtudiantMaster , CandidatRattrapage ,FiliereLMD
from .models import MasterECUE, NoteMaster
from openpyxl import load_workbook
from .services import (
    generer_bulletin_licence2_droit_prive_pdf,
)
from django.urls import reverse
from reportlab.lib import colors
from .models import (
    EtudiantLMD,
    NoteLMD,
    SessionAcademique
)
from .pdf_licence_qhse import generer_bulletin_qhse_pdf
from .models import UE, ECUE, NoteLMD, GrandeUnite
import pandas as pd
from .pdf_licence_qhse import generer_bulletin_qhse_pdf
from django.contrib.auth.decorators import login_required
from core.decorators import role_required

def niveau_list(request):
    niveaux = Niveau.objects.all()
    return render(request, "lmd/niveaux/list.html", {"niveaux": niveaux})


def niveau_add(request):
    if request.method == "POST":
        nom = request.POST.get("nom")
        Niveau.objects.create(nom=nom)
        return redirect("niveau_lmd_list")

    return render(request, "lmd/niveaux/add.html")

def filiere_list(request):
    filieres = FiliereLMD.objects.all()
    return render(request, "lmd/filieresLMD/list.html", {"filieres": filieres})


def filiere_add(request):
    if request.method == "POST":
        Filiere.objects.create(
            nom=request.POST.get("nom")
        )
    return redirect("filiere_lmd_list")


def filiere_delete(request, pk):
    filiere = get_object_or_404(FiliereLMD, pk=pk)
    filiere.delete()
    return redirect("filiere_list")



# UPDATE
def filiere_edit(request, pk):
    filiere = get_object_or_404(FiliereLMD, pk=pk)

    if request.method == "POST":
        filiere.code = request.POST.get("code")
        filiere.libelle = request.POST.get("libelle")
        filiere.save()
        return redirect("filiere_list")

    return render(request, "lmd/filieresLMD/edit.html", {
        "filiere": filiere
    })

def classe_list(request):
    classes = Classe.objects.all()
    return render(request, "lmd/classes/list.html", {
        "classes": classes
    })


def classe_add(request):
    if request.method == "POST":
        nom = request.POST.get("nom")
        niveau=request.POST.get("niveau"),
        filiere_id = request.POST.get("filiere")

        Classe.objects.create(
            nom=nom,
            niveau=request.POST.get("niveau"),   # 🔥 IMPORTANT
            filiere_id=filiere_id
        )
        return redirect("classe_lmd_list")

    return render(request, "lmd/classes/add.html", {
        "niveaux": Niveau.objects.all(),
        "filieres": FiliereLMD.objects.all()
    })


def ue_list(request):

    ues = UE.objects.select_related(
        "filiere"
    ).order_by(
        "filiere__libelle",
        "semestre",
        "code"
    )

    return render(
        request,
        "lmd/ue/list.html",
        {
            "ues": ues
        }
    )


def ue_add(request):

    if request.method == "POST":
        try:
            filiere_id = request.POST.get("filiere")
            grande_unite_id = request.POST.get("grande_unite")

            # 🔥 Vérification des ForeignKey (évite FOREIGN KEY constraint failed)
            filiere = get_object_or_404(FiliereLMD, id=filiere_id)
            grande_unite = get_object_or_404(GrandeUnite, id=grande_unite_id)

            UE.objects.create(
                code=request.POST.get("code"),
                libelle=request.POST.get("libelle"),
                credit=int(request.POST.get("credit", 0)),
                semestre=request.POST.get("semestre"),
                filiere=filiere,
                grande_unite=grande_unite,
            )

            messages.success(request, "UE ajoutée avec succès.")
            return redirect("ue_list")

        except ValueError:
            messages.error(request, "Le crédit doit être un nombre valide.")

        except Exception as e:
            messages.error(request, f"Erreur : {e}")

    return render(
        request,
        "lmd/ue/add.html",
        {
            "filieres": FiliereLMD.objects.all(),
            "grandes_unites": GrandeUnite.objects.all(),
        }
    )


def ue_edit(request, pk):

    ue = get_object_or_404(UE, pk=pk)

    if request.method == "POST":
        try:
            filiere_id = request.POST.get("filiere")
            grande_unite_id = request.POST.get("grande_unite")

            # 🔥 Sécurisation ForeignKey
            filiere = get_object_or_404(FiliereLMD, id=filiere_id)
            grande_unite = get_object_or_404(GrandeUnite, id=grande_unite_id)

            ue.code = request.POST.get("code")
            ue.libelle = request.POST.get("libelle")
            ue.credit = int(request.POST.get("credit", 0))
            ue.semestre = request.POST.get("semestre")

            ue.filiere = filiere
            ue.grande_unite = grande_unite

            ue.save()

            messages.success(request, "UE modifiée avec succès.")
            return redirect("ue_list")

        except ValueError:
            messages.error(request, "Le crédit doit être un nombre valide.")

        except Exception as e:
            messages.error(request, f"Erreur : {e}")

    return render(request, "lmd/ue/edit.html", {
        "ue": ue,
        "filieres": FiliereLMD.objects.all(),
        "grandes_unites": GrandeUnite.objects.all(),
    })

def ue_delete(request, pk):

    ue = UE.objects.get(pk=pk)

    if request.method == "POST":
        ue.delete()
        return redirect("ue_list")

    return render(
        request,
        "lmd/ue/delete.html",
        {
            "ue": ue
        }
    )

# =====================
# LISTE + FILTRE
# =====================
def ecue_list(request):

    ue_id = request.GET.get("ue")

    ecues = ECUE.objects.select_related("ue").all()

    if ue_id:
        ecues = ecues.filter(ue_id=ue_id)

    total_credits = ecues.aggregate(
     total=Sum("credit")
     )["total"] or 0

    return render(request, "lmd/ecue/list.html", {
        "ecues": ecues,
        "ues": UE.objects.all(),
        "ue_selected": ue_id,
        "total_credits": total_credits,
    })


# =====================
# CREATE
# =====================
def ecue_add(request):

    if request.method == "POST":

        ECUE.objects.create(
            ue_id=request.POST.get("ue"),
            code=request.POST.get("code"),
            libelle=request.POST.get("libelle"),
            coefficient=request.POST.get("coefficient"),
            credit=request.POST.get("credit")
        )

        return redirect("ecue_list")

    return render(request, "lmd/ecue/add.html", {
        "ues": UE.objects.all()
    })


# =====================
# UPDATE
# =====================
def ecue_edit(request, pk):

    ecue = get_object_or_404(ECUE, pk=pk)

    if request.method == "POST":

        ecue.ue_id = request.POST.get("ue")
        ecue.code = request.POST.get("code")
        ecue.libelle = request.POST.get("libelle")
        ecue.coefficient = request.POST.get("coefficient")
        ecue.credit = request.POST.get("credit")

        ecue.save()

        return redirect("ecue_list")

    return render(request, "lmd/ecue/edit.html", {
        "ecue": ecue,
        "ues": UE.objects.all()
    })


# =====================
# DELETE
# =====================
def ecue_delete(request, pk):

    ecue = get_object_or_404(ECUE, pk=pk)

    if request.method == "POST":
        ecue.delete()
        return redirect("ecue_list")

    return render(request, "lmd/ecue/delete.html", {
        "ecue": ecue
    })


from django.db.models import Q
from .models import NoteLMD, EtudiantLMD, ECUE


def note_lmd_list(request):

    notes = NoteLMD.objects.select_related(
        "etudiant",
        "ecue"
    )

    etudiant = request.GET.get("etudiant")
    ecue = request.GET.get("ecue")

    if etudiant:
        notes = notes.filter(
            Q(etudiant__nom__icontains=etudiant) |
            Q(etudiant__prenoms__icontains=etudiant) |
            Q(etudiant__matricule__icontains=etudiant)
        )

    if ecue:
        notes = notes.filter(ecue__code__icontains=ecue)

    return render(request, "lmd/notes/list.html", {
        "notes": notes,
        "ecues": ECUE.objects.all(),
    })


def note_lmd_add(request):

    if request.method == "POST":
        try:
            etudiant_id = request.POST.get("etudiant")
            ecue_id = request.POST.get("ecue")
            semestre = request.POST.get("semestre")
            session = request.POST.get("session")

            cc = float(request.POST.get("cc", "0").replace(",", "."))
            examen = float(request.POST.get("examen", "0").replace(",", "."))

            NoteLMD.objects.create(
                etudiant_id=etudiant_id,
                ecue_id=ecue_id,
                semestre=semestre,
                session=session,
                cc=cc,
                examen=examen,
            )

            messages.success(request, "La note a été enregistrée avec succès.")
            return redirect("note_lmd_list")

        except ValueError:
            messages.error(
                request,
                "Les notes de CC et d'examen doivent être des nombres valides."
            )

        except Exception as e:
            messages.error(request, f"Erreur : {e}")

    return render(
        request,
        "lmd/notes/form.html",
        {
            "etudiants": EtudiantLMD.objects.all(),
            "ecues": ECUE.objects.all(),
        },
    )


def note_lmd_edit(request, pk):

    note = get_object_or_404(NoteLMD, pk=pk)

    if request.method == "POST":
        try:
            note.etudiant_id = request.POST.get("etudiant")
            note.ecue_id = request.POST.get("ecue")
            note.semestre = request.POST.get("semestre")
            note.session = request.POST.get("session")

            # ✅ Conversion obligatoire en float
            note.cc = float(request.POST.get("cc", "0").replace(",", "."))
            note.examen = float(request.POST.get("examen", "0").replace(",", "."))

            note.save()

            messages.success(request, "Note modifiée avec succès.")
            return redirect("note_lmd_list")

        except ValueError:
            messages.error(request, "CC et Examen doivent être des nombres valides.")

    return render(request, "lmd/notes/form.html", {
        "note": note,
        "etudiants": EtudiantLMD.objects.all(),
        "ecues": ECUE.objects.all(),
    })
def note_lmd_delete(request, pk):

    note = NoteLMD.objects.get(pk=pk)
    note.delete()

    return redirect("note_lmd_list")

def bulletin_lmd_listRE(request):

    etudiants = EtudiantLMD.objects.select_related(
        "user",
        "filiere"
    )

    matricule = request.GET.get("matricule")
    nom = request.GET.get("nom")
    telephone = request.GET.get("telephone")
    classe = request.GET.get("classe")

    if matricule:
        etudiants = etudiants.filter(matricule__icontains=matricule)

    if nom:
        etudiants = etudiants.filter(
            Q(nom__icontains=nom) |
            Q(prenoms__icontains=nom)
        )

    if telephone:
        etudiants = etudiants.filter(telephone__icontains=telephone)

    if classe:
        etudiants = etudiants.filter(classe_id=classe)

    return render(request, "lmd/bulletins/list.html", {
        "etudiants": etudiants,
        "classes": ClasseLMD.objects.select_related("filiere"),
    })
    
    
@login_required(login_url="login")
@role_required("ADMIN")
def bulletin_lmd_list(request):

    matricule = request.GET.get("matricule")
    nom = request.GET.get("nom")
    telephone = request.GET.get("telephone")
    classe = request.GET.get("classe")

    etudiants = EtudiantLMD.objects.select_related(
        "user",
        "filiere"
    ).all()

    if matricule:
        etudiants = etudiants.filter(matricule__icontains=matricule)

    if nom:
        etudiants = etudiants.filter(
            Q(nom__icontains=nom) |
            Q(prenoms__icontains=nom)
        )

    if telephone:
        etudiants = etudiants.filter(telephone__icontains=telephone)

    if classe:
        etudiants = etudiants.filter(classe_id=classe)

    return render(request, "lmd/bulletins/list.html", {
        "etudiants": etudiants,
        "classes": ClasseLMD.objects.select_related("filiere"),
        
    })


def etudiant_lmd_addENC(request):

    if request.method == "POST":

        etudiant = EtudiantLMD.objects.create(
            matricule=request.POST.get("matricule"),
            nom=request.POST.get("nom"),
            prenoms=request.POST.get("prenoms"),
            sexe=request.POST.get("sexe"),
            date_naissance=request.POST.get("date_naissance"),
            telephone=request.POST.get("telephone"),
            email=request.POST.get("email"),
            niveau=request.POST.get("niveau"),   # 🔥 IMPORTANT
            filiere_id=request.POST.get("filiere"),
            annee_academique=request.POST.get("annee_academique"),
        )

        # 🔥 UE automatique selon filière
        ues = UE.objects.filter(filiere=etudiant.filiere)

        return redirect("etudiant_lmd_list")

    return render(request, "lmd/etudiants/add.html", {
        "niveaux": Niveau.objects.all(),
        "filieres": FiliereLMD.objects.all(),
        "ues": UE.objects.all(),
        "ecues": ECUE.objects.all(),
    })

def etudiant_lmd_addens(request):

    if request.method == "POST":

        from datetime import datetime

        date_naissance = request.POST.get("date_naissance")

        if date_naissance:
            try:
                date_naissance = datetime.strptime(date_naissance, "%Y-%m-%d").date()
            except ValueError:
                date_naissance = None
        else:
            date_naissance = None

        etudiant = EtudiantLMD.objects.create(
            matricule=request.POST.get("matricule"),
            nom=request.POST.get("nom"),
            prenoms=request.POST.get("prenoms"),
            sexe=request.POST.get("sexe"),
            date_naissance=date_naissance,
            telephone=request.POST.get("telephone"),
            email=request.POST.get("email"),
            niveau=request.POST.get("niveau"),
            filiere_id=request.POST.get("filiere"),
            annee_academique=request.POST.get("annee_academique"),
        )

        # ✔ récupération UE / ECUE
        ue_ids = request.POST.getlist("ue")
        ecue_ids = request.POST.getlist("ecue")

        # ✔ liaison (si ManyToMany)
        etudiant.ues.set(UE.objects.filter(id__in=ue_ids))
        etudiant.ecues.set(ECUE.objects.filter(id__in=ecue_ids))

        return redirect("etudiant_lmd_list")

    return render(request, "lmd/etudiants/add.html", {
        "niveaux": Niveau.objects.all(),
        "filieres": FiliereLMD.objects.all(),
        "ues": UE.objects.all(),
        "ecues": ECUE.objects.all(),
    })
def etudiant_lmd_add(request):

    if request.method == "POST":

        from datetime import datetime

        matricule = request.POST.get("matricule")
        annee_academique = request.POST.get("annee_academique")


        # ==========================
        # Vérification doublon
        # ==========================
        if EtudiantLMD.objects.filter(
            matricule=matricule
        ).exists():

            messages.error(
                request,
                "Ce matricule existe déjà."
            )

            return redirect("etudiant_lmd_add")


        # ==========================
        # Conversion date naissance
        # ==========================
        date_naissance = request.POST.get("date_naissance")

        if date_naissance:
            try:
                date_naissance = datetime.strptime(
                    date_naissance,
                    "%Y-%m-%d"
                ).date()

            except ValueError:
                date_naissance = None

        else:
            date_naissance = None



        # ==========================
        # Création étudiant
        # ==========================
        EtudiantLMD.objects.create(

            matricule=matricule,

            nom=request.POST.get("nom"),

            prenoms=request.POST.get("prenoms"),

            sexe=request.POST.get("sexe"),

            statut=request.POST.get("statut", "NF"),

            date_naissance=date_naissance,

            lieu_naissance=request.POST.get(
                "lieu_naissance"
            ),

            telephone=request.POST.get(
                "telephone"
            ),

            email=request.POST.get(
                "email"
            ),

            niveau=request.POST.get(
                "niveau"
            ),

            filiere_id=request.POST.get(
                "filiere"
            ),

            annee_academique=annee_academique,

        )


        messages.success(
            request,
            "Étudiant ajouté avec succès."
        )


        return redirect(
            "etudiant_lmd_list"
        )


    return render(
        request,
        "lmd/etudiants/add.html",
        {
            "niveaux": NiveauLMD.objects.all(),
            "filieres": FiliereLMD.objects.all(),
        }
    )
def calcul_moyenne_ecue(note):
    if not note:
        return 0

    cc = note.cc or 0
    examen = note.examen or 0

    return round((cc * 0.4) + (examen * 0.6), 2)

def calcul_moyenne_ue(etudiant, ue):

    ecues = ue.ecues.all()

    notes = NoteLMD.objects.filter(
        etudiant=etudiant,
        ecue__ue=ue
    )

    total = 0
    count = 0

    for note in notes:
        total += calcul_moyenne_ecue(note)
        count += 1

    if count == 0:
        return 0

    return round(total / count, 2)
def ue_validee(moyenne_ue):
    return moyenne_ue >= 10

def etudiant_lmd_editDDD(request, pk):
    etudiant = EtudiantLMD.objects.get(pk=pk)
    return render(request, "lmd/etudiants/edit.html", {"etudiant": etudiant})


def etudiant_lmd_delete(request, pk):
    etudiant = get_object_or_404(EtudiantLMD, pk=pk)

    if request.method == "POST":
        etudiant.delete()

    return redirect("etudiant_lmd_list")

def etudiant_lmd_update(request, pk):
    etudiant = get_object_or_404(EtudiantLMD, pk=pk)

    if request.method == "POST":
        try:
            matricule = request.POST.get("matricule")

            if EtudiantLMD.objects.exclude(id=etudiant.id).filter(matricule=matricule).exists():
                return redirect("etudiant_lmd_list")

            etudiant.matricule = matricule or ""
            etudiant.nom = request.POST.get("nom") or ""
            etudiant.prenoms = request.POST.get("prenoms") or ""
            etudiant.telephone = request.POST.get("telephone") or ""
            etudiant.email = request.POST.get("email") or ""

            etudiant.sexe = request.POST.get("sexe") or "M"
            etudiant.niveau = request.POST.get("niveau") or "L1"
            etudiant.statut = request.POST.get("statut") or "AF"
            etudiant.annee_academique = request.POST.get("annee_academique") or ""

            filiere_id = request.POST.get("filiere")
            if filiere_id:
                etudiant.filiere_id = filiere_id

            date_naissance = request.POST.get("date_naissance")
            if date_naissance:
                etudiant.date_naissance = date_naissance

            etudiant.save()

            # =========================
            # UE / ECUE
            # =========================
            ue_ids = request.POST.getlist("ues")
            ecue_ids = request.POST.getlist("ecue")

            etudiant.ues.set(ue_ids)
            etudiant.ecues.set(ecue_ids)

        except Exception as e:
            print("ERROR UPDATE ETUDIANT:", e)

    return redirect("etudiant_lmd_list")


def resultat_ue(request, etudiant_id):

    from .models import EtudiantLMD, UE

    etudiant = EtudiantLMD.objects.get(id=etudiant_id)
    ues = UE.objects.all()

    resultats = []

    for ue in ues:
        moyenne = calcul_moyenne_ue(etudiant, ue)

        resultats.append({
            "ue": ue,
            "moyenne": moyenne,
            "statut": statut_ue(moyenne)
        })

    return render(request, "lmd/resultats_ue.html", {
        "etudiant": etudiant,
        "resultats": resultats
    })
@login_required(login_url="login")
@role_required("ADMIN")
def bulletin_lmd_pdf(request, etudiant_id):

    file_path = os.path.join(
        settings.BASE_DIR,
        f"bulletin_{etudiant_id}.pdf"
    )

    generate_bulletin_lmd_pdf(etudiant_id, file_path)

    return FileResponse(open(file_path, "rb"))


def etudiant_lmd_list(request):

    etudiants = EtudiantLMD.objects.select_related(
        "user",
        "filiere"
    )

    matricule = request.GET.get("matricule")
    nom = request.GET.get("nom")
    telephone = request.GET.get("telephone")
    ue = request.GET.get("ue")

    if matricule:
        etudiants = etudiants.filter(
            matricule__icontains=matricule
        )

    if nom:
        etudiants = etudiants.filter(
            Q(nom__icontains=nom) |
            Q(prenoms__icontains=nom)
        )

    if telephone:
        etudiants = etudiants.filter(
            telephone__icontains=telephone
        )

    if ue:
        etudiants = etudiants.filter(
            note_lmd__ecue__code__icontains=ue
        ).distinct()

    # Trier les étudiants
    etudiants = etudiants.order_by("nom")

    # Pagination (10 étudiants par page)
    paginator = Paginator(etudiants, 10)

    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "lmd/etudiants/list.html", {
        "page_obj": page_obj,
         "ues": UE.objects.all(),
         "ecues": ECUE.objects.all(),
        "filieres": FiliereLMD.objects.all(),
    })



def note_lmdecue_add(request):

    filieres = FiliereLMD.objects.all()
    ecues = ECUE.objects.all()

    if request.method == "POST":

        filiere_id = request.POST["filiere"]
        niveau = request.POST["niveau"]
        semestre = request.POST["semestre"]
        session = request.POST["session"]
        ecue_id = request.POST["ecue"]

        etudiants = (
            EtudiantLMD.objects
            .filter(
                filiere_id=filiere_id,
                niveau=niveau
            )
            .order_by("nom", "prenoms")
        )

        return render(request,
            "lmd/notes/saisie_notes.html",
            {
                "etudiants": etudiants,
                "ecue_id": ecue_id,
                "semestre": semestre,
                "session": session,
            }
        )

    return render(
        request,
        "lmd/notes/choix.html",
        {
            "filieres": filieres,
            "ecues": ecues,
        }
    )

def note_lmd_save_batch(request):

    if request.method == "POST":

        ecue_id = request.POST.get("ecue_id")
        semestre = request.POST.get("semestre")
        session = request.POST.get("session")

        etudiants_ids = request.POST.getlist("etudiant_id")

        for etu_id in etudiants_ids:

            NoteLMD.objects.update_or_create(
                etudiant_id=etu_id,
                ecue_id=ecue_id,
                semestre=semestre,
                session=session,
                defaults={
                    "cc": request.POST.get(f"cc_{etu_id}"),
                    "examen": request.POST.get(f"examen_{etu_id}"),
                }
            )

        return redirect("note_lmd_list")


def note_lmd_listecue(request):

    notes = (
        NoteLMD.objects
        .select_related(
            "etudiant",
            "ecue",
            "etudiant__filiere",
        )
        .order_by(
            "etudiant__filiere__nom",
            "etudiant__niveau",
            "etudiant__nom",
        )
    )

    return render(
        request,
        "lmd/notes/listecue.html",
        {
            "notes": notes
        }
    )

def saisie_list(request):
    saisies = SaisieNoteLMD.objects.select_related(
        "filiere", "ecue"
    ).order_by("-date_creation")

    return render(request, "lmd/saisies/list.html", {
        "saisies": saisies
    })

def saisie_add(request):

    filieres = FiliereLMD.objects.all()
    niveaux = Niveau.objects.all()
    ecues = ECUE.objects.all()

    if request.method == "POST":
        SaisieNoteLMD.objects.create(
            filiere_id=request.POST.get("filiere"),
            niveau=request.POST.get("niveau"),
            ecue_id=request.POST.get("ecue"),
            semestre=request.POST.get("semestre"),
            session=request.POST.get("session"),
            created_by=request.user
        )

        return redirect("saisie_list")

    return render(request, "lmd/saisies/form.html", {
        "filieres": filieres,
        "niveaux": niveaux,
        "ecues": ecues,
    })



def saisie_edit(request, pk):

    saisie = get_object_or_404(SaisieNoteLMD, pk=pk)

    filieres = FiliereLMD.objects.all()
    ecues = ECUE.objects.all()

    if request.method == "POST":

        filiere_id = request.POST.get("filiere")
        ecue_id = request.POST.get("ecue")

        saisie.filiere = get_object_or_404(FiliereLMD, id=filiere_id)
        saisie.ecue = get_object_or_404(ECUE, id=ecue_id)

        saisie.niveau = request.POST.get("niveau")
        saisie.semestre = request.POST.get("semestre")
        saisie.session = request.POST.get("session")

        saisie.save()

        return redirect("saisie_list")

    return render(request, "lmd/saisies/edit.html", {
        "saisie": saisie,
        "filieres": filieres,
        "ecues": ecues
    })

def saisie_delete(request, pk):

    saisie = SaisieNoteLMD.objects.get(pk=pk)

    if request.method == "POST":
        saisie.delete()
        return redirect("saisie_list")

    return render(request, "lmd/saisies/delete.html", {
        "saisie": saisie
    })

def saisie_detail(request, pk):

    saisie = SaisieNoteLMD.objects.get(pk=pk)

    etudiants = EtudiantLMD.objects.filter(
        filiere=saisie.filiere,
        niveau=saisie.niveau
    ).order_by("nom")

    return render(request, "lmd/saisies/detail.html", {
        "saisie": saisie,
        "etudiants": etudiants
    })

    

# views.py



def filiereLMD_list(request):
    filieres = FiliereLMD.objects.all()
    return render(request, "lmd/filieresLMD/list.html", {"filieres": filieres})

def filiereLMD_add(request):

    if request.method == "POST":
        code = request.POST.get("code")
        libelle = request.POST.get("libelle")

        if not code or not libelle:
            messages.error(request, "Tous les champs sont obligatoires.")
            return redirect("filiere_add")

        if FiliereLMD.objects.filter(code=code).exists():
            messages.error(request, "Ce code existe déjà.")
            return redirect("filiere_add")

        FiliereLMD.objects.create(
            code=code,
            libelle=libelle
        )

        messages.success(request, "Filière ajoutée avec succès.")
        return redirect("filiere_list")

    return render(request, "lmd/filieresLMD/add.html")



from django.db import transaction

def saisie_note_etudiant(request, pk):

    saisie = get_object_or_404(SaisieNoteLMD, pk=pk)

    # étudiants concernés (filtre filière + niveau)
    etudiants = EtudiantLMD.objects.filter(
        filiere=saisie.filiere,
        niveau=saisie.niveau
    ).order_by("nom", "prenoms")

     # Aucun étudiant trouvé
    aucun_etudiant = not etudiants.exists()

    if request.method == "POST":

        with transaction.atomic():

            for etudiant in etudiants:

                note_value = request.POST.get(f"note_{etudiant.id}")

                if note_value != "" and note_value is not None:

                    NoteLMD.objects.update_or_create(
                        etudiant=etudiant,
                        ecue=saisie.ecue,
                        semestre=saisie.semestre,
                        session=saisie.session,
                        defaults={
                            "note": note_value,
                            "saisie": saisie
                        }
                    )

        return redirect("saisie_detail", pk=saisie.id)

    # récupérer notes existantes
    notes_existantes = {
        n.etudiant.id: {
        "cc": n.cc,
        "examen": n.examen
        }
        # n.etudiant.id: n.note
        for n in NoteLMD.objects.filter(
            ecue=saisie.ecue,
            semestre=saisie.semestre,
            session=saisie.session
        )
    }

    context = {
        "saisie": saisie,
        "etudiants": etudiants,
        "notes_existantes": notes_existantes,
    }

    return render(request, "lmd/saisie_note_etudiant.html", context)
    

from django.db import transaction


def enregistrer_notes(request, pk):
    saisie = get_object_or_404(SaisieNoteLMD, pk=pk)

    etudiants = EtudiantLMD.objects.filter(
        filiere=saisie.filiere,
        niveau=saisie.niveau
    )

    if request.method == "POST":
        with transaction.atomic():
            for etudiant in etudiants:

                cc = request.POST.get(f"cc_{etudiant.id}")
                examen = request.POST.get(f"examen_{etudiant.id}")

                try:
                    cc = float(cc)
                except (TypeError, ValueError):
                    cc = 0

                try:
                    examen = float(examen)
                except (TypeError, ValueError):
                    examen = 0

                # Ignore si aucune note
                if cc == 0 and examen == 0:
                    continue

                NoteLMD.objects.update_or_create(
                    etudiant=etudiant,
                    ecue=saisie.ecue,
                    semestre=saisie.semestre,
                    session=saisie.session,
                    defaults={
                        "cc": cc,
                        "examen": examen,
                    }
                )

        return redirect("saisie_detail", pk=saisie.id)

    notes = {
        n.etudiant_id: n
        for n in NoteLMD.objects.filter(
            ecue=saisie.ecue,
            semestre=saisie.semestre,
            session=saisie.session
        )
    }

    for etudiant in etudiants:
        etudiant.note = notes.get(etudiant.id)

    return render(request, "lmd/saisie_note_etudiant.html", {
        "saisie": saisie,
        "etudiants": etudiants,
    })

def filiereLMD_edit(request, pk):
    filiere = get_object_or_404(FiliereLMD, pk=pk)

    if request.method == "POST":
        code = request.POST.get("code")
        libelle = request.POST.get("libelle")

        # vérification doublon (optionnel mais recommandé)
        if FiliereLMD.objects.exclude(pk=pk).filter(code=code).exists():
            messages.error(request, "Ce code existe déjà.")
            return redirect("filiereLMD_edit", pk=pk)

        filiere.code = code
        filiere.libelle = libelle
        filiere.save()

        messages.success(request, "Filière modifiée avec succès.")
        return redirect("filiereLMD_list")

    return render(request, "lmd/filieresLMD/edit.html", {
        "filiere": filiere
    })

def filiereLMD_delete(request, pk):
    filiere = get_object_or_404(FiliereLMD, pk=pk)
    filiere.delete()
    return redirect("filiereLMD_list")


def ajouter_etudiants_saisie(request, saisie_id):

    saisie = get_object_or_404(
        SaisieNoteLMD,
        id=saisie_id
    )

    etudiants = EtudiantLMD.objects.filter(
        filiere=saisie.filiere,
        niveau=saisie.niveau
    )

    compteur = 0

    for etudiant in etudiants:

        existe = NoteLMD.objects.filter(
            etudiant=etudiant,
            ecue=saisie.ecue,
            semestre=saisie.semestre,
            session=saisie.session
        ).exists()

        if not existe:

            NoteLMD.objects.create(
                etudiant=etudiant,
                ecue=saisie.ecue,
                semestre=saisie.semestre,
                session=saisie.session,
                cc=0,
                examen=0
            )

            compteur += 1


    messages.success(
        request,
        f"{compteur} étudiant(s) ajouté(s)."
    )


    return redirect(
        "saisie_detail",
         pk=saisie.id
    )
    
def filiere_l3_detail(request, id):

    filiere = FiliereLMD.objects.get(id=id)

    ues = UE.objects.filter(
        filiere=filiere
    )

    etudiants = EtudiantLMD.objects.filter(
        filiere=filiere,
        niveau="L3"
    )

    return render(
        request,
        "lmd/l3_detail.html",
        {
            "filiere":filiere,
            "ues":ues,
            "etudiants":etudiants
        }
    )
    
@login_required
def filiere_master_detail(request,id):

    filiere = FiliereLMD.objects.get(id=id)


    ues = UE.objects.filter(
        filiere=filiere
    )


    etudiants = EtudiantLMD.objects.filter(
        filiere=filiere,
        niveau__in=["M1","M2"]
    )


    return render(
        request,
        "lmd/master_detail.html",
        {
            "filiere":filiere,
            "ues":ues,
            "etudiants":etudiants
        }
    )
    
def generer_rattrapages(etudiant):

    notes = NoteLMD.objects.filter(
        etudiant=etudiant,
        session="1"
    )

    for note in notes:
        
        if note.moyenne < 10:

            CandidatRattrapage.objects.create(

                etudiant=etudiant,

                ecue=note.ecue,

                ancienne_note=note.moyenne

            )
            
def meilleure_note(note1,note2):

    return max(
        note1,
        note2
    )
    

@login_required
def rattrapage_liste(request):

    semestre = request.GET.get("semestre")


    session_rattrapage = SessionAcademique.objects.filter(
        type_session="RATTRAPAGE",
        active=True
    ).first()


    if not session_rattrapage:
        return render(
            request,
            "lmd/rattrapage/liste.html",
            {
                "candidats": [],
                "semestre": semestre
            }
        )


    notes = NoteLMD.objects.filter(
        session="1",
        moyenne__lt=10
    ).select_related(
        "etudiant",
        "ecue"
    )


    for note in notes:

        candidat, created = CandidatRattrapage.objects.get_or_create(

            etudiant=note.etudiant,

            ecue=note.ecue,

            session=session_rattrapage,

            annee_academique="2025-2026",

            defaults={
                "ancienne_note": note.moyenne,
                "statut":"EN_ATTENTE"
            }
        )


        candidat.ancienne_note = note.moyenne
        candidat.save()



    candidats = CandidatRattrapage.objects.filter(
        session=session_rattrapage
    ).select_related(
        "etudiant",
        "ecue"
    )


    if semestre:

        candidats = candidats.filter(
            ecue__ue__semestre=semestre
        )


    return render(
        request,
        "lmd/rattrapage/liste.html",
        {
            "candidats": candidats,
            "semestre": semestre
        }
    )
    
@login_required
def saisie_rattrapage(request):

    session = SessionAcademique.objects.filter(
        type_session="RATTRAPAGE",
        active=True
    ).first()


    if not session:

        return render(
            request,
            "lmd/rattrapage/saisie.html",
            {
                "candidats": []
            }
        )


    candidats = CandidatRattrapage.objects.filter(
        session=session,
        statut="EN_ATTENTE"
    ).select_related(
        "etudiant",
        "ecue"
    )


    # ============================
    # ENREGISTREMENT DES NOTES
    # ============================

    if request.method == "POST":


        for candidat in candidats:


            note_key = f"note_{candidat.id}"


            nouvelle_note = request.POST.get(note_key)


            if nouvelle_note not in [None, ""]:


                nouvelle_note = float(nouvelle_note)


                candidat.nouvelle_note = nouvelle_note


                if nouvelle_note >= 10:

                    candidat.statut = "VALIDE"

                else:

                    candidat.statut = "ECHEC"


                candidat.save()



        messages.success(
            request,
            "Les notes de rattrapage ont été enregistrées."
        )


        return redirect(
            "rattrapage_liste"
        )



    return render(

        request,

        "lmd/rattrapage/saisie.html",

        {
            "candidats": candidats
        }

    )
    
    
@login_required
def liste_rattrapage(request):

    session_rattrapage = SessionAcademique.objects.filter(
        type_session="RATTRAPAGE",
        active=True
    ).first()


    if not session_rattrapage:
        return render(
            request,
            "lmd/rattrapage/liste.html",
            {
                "candidats": []
            }
        )


    # ===============================
    # Création automatique des candidats
    # depuis les notes session normale
    # ===============================

    notes = NoteLMD.objects.filter(
        session="1",
        moyenne__lt=10
    ).select_related(
        "etudiant",
        "ecue"
    )


    for note in notes:

        candidat, created = CandidatRattrapage.objects.get_or_create(

            etudiant=note.etudiant,

            ecue=note.ecue,

            session=session_rattrapage,

            annee_academique="2025-2026",

            defaults={
                "ancienne_note": note.moyenne,
                "statut": "EN_ATTENTE"
            }
        )


        # Mise à jour de l'ancienne note
        if not created:

            candidat.ancienne_note = note.moyenne

            candidat.save()



    # ===============================
    # Récupération des candidats
    # ===============================

    candidats = CandidatRattrapage.objects.filter(

        session=session_rattrapage

    ).select_related(

        "etudiant",
        "ecue",
        "session"

    ).order_by(

        "etudiant__nom",
        "ecue__libelle"

    )



    return render(

        request,

        "lmd/rattrapage/liste.html",

        {
            "candidats": candidats
        }

    )

def deliberation_rattrapage(request):

    candidats = CandidatRattrapage.objects.select_related(
        "etudiant",
        "ecue",
        "session"
    ).all()

    return render(
        request,
        "lmd/rattrapage/deliberation.html",
        {
            "candidats": candidats
        }
    )
    
@login_required(login_url="login")
@role_required("ADMIN")    
def bulletin_rattrapage_list(request):

    candidats = CandidatRattrapage.objects.select_related(
        "etudiant",
        "ecue",
        "session"
    ).all()
    
    return render(
        request,
        "lmd/rattrapage/bulletins.html",
        {
            "candidats": candidats
        }
    )



def l3_gestion_dashboard(request):

    return render(
        request,
        "lmd/l3/gestion/dashboard.html"
    )


def l3_droit_etudiantsGGGG(request):

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Droit Privé"
    )

    etudiants = EtudiantLMD.objects.filter(
        filiere=filiere,
        niveau="L3",
        actif=True
    ).order_by(
        "nom",
        "prenoms"
    )


    context = {
        "filiere": filiere,
        "etudiants": etudiants,
        "total_etudiants": etudiants.count(),
    }


    return render(
        request,
        "lmd/l3/droit/etudiants.html",
        context
    )

from django.db.models import Q
from django.shortcuts import get_object_or_404, render


def droit_prive_etudiants(request):

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Droit Privé"
    )


    # Tous les étudiants Droit Privé (L1, L2, L3)
    etudiants = EtudiantLMD.objects.filter(
        filiere=filiere,
        actif=True
    )


    # Recherche
    q = request.GET.get("q")

    if q:

        etudiants = etudiants.filter(
            Q(matricule__icontains=q)
            |
            Q(nom__icontains=q)
            |
            Q(prenoms__icontains=q)
        )


    # Filtre par niveau
    niveau = request.GET.get("niveau")

    if niveau:

        etudiants = etudiants.filter(
            niveau=niveau
        )


    # Tri
    etudiants = etudiants.order_by(
        "niveau",
        "nom",
        "prenoms"
    )


    context = {

        "filiere": filiere,

        "etudiants": etudiants,

        "total_etudiants": etudiants.count(),

    }


    return render(
        request,
         "lmd/l3/droit/etudiants.html",
        context
    )


def l3_droit_etudiant_add(request):

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Droit Privé"
    )


    if request.method=="POST":

        form = EtudiantDroitForm(request.POST)


        if form.is_valid():

            etudiant=form.save(commit=False)


            etudiant.filiere=filiere

            # etudiant.niveau="L3"

            etudiant.statut="AF"


            etudiant.save()


            return redirect(
                "droit_prive_etudiants"
            )


    else:

        form=EtudiantDroitForm()



    return render(
        request,
        "lmd/l3/droit/etudiant_form.html",
        {
            "form":form,
            "titre":"Ajouter étudiant en Droit Privé"
        }
    )

def l3_droit_etudiant_update(request,pk):

    etudiant=get_object_or_404(
        EtudiantLMD,
        pk=pk
    )

    form=EtudiantDroitForm(
        request.POST or None,
        instance=etudiant
    )

    if form.is_valid():

        form.save()

        return redirect(
            "droit_prive_etudiants"
        )

    return render(
        request,
        "lmd/l3/droit/etudiant_form.html",
        {
            "form":form,
            "titre":"Modifier étudiant"
        }
    )

def l3_droit_etudiant_delete(request,pk):

    etudiant=get_object_or_404(
        EtudiantLMD,
        pk=pk
    )


    if request.method=="POST":

        etudiant.delete()

        return redirect(
            "droit_prive_etudiants"
        )


    return render(
        request,
        "lmd/l3/droit/delete.html",
        {
            "etudiant":etudiant
        }
    )


from django.shortcuts import render, redirect, get_object_or_404
from .models import UE, ECUE, FiliereLMD



# =========================================================
# AJOUTER UNE UE - L3 DROIT PRIVÉ
# =========================================================
from django.shortcuts import render, get_object_or_404
from .models import UE, FiliereLMD


def l3_droit_ueAAA(request):

    # Récupérer la filière Droit Privé
    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Droit Privé"
    )


    # Récupérer les filtres depuis l'URL

    niveau = request.GET.get(
        "niveau",
        "L1"
    )


    semestre = request.GET.get(
        "semestre",
        "S1"
    )


    # Sécurité niveau

    if niveau not in ["L1", "L2", "L3"]:
        niveau = "L1"


    # Sécurité semestre

    if semestre not in [
        "S1",
        "S2",
        "S3",
        "S4",
        "S5",
        "S6"
    ]:
        semestre = "S1"



    # Récupération des UE correspondantes

    ues = (
        UE.objects
        .filter(
            filiere=filiere,
            niveau=niveau,
            semestre=semestre
        )
        .prefetch_related(
            "ecues"
        )
        .order_by(
            "code"
        )
    )



    context = {

        "filiere": filiere,

        "ues": ues,

        "niveau": niveau,

        "semestre": semestre,

    }


    return render(
        request,
        "lmd/l3/droit/ue.html",
        context
    )   
def l3_droit_ue(request):

    # =========================================================
    # 1. RÉCUPÉRER LA FILIÈRE DROIT PRIVÉ
    # =========================================================

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Droit Privé"
    )

    # =========================================================
    # 2. RÉCUPÉRER LES FILTRES
    # =========================================================

    niveau = request.GET.get("niveau", "L1")
    semestre = request.GET.get("semestre", "S1")

    # =========================================================
    # 3. SÉCURISER LE NIVEAU
    # =========================================================

    niveaux_valides = ["L1", "L2", "L3"]

    if niveau not in niveaux_valides:
        niveau = "L1"

    # =========================================================
    # 4. SÉCURISER LE SEMESTRE
    # =========================================================

    semestres_valides = [
        "S1",
        "S2",
        "S3",
        "S4",
        "S5",
        "S6",
    ]

    if semestre not in semestres_valides:
        semestre = "S1"

    # =========================================================
    # 5. RÉCUPÉRER LES UE
    # =========================================================

    ues = (
        UE.objects
        .filter(
            filiere=filiere,
            niveau=niveau,
            semestre=semestre
        )
        .prefetch_related("ecues")
        .order_by("code")
    )

    # =========================================================
    # 6. CONTEXTE
    # =========================================================

    context = {
        "filiere": filiere,
        "ues": ues,
        "niveau": niveau,
        "semestre": semestre,
    }

    # =========================================================
    # 7. AFFICHAGE
    # =========================================================

    return render(
        request,
        "lmd/l3/droit/ue.html",
        context
    )    

def l3_droit_ue_addAAAA(request):

    filiere = get_object_or_404(FiliereLMD, libelle="Droit Privé")

    if request.method == "POST":

        niveau = request.POST.get("niveau")
        semestre = request.POST.get("semestre")
        grande_unite_id = request.POST.get("grande_unite")

        UE.objects.create(
            code=request.POST.get("code"),
            libelle=request.POST.get("libelle"),
            credit=request.POST.get("credit"),
            niveau=niveau,
            semestre=semestre,
            filiere=filiere,
            grande_unite_id=grande_unite_id or None,
        )

        return redirect("l3_droit_ue")

    return render(
        request,
        "lmd/l3/droit/ue_form.html",
        {
            "titre": "Ajouter une UE - Droit Privé",
            "filiere": filiere,
            "grandes_unites": GrandeUnite.objects.filter(filiere=filiere),
        }
    )

def l3_droit_ue_add(request):

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Droit Privé"
    )

    if request.method == "POST":

        # =========================================================
        # RÉCUPÉRATION DES DONNÉES
        # =========================================================

        niveau = request.POST.get("niveau")
        semestre = request.POST.get("semestre")
        session = request.POST.get("session")
        code = request.POST.get("code")
        libelle = request.POST.get("libelle")
        credit = request.POST.get("credit")
        ordre = request.POST.get("ordre")
        grande_unite_id = request.POST.get("grande_unite")

        # =========================================================
        # CRÉATION DE L'UE
        # =========================================================

        UE.objects.create(
            code=code,
            libelle=libelle,
            credit=credit,
            niveau=niveau,
            semestre=semestre,
            session=session,
            ordre=ordre,
            filiere=filiere,
            grande_unite_id=grande_unite_id or None,
        )

        # =========================================================
        # REDIRECTION
        # =========================================================

        return redirect(
            f"/lmd/l3/droit/ue/?niveau={niveau}&semestre={semestre}&session={session}"
        )

    # =============================================================
    # AFFICHAGE DU FORMULAIRE
    # =============================================================

    niveau = request.GET.get("niveau", "")
    semestre = request.GET.get("semestre", "")
    session = request.GET.get("session", "1")

    grandes_unites = GrandeUnite.objects.filter(
        filiere=filiere
    ).order_by(
        "niveau",
        "semestre",
        "ordre"
    )

    return render(
        request,
        "lmd/l3/droit/ue_form.html",
        {
            "titre": "Ajouter une UE - Droit Privé",
            "filiere": filiere,
            "niveau": niveau,
            "semestre": semestre,
            "session": session,
            "grandes_unites": grandes_unites,
        }
    )


# =========================================================
# MODIFIER UNE UE - L3 DROIT PRIVÉ
# =========================================================

def l3_droit_ue_updateAAA(request, pk):

    ue = get_object_or_404(
        UE,
        pk=pk,
        filiere__libelle="Droit Privé"
    )

    if request.method == "POST":

        ue.code = request.POST.get("code")
        ue.libelle = request.POST.get("libelle")
        ue.credit = request.POST.get("credit")
        ue.semestre = request.POST.get("semestre")

        ue.save()

        return redirect("l3_droit_ue")

    return render(
        request,
        "lmd/l3/droit/ue_form.html",
        {
            "titre": "Modifier UE - L3 Droit Privé",
            "ue": ue,
        }
    )

def l3_droit_ue_update(request, pk):

    ue = get_object_or_404(
        UE,
        pk=pk,
        filiere__libelle="Droit Privé"
    )

    if request.method == "POST":

        ue.code = request.POST.get("code")
        ue.libelle = request.POST.get("libelle")
        ue.credit = request.POST.get("credit")
        ue.semestre = request.POST.get("semestre")
        ue.niveau = request.POST.get("niveau")

        ue.save()

        return redirect("l3_droit_ue")

    return render(
        request,
        "lmd/l3/droit/ue_form.html",
        {
            "titre": "Modifier une UE - Droit Privé",
            "ue": ue,
        }
    )
# =========================================================
# SUPPRIMER UNE UE - L3 DROIT PRIVÉ
# =========================================================

def l3_droit_ue_deleteAAA(request, pk):

    ue = get_object_or_404(
        UE,
        pk=pk,
        filiere__libelle="Droit Privé"
    )

    if request.method == "POST":
        ue.delete()
        return redirect("l3_droit_ue")

    return render(
        request,
        "lmd/l3/droit/ue_confirm_delete.html",
        {
            "ue": ue,
        }
    )

def l3_droit_ue_delete(request, pk):

    # =========================================================
    # 1. RÉCUPÉRER L'UE
    # =========================================================

    ue = get_object_or_404(
        UE,
        pk=pk,
        filiere__libelle="Droit Privé"
    )

    # =========================================================
    # 2. SUPPRESSION
    # =========================================================

    if request.method == "POST":

        # Conserver les filtres avant suppression
        niveau = ue.niveau
        semestre = ue.semestre

        ue.delete()

        # Retour à la liste
        return redirect(
            f"{reverse('l3_droit_ue')}?niveau={niveau}&semestre={semestre}"
        )

    # =========================================================
    # 3. PAGE DE CONFIRMATION
    # =========================================================

    return render(
        request,
        "lmd/l3/droit/ue_confirm_delete.html",
        {
            "ue": ue,
        }
    )
    
def l3_droit_ue_deleteAAZ(request, pk):

    ue = get_object_or_404(
        UE,
        pk=pk,
        filiere__libelle="Droit Privé"
    )

    if request.method == "POST":

        ue.delete()

        return redirect("l3_droit_ue")

    return render(
        request,
        "lmd/l3/droit/ue_confirm_delete.html",
        {
            "ue": ue,
        }
    )

def l3_droit_prive_notesQQQ(request):

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Droit Privé"
    )


    # récupération des filtres
    niveau = request.GET.get("niveau", "L1")
    semestre = request.GET.get("semestre", "S1")


    # étudiants du niveau choisi
    etudiants = EtudiantLMD.objects.filter(
        filiere=filiere,
        niveau=niveau
    ).order_by(
        "nom",
        "prenoms"
    )


    # ECUE du niveau + semestre
    ecues = ECUE.objects.filter(
        ue__filiere=filiere,
        ue__niveau=niveau,
        semestre=semestre
    ).order_by(
        "code"
    )


    # Notes existantes
    notes = NoteLMD.objects.filter(
        etudiant__in=etudiants,
        ecue__in=ecues
    )



    # dictionnaire des notes
    notes_dict = {}

    for note in notes:

        notes_dict[
            note.etudiant_id
        ] = {

            "cc": note.cc,

            "examen": note.examen,

            "ecue": note.ecue_id

        }



    # préparation pour le template
    etudiants_notes = []


    for etudiant in etudiants:


        ancienne_note = notes_dict.get(
            etudiant.id,
            {}
        )


        etudiants_notes.append({

            "etudiant": etudiant,

            "cc": ancienne_note.get("cc",""),

            "examen": ancienne_note.get("examen",""),

        })



    return render(
        request,
        "lmd/l3/droit/notes.html",
        {
            "niveau": niveau,
            "semestre": semestre,

            "ecues": ecues,

            "etudiants_notes": etudiants_notes,
        }
    )

def l3_droit_prive_noteseeeee(request):

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Droit Privé"
    )


    niveau = request.GET.get("niveau", "L1")
    semestre = request.GET.get("semestre", "S1")


    etudiants = EtudiantLMD.objects.filter(
        filiere=filiere,
        niveau=niveau
    ).order_by(
        "nom",
        "prenoms"
    )


    ecues = ECUE.objects.filter(
        ue__filiere=filiere,
        ue__niveau=niveau,
        semestre=semestre
    ).order_by(
        "code"
    )


    notes = NoteLMD.objects.filter(
        etudiant__in=etudiants,
        ecue__in=ecues
    )


    notes_dict = {}

    for note in notes:

        notes_dict[
            note.etudiant_id
        ] = note



    etudiants_notes = []


    for etudiant in etudiants:


        note = notes_dict.get(
            etudiant.id
        )


        etudiants_notes.append({

            "etudiant": etudiant,

            "cc": note.cc if note else "",

            "examen": note.examen if note else "",

        })



    return render(
        request,
        "lmd/l3/droit/notes.html",
        {
            "niveau": niveau,
            "semestre": semestre,
            "ecues": ecues,
            "etudiants_notes": etudiants_notes,
        }
    )


def l3_droit_prive_notes(request):

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Droit Privé"
    )


    niveau = request.GET.get("niveau", "L1")
    semestre = request.GET.get("semestre", "S1")


    # Etudiants
    etudiants = EtudiantLMD.objects.filter(
        filiere=filiere,
        niveau=niveau
    ).order_by(
        "nom",
        "prenoms"
    )


    # ECUE
    ecues = ECUE.objects.filter(
        ue__filiere=filiere,
        ue__niveau=niveau,
        ue__semestre=semestre
    ).order_by(
        "code"
    )


    # Notes existantes
    notes = NoteLMD.objects.filter(
        etudiant__in=etudiants,
        ecue__in=ecues
    )


    notes_dict = {}

    for note in notes:

        notes_dict[
            note.etudiant_id
        ] = note



    etudiants_notes = []


    for etudiant in etudiants:

        note = notes_dict.get(
            etudiant.id
        )


        etudiants_notes.append({

            "etudiant": etudiant,

            "cc": note.cc if note else "",

            "examen": note.examen if note else "",

        })


    print("====== DEBUG ======")
    print("Niveau :", niveau)
    print("Semestre :", semestre)
    print("Etudiants :", etudiants.count())
    print("ECUE :", ecues.count())
    print("===================")


    return render(
        request,
        "lmd/l3/droit/notes.html",
        {
            "niveau": niveau,
            "semestre": semestre,
            "ecues": ecues,
            "etudiants_notes": etudiants_notes,
        }
    )
def l3_droit_prive_notes_detail(request):

    notes = (
        NoteLMD.objects
        .filter(etudiant__filiere__libelle="Droit Privé")
        .select_related("etudiant", "ecue")
        .order_by(
            "etudiant__nom",
            "etudiant__prenoms",
            "ecue__code"
        )
    )

    return render(
        request,
        "lmd/l3/droit/detail_notes.html",
        {
            "notes": notes,
        }
    )

def l3_droit_notes(request):

    # =========================================================
    # FILIÈRE
    # =========================================================

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Droit Privé"
    )

    # =========================================================
    # NIVEAU
    # =========================================================

    niveau = request.GET.get("niveau", "L1")

    if niveau not in ["L1", "L2", "L3"]:
        niveau = "L1"

    # =========================================================
    # SEMESTRE
    # =========================================================

    semestre = request.GET.get("semestre", "S1")

    if semestre not in ["S1", "S2", "S3", "S4", "S5", "S6"]:
        semestre = "S1"

    # =========================================================
    # ÉTUDIANTS
    # =========================================================

    etudiants = (
        EtudiantLMD.objects
        .filter(
            filiere=filiere,
            niveau=niveau
        )
        .order_by(
            "nom",
            "prenoms"
        )
    )

    # =========================================================
    # NOTES EXISTANTES
    # =========================================================

    notes = (
        NoteLMD.objects
        .filter(
            etudiant__filiere=filiere,
            etudiant__niveau=niveau,
            semestre=semestre,
            session="1"
        )
        .select_related(
            "etudiant",
            "ecue",
            "ecue__ue"
        )
        .order_by(
            "etudiant__nom",
            "ecue__code"
        )
    )

    # =========================================================
    # ECUE DU SEMESTRE SÉLECTIONNÉ
    # =========================================================

    ecues = (
        ECUE.objects
        .filter(
            ue__filiere=filiere,
            ue__semestre=semestre
        )
        .select_related("ue")
        .order_by(
            "ue__ordre",
            "code"
        )
    )

    # =========================================================
    # CONTEXTE
    # =========================================================

    return render(
        request,
        "lmd/l3/droit/notes.html",
        {
            "etudiants": etudiants,
            "notes": notes,
            "ecues": ecues,
            "niveau": niveau,
            "semestre": semestre,
        }
    )

def droit_prive_notes_detail(request):

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Droit Privé"
    )


    niveau = request.GET.get(
        "niveau",
        "L1"
    )


    semestre = request.GET.get(
        "semestre",
        "S1"
    )



    notes = (
        NoteLMD.objects
        .filter(
            etudiant__filiere=filiere,
            etudiant__niveau=niveau,
            ecue__semestre=semestre
        )
        .select_related(
            "etudiant",
            "ecue"
        )
        .order_by(
            "etudiant__nom",
            "ecue__code"
        )
    )



    return render(
        request,
         "lmd/l3/droit/detail_notes.html",
        {
            "notes":notes,
            "niveau":niveau,
            "semestre":semestre,
        }
    )


     
def droit_prive_note_add(request):

    if request.method != "POST":
        return redirect("droit_prive_notes")

    # =========================================================
    # RÉCUPÉRATION DES DONNÉES
    # =========================================================

    ecue_id = request.POST.get("ecue_id")
    niveau = request.POST.get("niveau")
    semestre = request.POST.get("semestre")

    # Session 1 par défaut
    session = "1"

    # =========================================================
    # VÉRIFICATIONS
    # =========================================================

    if not ecue_id:
        messages.error(
            request,
            "Veuillez sélectionner une ECUE."
        )

        return redirect(
            f"{reverse('droit_prive_notes')}?"
            f"niveau={niveau}&semestre={semestre}"
        )

    if semestre not in ["S1", "S2", "S3", "S4", "S5", "S6"]:
        messages.error(
            request,
            "Semestre invalide."
        )

        return redirect("droit_prive_notes")

    # =========================================================
    # RÉCUPÉRER L'ECUE
    # =========================================================

    ecue = get_object_or_404(
        ECUE,
        id=ecue_id
    )

    # =========================================================
    # VÉRIFIER QUE L'ECUE CORRESPOND BIEN AU SEMESTRE
    # =========================================================

    if ecue.ue.semestre != semestre:
        messages.error(
            request,
            f"L'ECUE sélectionnée n'appartient pas au {semestre}."
        )

        return redirect(
            f"{reverse('droit_prive_notes')}?"
            f"niveau={niveau}&semestre={semestre}"
        )

    # =========================================================
    # PARCOURIR LES ÉTUDIANTS
    # =========================================================

    for key, value in request.POST.items():

        if not key.startswith("cc_"):
            continue

        etudiant_id = key.replace("cc_", "")

        # =====================================================
        # RÉCUPÉRER L'ÉTUDIANT
        # =====================================================

        etudiant = get_object_or_404(
            EtudiantLMD,
            id=etudiant_id
        )

        # =====================================================
        # CC
        # =====================================================

        cc_value = value.strip()

        if cc_value == "":
            cc = 0
        else:
            try:
                cc = float(cc_value)
            except (ValueError, TypeError):
                cc = 0

        # =====================================================
        # EXAMEN
        # =====================================================

        examen_value = request.POST.get(
            f"examen_{etudiant_id}",
            ""
        ).strip()

        if examen_value == "":
            examen = 0
        else:
            try:
                examen = float(examen_value)
            except (ValueError, TypeError):
                examen = 0

        # =====================================================
        # ENREGISTREMENT
        # =====================================================

        NoteLMD.objects.update_or_create(

            etudiant=etudiant,

            ecue=ecue,

            semestre=semestre,

            session=session,

            defaults={
                "cc": cc,
                "examen": examen,
            }
        )

    # =========================================================
    # MESSAGE
    # =========================================================

    messages.success(
        request,
        f"Les notes du {semestre} ont été enregistrées avec succès."
    )

    # =========================================================
    # RETOUR À LA PAGE AVEC LES FILTRES
    # =========================================================

    return redirect(
        f"{reverse('droit_prive_notes')}?"
        f"niveau={niveau}&semestre={semestre}"
    )
 

def droit_prive_note_addAAAA(request):

    if request.method != "POST":
        return redirect("droit_prive_notes")

    # =========================================================
    # RÉCUPÉRATION DES DONNÉES
    # =========================================================

    ecue_id = request.POST.get("ecue_id")
    niveau = request.POST.get("niveau")
    semestre = request.POST.get("semestre")

    # Session envoyée par le formulaire
    # Session 1 par défaut
    session = request.POST.get("session", "1")

    # =========================================================
    # VÉRIFICATION DU NIVEAU
    # =========================================================

    if niveau not in ["L1", "L2", "L3"]:
        messages.error(
            request,
            "Niveau invalide."
        )

        return redirect("droit_prive_notes")

    # =========================================================
    # VÉRIFICATION DU SEMESTRE
    # =========================================================

    if semestre not in [
        "S1",
        "S2",
        "S3",
        "S4",
        "S5",
        "S6"
    ]:
        messages.error(
            request,
            "Semestre invalide."
        )

        return redirect(
            f"{reverse('droit_prive_notes')}?"
            f"niveau={niveau}&semestre=S1&session={session}"
        )

    # =========================================================
    # VÉRIFICATION DE LA SESSION
    # =========================================================

    if session not in ["1", "2"]:
        messages.error(
            request,
            "Session invalide."
        )

        return redirect(
            f"{reverse('droit_prive_notes')}?"
            f"niveau={niveau}&semestre={semestre}&session=1"
        )

    # =========================================================
    # VÉRIFICATION DE L'ECUE
    # =========================================================

    if not ecue_id:

        messages.error(
            request,
            "Veuillez sélectionner une ECUE."
        )

        return redirect(
            f"{reverse('droit_prive_notes')}?"
            f"niveau={niveau}&semestre={semestre}"
            f"&session={session}"
        )

    # =========================================================
    # RÉCUPÉRER L'ECUE
    # =========================================================

    ecue = get_object_or_404(
        ECUE,
        id=ecue_id
    )

    # =========================================================
    # VÉRIFIER LE NIVEAU DE L'ECUE
    # =========================================================

    if ecue.ue.niveau != niveau:

        messages.error(
            request,
            "L'ECUE sélectionnée ne correspond pas au niveau choisi."
        )

        return redirect(
            f"{reverse('droit_prive_notes')}?"
            f"niveau={niveau}&semestre={semestre}"
            f"&session={session}"
        )

    # =========================================================
    # VÉRIFIER LE SEMESTRE DE L'ECUE
    # =========================================================

    if ecue.ue.semestre != semestre:

        messages.error(
            request,
            f"L'ECUE sélectionnée n'appartient pas au {semestre}."
        )

        return redirect(
            f"{reverse('droit_prive_notes')}?"
            f"niveau={niveau}&semestre={semestre}"
            f"&session={session}"
        )

    # =========================================================
    # PARCOURIR LES ÉTUDIANTS
    # =========================================================

    nombre_notes = 0

    for key, value in request.POST.items():

        # On récupère uniquement les champs CC
        if not key.startswith("cc_"):
            continue

        # =====================================================
        # RÉCUPÉRER L'ID DE L'ÉTUDIANT
        # =====================================================

        etudiant_id = key.replace("cc_", "")

        etudiant = get_object_or_404(
            EtudiantLMD,
            id=etudiant_id
        )

        # =====================================================
        # VÉRIFIER QUE L'ÉTUDIANT APPARTIENT AU BON NIVEAU
        # =====================================================

        if etudiant.niveau != niveau:

            continue

        # =====================================================
        # CC
        # =====================================================

        cc_value = value.strip()

        if cc_value == "":
            cc = 0
        else:
            try:
                cc = float(cc_value)
            except (ValueError, TypeError):
                cc = 0

        # =====================================================
        # EXAMEN
        # =====================================================

        examen_value = request.POST.get(
            f"examen_{etudiant_id}",
            ""
        ).strip()

        if examen_value == "":
            examen = 0
        else:
            try:
                examen = float(examen_value)
            except (ValueError, TypeError):
                examen = 0

        # =====================================================
        # VALIDATION DES NOTES
        # =====================================================

        if cc < 0 or cc > 20:

            messages.error(
                request,
                f"La note de CC de "
                f"{etudiant.nom} {etudiant.prenoms} "
                f"doit être comprise entre 0 et 20."
            )

            return redirect(
                f"{reverse('droit_prive_notes')}?"
                f"niveau={niveau}&semestre={semestre}"
                f"&session={session}"
            )

        if examen < 0 or examen > 20:

            messages.error(
                request,
                f"La note d'examen de "
                f"{etudiant.nom} {etudiant.prenoms} "
                f"doit être comprise entre 0 et 20."
            )

            return redirect(
                f"{reverse('droit_prive_notes')}?"
                f"niveau={niveau}&semestre={semestre}"
                f"&session={session}"
            )

        # =====================================================
        # ENREGISTREMENT DE LA NOTE
        # =====================================================

        NoteLMD.objects.update_or_create(

            etudiant=etudiant,

            ecue=ecue,

            semestre=semestre,

            session=session,

            defaults={
                "cc": cc,
                "examen": examen,
            }
        )

        nombre_notes += 1

    # =========================================================
    # MESSAGE DE SUCCÈS
    # =========================================================

    if session == "1":

        libelle_session = "Session 1"

    else:

        libelle_session = "Session 2"

    messages.success(
        request,
        f"{nombre_notes} note(s) du {semestre} "
        f"- {libelle_session} ont été enregistrée(s) "
        f"avec succès."
    )

    # =========================================================
    # RETOUR À LA PAGE AVEC LES FILTRES
    # =========================================================

    return redirect(
        f"{reverse('droit_prive_notes')}?"
        f"niveau={niveau}"
        f"&semestre={semestre}"
        f"&session={session}"
    )


 
@login_required(login_url="login")
@role_required("ADMIN")
def l3_droit_bulletins(request):

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Droit Privé"
    )

    etudiants = EtudiantLMD.objects.filter(
        filiere=filiere,
       
    ).order_by(
        "nom",
        "prenoms"
    )

    return render(
        request,
        "lmd/l3/droit/bulletins.html",
        {
            "etudiants": etudiants,
            "filiere": filiere
        }
    )




def l3_gestion_etudiants(request):

    return render(
        request,
        "lmd/l3/gestion/etudiants.html"
    )


from django.shortcuts import render, redirect, get_object_or_404
from .models import UE, FiliereLMD

def l3_gestion_ue(request):

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Sciences de Gestion"
    )

    # Par défaut : Semestre 1
    semestre = request.GET.get("semestre", "S1")

    ues = UE.objects.filter(
        filiere=filiere,
        semestre=semestre
    ).prefetch_related(
        "ecues"
    )

    return render(
        request,
        "lmd/l3/gestion/ue.html",
        {
            "filiere": filiere,
            "ues": ues,
            "semestre": semestre,
        }
    )


def l3_gestion_notesrrr(request, ecue_id=None):

    # =========================================================
    # 1. VÉRIFIER QU'UN ECUE A ÉTÉ SÉLECTIONNÉ
    # =========================================================

    if ecue_id is None:
        return redirect("l3_gestion_saisie_notes")

    # =========================================================
    # 2. RÉCUPÉRER L'ECUE
    # =========================================================

    ecue = get_object_or_404(
        ECUE.objects.select_related(
            "ue",
            "ue__filiere",
        ),
        id=ecue_id,
    )

    ue = ecue.ue
    filiere = ue.filiere
    niveau = ue.niveau

    # =========================================================
    # 3. VÉRIFICATION FILIÈRE
    # =========================================================

    if filiere.libelle != "Sciences de Gestion":
        messages.error(
            request,
            "Cette ECUE n'appartient pas à la filière Sciences de Gestion."
        )

        return redirect("l3_gestion_saisie_notes")

    # =========================================================
    # 4. RÉCUPÉRER LES ÉTUDIANTS
    # =========================================================

    etudiants = (
        EtudiantLMD.objects
        .filter(
            filiere=filiere,
            niveau=niveau,
        )
        .order_by(
            "nom",
            "prenoms",
        )
    )

    # =========================================================
    # 5. ENREGISTREMENT DES NOTES
    # =========================================================

    if request.method == "POST":

        nombre_notes = 0

        for etudiant in etudiants:

            cc_value = request.POST.get(
                f"cc_{etudiant.id}",
                ""
            ).strip()

            examen_value = request.POST.get(
                f"examen_{etudiant.id}",
                ""
            ).strip()

            # -------------------------------------------------
            # Aucun champ rempli
            # -------------------------------------------------

            if not cc_value and not examen_value:
                continue

            # -------------------------------------------------
            # Conversion des notes
            # -------------------------------------------------

            try:

                cc = float(cc_value) if cc_value else 0
                examen = float(examen_value) if examen_value else 0

            except (ValueError, TypeError):

                messages.error(
                    request,
                    f"Note invalide pour {etudiant.nom} "
                    f"{etudiant.prenoms}."
                )

                continue

            # -------------------------------------------------
            # Vérification des notes
            # -------------------------------------------------

            if cc < 0 or cc > 20:

                messages.error(
                    request,
                    f"Le CC de {etudiant.nom} {etudiant.prenoms} "
                    f"doit être compris entre 0 et 20."
                )

                continue

            if examen < 0 or examen > 20:

                messages.error(
                    request,
                    f"L'examen de {etudiant.nom} {etudiant.prenoms} "
                    f"doit être compris entre 0 et 20."
                )

                continue

            # -------------------------------------------------
            # CALCUL MOYENNE
            # -------------------------------------------------

            moyenne = round(
                (cc * 0.40) + (examen * 0.60),
                2,
            )

            # -------------------------------------------------
            # CRÉATION / MODIFICATION DE LA NOTE
            # -------------------------------------------------

            NoteLMD.objects.update_or_create(

                etudiant=etudiant,

                ecue=ecue,

                semestre="S1",

                session="1",

                defaults={
                    "cc": cc,
                    "examen": examen,
                    "moyenne": moyenne,
                },
            )

            nombre_notes += 1

        # =====================================================
        # MESSAGE DE CONFIRMATION
        # =====================================================

        messages.success(
            request,
            f"{nombre_notes} note(s) enregistrée(s) avec succès."
        )

        # =====================================================
        # RETOUR SUR LA MÊME ECUE
        # =====================================================

        return redirect(
            "l3_gestion_notes",
            ecue_id=ecue.id,
        )

    # =========================================================
    # 6. RÉCUPÉRER LES NOTES EXISTANTES
    # =========================================================

    notes = (
        NoteLMD.objects
        .filter(
            ecue=ecue,
            semestre="S1",
            session="1",
        )
    )

    # =========================================================
    # 7. TRANSFORMER LES NOTES EN DICTIONNAIRE
    # =========================================================

    notes_dict = {
        note.etudiant_id: note
        for note in notes
    }

    # =========================================================
    # 8. CONTEXTE
    # =========================================================

    context = {

        "ecue": ecue,

        "ue": ue,

        "filiere": filiere,

        "niveau": niveau,

        "etudiants": etudiants,

        "notes_dict": notes_dict,

    }

    # =========================================================
    # 9. AFFICHAGE
    # =========================================================

    return render(
        request,
        "lmd/l3/gestion/notes/saisie.html",
        context,
    )  
    
def l3_gestion_notes(request, ecue_id=None):

    # =========================================================
    # 1. VÉRIFIER QU'UN ECUE A ÉTÉ SÉLECTIONNÉ
    # =========================================================

    if ecue_id is None:
        return redirect("l3_gestion_saisie_notes")

    # =========================================================
    # 2. RÉCUPÉRER L'ECUE
    # =========================================================

    ecue = get_object_or_404(
        ECUE.objects.select_related(
            "ue",
            "ue__filiere",
        ),
        id=ecue_id,
    )

    ue = ecue.ue
    filiere = ue.filiere

    # =========================================================
    # 3. RÉCUPÉRER LE SEMESTRE DE L'UE
    # =========================================================

    # semestre = ue.semestre
    # semestre = etudiant.semestre
    for etudiant in etudiants:
         semestre_etudiant = etudiant.semestre

    # =========================================================
    # 4. RÉCUPÉRER LE NIVEAU
    # =========================================================

    niveau = ue.niveau

    # =========================================================
    # 5. VÉRIFICATION FILIÈRE
    # =========================================================

    if filiere.libelle != "Sciences de Gestion":

        messages.error(
            request,
            "Cette ECUE n'appartient pas à la filière Sciences de Gestion."
        )

        return redirect(
            "l3_gestion_saisie_notes"
        )

    # =========================================================
    # 6. RÉCUPÉRER LES ÉTUDIANTS
    # =========================================================

    etudiants = (
        EtudiantLMD.objects
        .filter(
            filiere=filiere,
            niveau=niveau,
        )
        .order_by(
            "nom",
            "prenoms",
        )
    )

    # =========================================================
    # 7. ENREGISTREMENT DES NOTES
    # =========================================================

    if request.method == "POST":

        nombre_notes = 0

        for etudiant in etudiants:

            # -------------------------------------------------
            # RÉCUPÉRER CC
            # -------------------------------------------------

            cc_value = request.POST.get(
                f"cc_{etudiant.id}",
                ""
            ).strip()

            # -------------------------------------------------
            # RÉCUPÉRER EXAMEN
            # -------------------------------------------------

            examen_value = request.POST.get(
                f"examen_{etudiant.id}",
                ""
            ).strip()

            # -------------------------------------------------
            # AUCUNE NOTE SAISIE
            # -------------------------------------------------

            if not cc_value and not examen_value:
                continue

            # -------------------------------------------------
            # CONVERSION
            # -------------------------------------------------

            try:

                cc = float(cc_value) if cc_value else 0
                examen = float(examen_value) if examen_value else 0

            except (ValueError, TypeError):

                messages.error(
                    request,
                    f"Note invalide pour "
                    f"{etudiant.nom} {etudiant.prenoms}."
                )

                continue

            # -------------------------------------------------
            # VÉRIFICATION CC
            # -------------------------------------------------

            if cc < 0 or cc > 20:

                messages.error(
                    request,
                    f"Le CC de {etudiant.nom} {etudiant.prenoms} "
                    f"doit être compris entre 0 et 20."
                )

                continue

            # -------------------------------------------------
            # VÉRIFICATION EXAMEN
            # -------------------------------------------------

            if examen < 0 or examen > 20:

                messages.error(
                    request,
                    f"L'examen de {etudiant.nom} {etudiant.prenoms} "
                    f"doit être compris entre 0 et 20."
                )

                continue

            # -------------------------------------------------
            # CRÉATION / MODIFICATION
            # -------------------------------------------------

            NoteLMD.objects.update_or_create(

                etudiant=etudiant,

                ecue=ecue,

                # IMPORTANT :
                # Le semestre vient automatiquement de l'UE
                # semestre=semestre,
                semestre=semestre_etudiant,
                

                # Session normale
                session="1",

                defaults={
                    "cc": cc,
                    "examen": examen,
                },
            )

            nombre_notes += 1

        # =====================================================
        # MESSAGE
        # =====================================================

        messages.success(
            request,
            f"{nombre_notes} note(s) du {semestre} "
            f"enregistrée(s) avec succès."
        )

        # =====================================================
        # RETOUR SUR L'ECUE
        # =====================================================

        return redirect(
            "l3_gestion_notes",
            ecue_id=ecue.id,
        )

    # =========================================================
    # 8. RÉCUPÉRER LES NOTES EXISTANTES
    # =========================================================

    notes = (
        NoteLMD.objects
        .filter(
            ecue=ecue,
            semestre=semestre,
            session="1",
        )
    )

    # =========================================================
    # 9. TRANSFORMER EN DICTIONNAIRE
    # =========================================================

    notes_dict = {
        note.etudiant_id: note
        for note in notes
    }

    # =========================================================
    # 10. CONTEXTE
    # =========================================================

    context = {

        "ecue": ecue,

        "ue": ue,

        "filiere": filiere,

        "niveau": niveau,

        "semestre": semestre,

        "etudiants": etudiants,

        "notes_dict": notes_dict,

    }

    # =========================================================
    # 11. AFFICHAGE
    # =========================================================

    return render(
        request,
        "lmd/l3/gestion/notes/saisie.html",
        context,
    )

@login_required(login_url="login")
@role_required("ADMIN")
def l3_gestion_bulletins(request):

    return render(
        request,
        "lmd/l3/gestion/bulletins.html"
    )


def l3_gestion_notes_selectioneeee(request):

    ecues = ECUE.objects.filter(
        ue__filiere__libelle__icontains="Sciences de Gestion"
    )

    return render(
        request,
        "lmd/l3/gestion/notes/ecue_list.html",
        {
            "ecues": ecues
        }
    )
  
  
def l3_gestion_notes_selection(request):

    # =========================================================
    # 1. RÉCUPÉRER LES FILTRES
    # =========================================================

    niveau_selectionne = request.GET.get("niveau", "").strip()
    ue_selectionnee = request.GET.get("ue", "").strip()
    ecue_selectionnee = request.GET.get("ecue", "").strip()

    # =========================================================
    # 2. FILTRER UNIQUEMENT SCIENCES DE GESTION
    # =========================================================

    base_ues = UE.objects.filter(
        filiere__libelle__icontains="Sciences de Gestion"
    )

    # =========================================================
    # 3. UE
    # =========================================================

    ues = UE.objects.none()

    if niveau_selectionne:

        ues = (
            base_ues
            .filter(
                niveau=niveau_selectionne
            )
            .order_by("ordre", "code")
        )

    # =========================================================
    # 4. ECUE
    # =========================================================

    ecues = ECUE.objects.none()

    if ue_selectionnee:

        ecues = (
            ECUE.objects
            .filter(
                ue_id=ue_selectionnee,
                ue__filiere__libelle__icontains="Sciences de Gestion",
            )
            .select_related("ue")
            .order_by("ordre", "code")
        )

    # =========================================================
    # 5. OBJET UE SÉLECTIONNÉ
    # =========================================================

    ue_selectionnee_obj = None

    if ue_selectionnee:

        ue_selectionnee_obj = (
            UE.objects
            .filter(
                id=ue_selectionnee,
                filiere__libelle__icontains="Sciences de Gestion",
            )
            .first()
        )

    # =========================================================
    # 6. OBJET ECUE SÉLECTIONNÉ
    # =========================================================

    ecue_selectionnee_obj = None

    if ecue_selectionnee:

        ecue_selectionnee_obj = (
            ECUE.objects
            .filter(
                id=ecue_selectionnee,
                ue__filiere__libelle__icontains="Sciences de Gestion",
            )
            .select_related("ue")
            .first()
        )

    # =========================================================
    # 7. CONTEXTE
    # =========================================================

    context = {

        "niveau_selectionne": niveau_selectionne,

        "ues": ues,

        "ue_selectionnee": ue_selectionnee,

        "ue_selectionnee_obj": ue_selectionnee_obj,

        "ecues": ecues,

        "ecue_selectionnee": ecue_selectionnee,

        "ecue_selectionnee_obj": ecue_selectionnee_obj,

    }

    # =========================================================
    # 8. AFFICHAGE
    # =========================================================

    return render(
        request,
        "lmd/l3/gestion/notes/ecue_list.html",
        context
    ) 
# ================= MASTER =================

def master_dashboard(request):

    return render(
        request,
        "lmd/master/dashboard.html"
    )

def master_filiere_list(request):

    return render(
        request,
        "lmd/master/filieres.html"
    )

    
def master_etudiant_list(request):

    etudiants = EtudiantMaster.objects.select_related(
        "programme"
    ).all()


    return render(
        request,
        "lmd/master/etudiants/list.html",
        {
            "etudiants":etudiants
        }
    )
    
def master_etudiant_add(request):

    if request.method == "POST":

        form = MasterEtudiantForm(request.POST)

        if form.is_valid():

            form.save()

            return redirect(
                "master_etudiant_list"
            )

    else:

        form = MasterEtudiantForm()


    return render(
        request,
        "lmd/master/etudiants/form.html",
        {
            "form": form,
            "titre": "Ajouter étudiant Master"
        }
    )


def master_etudiant_edit(request, id):

    etudiant = get_object_or_404(
        EtudiantMaster,
        id=id
    )


    if request.method == "POST":

        form = MasterEtudiantForm(
            request.POST,
            instance=etudiant
        )


        if form.is_valid():

            form.save()

            return redirect(
                "master_etudiant_list"
            )


    else:

        form = MasterEtudiantForm(
            instance=etudiant
        )


    return render(
        request,
        "lmd/master/etudiants/form.html",
        {
            "form": form,
            "titre": "Modifier étudiant Master"
        }
    )

def master_etudiant_delete(request, id):

    etudiant = get_object_or_404(
        EtudiantMaster,
        id=id
    )


    if request.method == "POST":

        etudiant.delete()

        return redirect(
            "master_etudiant_list"
        )


    return render(
        request,
        "lmd/master/etudiants/delete.html",
        {
            "etudiant": etudiant
        }
    )

from .forms import MasterProgrammeForm


def master_programme_add(request):

    if request.method == "POST":

        form = MasterProgrammeForm(request.POST)


        if form.is_valid():

            form.save()

            return redirect(
                "master_etudiant_list"
            )


    else:

        form = MasterProgrammeForm()


    return render(
        request,
        "lmd/master/programmes/form.html",
        {
            "form": form,
            "titre": "Ajouter Programme Master"
        }
    )


def master_programme_edit(request, id):

    programme = get_object_or_404(
        MasterProgramme,
        id=id
    )

    if request.method == "POST":

        form = MasterProgrammeForm(
            request.POST,
            instance=programme
        )

        if form.is_valid():

            form.save()

            return redirect(
                "master_programme_list"
            )

    else:

        form = MasterProgrammeForm(
            instance=programme
        )

    return render(
        request,
        "lmd/master/programme_form.html",
        {
            "form": form,
            "titre": "Modifier Programme Master"
        }
    )
    
def master_programme_delete(request, id):

    programme = get_object_or_404(
        MasterProgramme,
        id=id
    )

    programme.delete()

    return redirect(
        "master_programme_list"
    ) 
    
# def master_ue_list(request):

#     return render(
#         request,
#         "lmd/master/ue_list.html"
#     )

def master_ue_listPASS(request):
    ues = MasterUE.objects.select_related("programme").all()

    return render(
        request,
        "lmd/master/ue_list.html",
        {
            "ues": ues,
        },
    )

def master_ue_list(request):

    programme_id = request.GET.get("programme")


    programme = None


    if programme_id:

        programme = get_object_or_404(
            MasterProgramme,
            id=programme_id
        )


        ues = MasterUE.objects.filter(
            programme=programme
        )


    else:

        ues = MasterUE.objects.all()



    return render(
        request,
         "lmd/master/ue_list.html",
        {
            "ues":ues,
            "programme":programme
        }
    )

def master_saisie_notes_ecuePASS(request, id):

    ecue = MasterECUE.objects.select_related(
        "ue",
        "ue__programme"
    ).get(id=id)


    etudiants = EtudiantMaster.objects.filter(
        programme=ecue.ue.programme
    )



    # Notes existantes
    notes_existantes = NoteMaster.objects.filter(
        ecue=ecue
    )



    notes_par_etudiant = {
        note.etudiant_id: note
        for note in notes_existantes
    }



    # Association étudiant + note
    for etudiant in etudiants:

        etudiant.note_existante = notes_par_etudiant.get(
            etudiant.id
        )



    if request.method == "POST":


        for etudiant in etudiants:


            cc = request.POST.get(
                f"cc_{etudiant.id}"
            )


            examen = request.POST.get(
                f"examen_{etudiant.id}"
            )


            if cc is not None or examen is not None:


                NoteMaster.objects.update_or_create(

                    etudiant=etudiant,

                    ecue=ecue,

                    defaults={

                        "cc": cc or 0,

                        "examen": examen or 0

                    }

                )



        messages.success(
            request,
            "Notes enregistrées avec succès"
        )


        return redirect(
            "master_saisie_notes",
            id=id
        )



    return render(
        request,
        "lmd/master/saisie_notes_ecue.html",
        {
            "ecue": ecue,
            "etudiants": etudiants
        }
    )

def master_saisie_notes_ecue(request, id):

    ecue = MasterECUE.objects.select_related(
        "ue",
        "ue__programme"
    ).get(id=id)


    etudiants = EtudiantMaster.objects.filter(
        programme=ecue.ue.programme
    )



    notes_existantes = NoteMaster.objects.filter(
        ecue=ecue
    )


    notes_dict = {
        note.etudiant_id: note
        for note in notes_existantes
    }


    for etudiant in etudiants:

        etudiant.note_existante = notes_dict.get(
            etudiant.id
        )



    if request.method == "POST":


        erreurs = []


        for etudiant in etudiants:


            cc = request.POST.get(
                f"cc_{etudiant.id}"
            )


            examen = request.POST.get(
                f"examen_{etudiant.id}"
            )



            if cc == "" and examen == "":
                continue



            cc = float(cc or 0)

            examen = float(examen or 0)



            # Validation

            if cc < 0 or cc > 20:

                erreurs.append(
                    f"CC invalide pour {etudiant.nom}"
                )

                continue



            if examen < 0 or examen > 20:

                erreurs.append(
                    f"Examen invalide pour {etudiant.nom}"
                )

                continue




            NoteMaster.objects.update_or_create(

                etudiant=etudiant,

                ecue=ecue,

                defaults={

                    "cc":cc,

                    "examen":examen

                }

            )




        if erreurs:


            for erreur in erreurs:

                messages.error(
                    request,
                    erreur
                )


        else:


            messages.success(
                request,
                "Les notes ont été enregistrées avec succès."
            )



        return redirect(
            "master_saisie_notes",
            id=id
        )



    return render(
        request,
        "lmd/master/saisie_notes_ecue.html",
        {
            "ecue":ecue,
            "etudiants":etudiants
        }
    )

def master_saisie_notes(request):

    programmes = MasterProgramme.objects.prefetch_related(
        "ues__ecues"
    ).all()


    return render(
        request,
        "lmd/master/notes.html",
        {
            "programmes": programmes
        }
    )


from django.shortcuts import render
from .models import EtudiantLMD


from .models import EtudiantMaster


from django.shortcuts import render
from .models import EtudiantMaster

@login_required(login_url="login")
@role_required("ADMIN")
def master_bulletin_list(request):

    etudiants = EtudiantMaster.objects.select_related(
        "programme",
        "programme__filiere"
    ).filter(
        programme__specialite__in=[
            "DROIT",
            "GESTION",
            "QHSE"
        ]
    )


    # ==========================
    # FILTRE MATRICULE
    # ==========================
    matricule = request.GET.get("matricule")

    if matricule:
        etudiants = etudiants.filter(
            matricule__icontains=matricule
        )


    # ==========================
    # FILTRE NIVEAU
    # ==========================
    niveau = request.GET.get("niveau")

    if niveau:
        etudiants = etudiants.filter(
            niveau=niveau
        )


    # ==========================
    # FILTRE FILIERE
    # ==========================
    specialite = request.GET.get("specialite")

    if specialite:
        etudiants = etudiants.filter(
            programme__specialite=specialite
        )


    context = {

        "etudiants": etudiants,

        "matricule": matricule or "",

        "niveau": niveau or "",

        "specialite": specialite or "",

    }


    return render(
        request,
        "lmd/master/bulletins.html",
        context
    )
    
@login_required(login_url="login")
@role_required("ADMIN")
def master_bulletin_pdf(request, id, semestre):

    etudiant = get_object_or_404(
        EtudiantMaster,
        id=id
    )


    pdf_dir = os.path.join(
        settings.MEDIA_ROOT,
        "bulletins"
    )

    os.makedirs(
        pdf_dir,
        exist_ok=True
    )


    file_path = os.path.join(
        pdf_dir,
        f"master_{etudiant.matricule}_{semestre}.pdf"
    )


    generer_bulletin_masters_pdf(
        etudiant,
        semestre,
        file_path
    )


    return FileResponse(
        open(file_path,"rb"),
        content_type="application/pdf"
    )


def l3_droit_ecue_addAZ(request, pk):

    ue = get_object_or_404(
        UE,
        id=pk,
        filiere__libelle="Droit Privé"
    )


    if request.method == "POST":

        ECUE.objects.create(

            ue=ue,

            code=request.POST.get("code"),

            libelle=request.POST.get("libelle"),

            coefficient=request.POST.get("coefficient"),

            credit=request.POST.get("credit"),

        )


        return redirect(
            "l3_droit_ecue",
            ue.id
        )


    return render(
        request,
        "lmd/l3/droit/ecue_form.html",
        {
            "ue":ue,
            "titre":"Ajouter ECUE - Droit Privé"
        }
    )

def l3_droit_ecue_add(request, pk):

    ue = get_object_or_404(
        UE,
        id=pk,
        filiere__libelle="Droit Privé"
    )

    if request.method == "POST":

        form = ECUEForm(request.POST)

        if form.is_valid():

            ecue = form.save(commit=False)

            ecue.ue = ue

            ecue.save()

            return redirect(
                "l3_droit_ecue",
                ue.id
            )

    else:

        form = ECUEForm()

    return render(
        request,
        "lmd/l3/droit/ecue_form.html",
        {
            "ue": ue,
            "form": form,
            "titre": "Ajouter ECUE - Droit Privé"
        }
    )

def l3_droit_saisie_notes(request, ecue_id):
    ecue = get_object_or_404(
        ECUE,
        id=ecue_id
    )

    etudiants = EtudiantLMD.objects.filter(
        filiere=ecue.ue.filiere,
        niveau="L3"
    )

    if request.method == "POST":

        for etudiant in etudiants:

            cc = request.POST.get(f"cc_{etudiant.id}")
            examen = request.POST.get(f"examen_{etudiant.id}")

            if cc != "" or examen != "":

                NoteLMD.objects.update_or_create(
                    etudiant=etudiant,
                    ecue=ecue,
                    defaults={
                        "cc": float(cc) if cc else 0,
                        "examen": float(examen) if examen else 0
                    }
                )

        return redirect("l3_droit_ecue", pk=ecue.ue.id)

    notes = {
        note.etudiant_id: note
        for note in NoteLMD.objects.filter(ecue=ecue)
    }

    return render(
        request,
        "lmd/l3/droit/saisie_notes.html",
        {
            "ecue": ecue,
            "etudiants": etudiants,
            "notes": notes,
        }
    )
# ==============================
# Liste des ECUE d'une UE
# ==============================

def l3_droit_ecue(request, pk):

    ue = get_object_or_404(
        UE,
        pk=pk,
        filiere__libelle="Droit Privé"
    )


    ecues = (
        ECUE.objects
        .filter(
            ue=ue
        )
        .order_by("code")
    )


    return render(
        request,
        "lmd/l3/droit/ecue.html",
        {
            "ue": ue,
            "ecues": ecues,
        }
    )
# ==============================
# Ajouter ECUE
# ==============================


# ==============================
# Modifier ECUE
# ==============================

def l3_droit_ecue_update(request, pk):

    ecue = get_object_or_404(
        ECUE,
        pk=pk
    )


    if request.method == "POST":

        form = ECUEForm(
            request.POST,
            instance=ecue
        )

        if form.is_valid():

            form.save()

            return redirect(
                "l3_droit_ecue",
                pk=ecue.ue.pk
            )


    else:

        form = ECUEForm(
            instance=ecue
        )


    return render(
        request,
        "lmd/l3/droit/ecue_form.html",
        {
            "form": form,
            "titre": "Modifier ECUE",
            "ue": ecue.ue,
            "ecue": ecue
        }
    )

# ==============================
# Supprimer ECUE
# ==============================

def l3_droit_ecue_delete(request, pk):

    ecue = get_object_or_404(
        ECUE,
        id=pk
    )

    ue_id = ecue.ue.id

    ecue.delete()


    return redirect(
        "l3_droit_ecue",
        pk=ue_id
    )
  
def l3_sciences_gestion_etudiants(request):

    etudiants = EtudiantLMD.objects.filter(
        filiere__libelle__icontains="Sciences de Gestion"
    ).order_by(
        "niveau",
        "nom",
        "prenoms"
    )

    print("======================================")
    print("NOMBRE ETUDIANTS :", etudiants.count())

    for e in etudiants:
        print(
            e.id,
            e.matricule,
            e.nom,
            e.prenoms,
            e.niveau,
            e.filiere.libelle
        )

    print("======================================")

    return render(
        request,
        "lmd/l3/gestion/etudiants.html",
        {
            "etudiants": etudiants,
        }
    )

from django.http import HttpResponse

def imprimer_bulletin_l3_droit_prive(request, id, semestre):

    etudiant = get_object_or_404(
        EtudiantLMD,
        id=id
    )

    pdf_dir = os.path.join(
        settings.MEDIA_ROOT,
        "bulletins"
    )

    os.makedirs(
        pdf_dir,
        exist_ok=True
    )


    file_path = os.path.join(
        pdf_dir,
        f"bulletin_{etudiant.matricule}_{semestre}.pdf"
    )


    generer_bulletin_droit_prive_pdf(
        etudiant,
        semestre,
        file_path
    )


    return FileResponse(
        open(file_path, "rb"),
        content_type="application/pdf"
    )
    
def liste_bulletins_l3_droit_priveFFF(request):

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Droit Privé"
    )

    etudiants = EtudiantLMD.objects.filter(
        filiere=filiere,
        niveau="L3"
    ).order_by(
        "nom",
        "prenoms"
    )

    return render(
        request,
        "lmd/l3/droit/bulletins.html",
        {
            "etudiants": etudiants,
            "filiere": filiere
        }
    )

@login_required(login_url="login")
@role_required("ADMIN")
def liste_bulletins_l3_droit_prive(request):

    etudiants = EtudiantLMD.objects.filter(
        filiere__libelle="Droit Privé"
    ).order_by(
        "niveau",
        "nom",
        "prenoms"
    )


    print("Nombre étudiants :", etudiants.count())


    for e in etudiants:
        print(
            e.nom,
            e.prenoms,
            e.niveau,
            e.filiere.libelle
        )


    return render(
        request,
        "lmd/l3/droit/bulletins.html",
        {
            "etudiants": etudiants
        }
    )

def l3_gestion_etudiant_add(request):

    if request.method == "POST":

        form = EtudiantGestionForm(request.POST)

        print("=== POST RECU ===")
        print(request.POST)

        if form.is_valid():

            print("=== FORMULAIRE VALIDE ===")

            etudiant = form.save(commit=False)

            etudiant.niveau = "L3"

            etudiant.filiere = FiliereLMD.objects.get(
                libelle__icontains="Sciences de Gestion"
            )

            etudiant.save()

            print("=== ETUDIANT ENREGISTRE ===")

            return redirect(
                "l3_gestion_etudiant_list"
            )

        else:

            print("=== ERREURS FORMULAIRE ===")
            print(form.errors)

    else:

        form = EtudiantGestionForm()


    return render(
        request,
        "lmd/l3/gestion/etud_form.html",
        {
            "form": form
        }
    )

def l3_gestion_etudiant_edit(request, pk):

    etudiant = get_object_or_404(
        EtudiantLMD,
        pk=pk
    )


    form = EtudiantGestionForm(
        request.POST or None,
        instance=etudiant
    )


    if form.is_valid():

        form.save()

        return redirect(
            "l3_gestion_etudiant_list"
        )


    return render(
        request,
        "lmd/l3/gestion/etud_form.html",
        {
            "form":form
        }
    )

def l3_gestion_etudiant_delete(request, pk):

    etudiant = get_object_or_404(
        EtudiantLMD,
        pk=pk
    )


    if request.method == "POST":

        etudiant.delete()

        return redirect(
            "l3_gestion_etudiant_list"
        )


    return render(
        request,
        "lmd/l3/gestion/delete.html",
        {
            "etudiant":etudiant
        }
    )

def l3_gestion_ue_list(request):

    # Niveau sélectionné
    niveau = request.GET.get("niveau", "L3")

    # Semestre sélectionné
    semestre = request.GET.get("semestre", "S1")

    # Filière Sciences de Gestion
    filiere = FiliereLMD.objects.filter(
        libelle__icontains="Sciences de Gestion"
    ).first()

    if not filiere:
        messages.error(
            request,
            "La filière Sciences de Gestion est introuvable."
        )
        return redirect("home")

    # UE selon filière + niveau + semestre
    ues = UE.objects.filter(
        filiere=filiere,
        niveau=niveau,
        semestre=semestre
    ).prefetch_related(
        "ecues"
    ).order_by(
        "ordre",
        "code"
    )

    return render(
        request,
        "lmd/l3/gestion/ue/list.html",
        {
            "ues": ues,
            "niveau": niveau,
            "semestre": semestre,
            "filiere": filiere,
        }
    )
    
def l3_gestion_ue_addAAAA(request):

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Sciences de Gestion"
    )

    niveau = request.GET.get("niveau")
    semestre = request.GET.get("semestre")

    grandes_unites = GrandeUnite.objects.filter(
        filiere=filiere
    )

    if niveau:
        grandes_unites = grandes_unites.filter(
            niveau=niveau
        )

    if semestre:
        grandes_unites = grandes_unites.filter(
            semestre=semestre
        )

    grandes_unites = grandes_unites.order_by(
        "ordre"
    )

    if request.method == "POST":

        niveau = request.POST.get("niveau")
        semestre = request.POST.get("semestre")
        grande_unite_id = request.POST.get("grande_unite")

        grande_unite = get_object_or_404(
            GrandeUnite,
            id=grande_unite_id,
            filiere=filiere
        )

        UE.objects.create(
            code=request.POST.get("code"),
            libelle=request.POST.get("libelle"),
            credit=request.POST.get("credit"),
            ordre=request.POST.get("ordre") or 1,
            niveau=niveau,
            semestre=semestre,
            filiere=filiere,
            grande_unite=grande_unite,
        )

        messages.success(
            request,
            "UE ajoutée avec succès."
        )

        return redirect(
            f"{reverse('l3_gestion_ue_list')}"
            f"?niveau={niveau}&semestre={semestre}"
        )

    return render(
        request,
        "lmd/l3/gestion/ue/form.html",
        {
            "titre": "Nouvelle UE - Sciences de Gestion",
            "filiere": filiere,
            "niveau": niveau,
            "semestre": semestre,
            "grandes_unites": grandes_unites,
        }
    )

def l3_gestion_ue_add(request):

    # =========================================================
    # FILIÈRE
    # =========================================================

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Sciences de Gestion"
    )

    # =========================================================
    # RÉCUPÉRATION DES FILTRES
    # =========================================================

    niveau = request.GET.get("niveau", "")
    semestre = request.GET.get("semestre", "")
    session = request.GET.get("session", "1")

    # =========================================================
    # GRANDES UNITÉS
    # =========================================================

    grandes_unites = GrandeUnite.objects.filter(
        filiere=filiere
    )

    if niveau:
        grandes_unites = grandes_unites.filter(
            niveau=niveau
        )

    if semestre:
        grandes_unites = grandes_unites.filter(
            semestre=semestre
        )

    grandes_unites = grandes_unites.order_by(
        "ordre"
    )

    # =========================================================
    # TRAITEMENT DU FORMULAIRE
    # =========================================================

    if request.method == "POST":

        # -----------------------------------------------------
        # RÉCUPÉRATION DES DONNÉES
        # -----------------------------------------------------

        niveau = request.POST.get("niveau")
        semestre = request.POST.get("semestre")
        session = request.POST.get("session", "1")

        code = request.POST.get("code")
        libelle = request.POST.get("libelle")
        credit = request.POST.get("credit")
        ordre = request.POST.get("ordre") or 1

        grande_unite_id = request.POST.get(
            "grande_unite"
        )

        # -----------------------------------------------------
        # VÉRIFICATION GRANDE UNITÉ
        # -----------------------------------------------------

        grande_unite = get_object_or_404(
            GrandeUnite,
            id=grande_unite_id,
            filiere=filiere,
            niveau=niveau,
            semestre=semestre
        )

        # -----------------------------------------------------
        # CRÉATION DE L'UE
        # -----------------------------------------------------

        UE.objects.create(
            code=code,
            libelle=libelle,
            credit=credit,
            ordre=ordre,
            niveau=niveau,
            semestre=semestre,
            session=session,
            filiere=filiere,
            grande_unite=grande_unite,
        )

        # -----------------------------------------------------
        # MESSAGE
        # -----------------------------------------------------

        messages.success(
            request,
            "UE ajoutée avec succès."
        )

        # -----------------------------------------------------
        # REDIRECTION
        # -----------------------------------------------------

        return redirect(
            f"{reverse('l3_gestion_ue_list')}"
            f"?niveau={niveau}"
            f"&semestre={semestre}"
            f"&session={session}"
        )

    # =========================================================
    # AFFICHAGE DU FORMULAIRE
    # =========================================================

    return render(
        request,
        "lmd/l3/gestion/ue/form.html",
        {
            "titre": "Nouvelle UE - Sciences de Gestion",
            "filiere": filiere,
            "niveau": niveau,
            "semestre": semestre,
            "session": session,
            "grandes_unites": grandes_unites,
        }
    )

def l3_gestion_ecue_list(request, ue_id):

    ue = get_object_or_404(UE, pk=ue_id)

    ecues = ue.ecues.all().order_by("code")

    return render(
        request,
        "lmd/l3/gestion/ecue/list.html",
        {
            "ue": ue,
            "ecues": ecues
        }
    )

def l3_gestion_ecue_add(request, ue_id):

    ue = get_object_or_404(UE, pk=ue_id)

    if request.method == "POST":

        form = ECUEForm(request.POST)

        if form.is_valid():

            ecue = form.save(commit=False)

            ecue.ue = ue

            ecue.save()

            return redirect(
                "l3_gestion_ecue_list",
                ue.id
            )

    else:

        form = ECUEForm()

    return render(
        request,
        "lmd/l3/gestion/ecue/form.html",
        {
            "form": form,
            "ue": ue
        }
    )
    


def l3_gestion_saisie_notes(request, ecue_id):

    # =========================================================
    # 1. RÉCUPÉRER L'ECUE
    # =========================================================

    ecue = get_object_or_404(
        ECUE.objects.select_related(
            "ue",
            "ue__filiere",
        ),
        id=ecue_id,
    )

    ue = ecue.ue
    filiere = ue.filiere
    niveau = ue.niveau

    # =========================================================
    # 2. RÉCUPÉRER LE SEMESTRE
    # =========================================================
    #
    # POST : le semestre vient du formulaire
    # GET  : le semestre vient de l'URL
    #
    # Si aucun semestre n'est fourni, on utilise celui de l'UE
    # comme valeur par défaut.
    # =========================================================

    if request.method == "POST":
        semestre = request.POST.get("semestre")
    else:
        semestre = request.GET.get("semestre")

    # ---------------------------------------------------------
    # Valeur par défaut
    # ---------------------------------------------------------

    if not semestre:
        semestre = ue.semestre

    # ---------------------------------------------------------
    # Vérification
    # ---------------------------------------------------------

    if semestre not in ["S1", "S2"]:

        messages.error(
            request,
            "Veuillez sélectionner un semestre valide."
        )

        return redirect(
            "l3_gestion_saisie_notes",
            ecue_id=ecue.id,
        )

    # =========================================================
    # 3. RÉCUPÉRER LES ÉTUDIANTS
    # =========================================================

    etudiants = (
        EtudiantLMD.objects
        .filter(
            filiere=filiere,
            niveau=niveau,
        )
        .order_by(
            "nom",
            "prenoms",
        )
    )

    # =========================================================
    # 4. ENREGISTREMENT DES NOTES
    # =========================================================

    if request.method == "POST":

        nombre_notes = 0

        for etudiant in etudiants:

            # -------------------------------------------------
            # CC
            # -------------------------------------------------

            cc_value = request.POST.get(
                f"cc_{etudiant.id}",
                ""
            ).strip()

            # -------------------------------------------------
            # EXAMEN
            # -------------------------------------------------

            examen_value = request.POST.get(
                f"examen_{etudiant.id}",
                ""
            ).strip()

            # -------------------------------------------------
            # AUCUNE NOTE
            # -------------------------------------------------

            if not cc_value and not examen_value:
                continue

            # -------------------------------------------------
            # CONVERSION
            # -------------------------------------------------

            try:

                cc = float(cc_value) if cc_value else 0
                examen = float(examen_value) if examen_value else 0

            except (ValueError, TypeError):

                messages.error(
                    request,
                    f"Note invalide pour "
                    f"{etudiant.nom} {etudiant.prenoms}."
                )

                continue

            # -------------------------------------------------
            # VALIDATION CC
            # -------------------------------------------------

            if not 0 <= cc <= 20:

                messages.error(
                    request,
                    f"Le CC de {etudiant.nom} {etudiant.prenoms} "
                    f"doit être compris entre 0 et 20."
                )

                continue

            # -------------------------------------------------
            # VALIDATION EXAMEN
            # -------------------------------------------------

            if not 0 <= examen <= 20:

                messages.error(
                    request,
                    f"L'examen de {etudiant.nom} {etudiant.prenoms} "
                    f"doit être compris entre 0 et 20."
                )

                continue

            # -------------------------------------------------
            # CALCUL MOYENNE
            # CC      = 40 %
            # EXAMEN  = 60 %
            # -------------------------------------------------

            moyenne = round(
                (cc * 0.40) + (examen * 0.60),
                2
            )

            # =================================================
            # ENREGISTRER LA NOTE
            # =================================================

            NoteLMD.objects.update_or_create(

                etudiant=etudiant,

                ecue=ecue,

                semestre=semestre,

                session="1",

                defaults={
                    "cc": cc,
                    "examen": examen,
                    "moyenne": moyenne,
                }
            )

            nombre_notes += 1

        # =====================================================
        # MESSAGE
        # =====================================================

        messages.success(
            request,
            f"{nombre_notes} note(s) du {semestre} "
            f"enregistrée(s) avec succès."
        )

        # =====================================================
        # RETOUR SUR LA MÊME PAGE
        # =====================================================
        #
        # IMPORTANT :
        # On conserve ecue_id ET semestre.
        # =====================================================

        return redirect(
            f"{request.path}?semestre={semestre}"
        )

    # =========================================================
    # 5. RÉCUPÉRER LES NOTES DU SEMESTRE SÉLECTIONNÉ
    # =========================================================

    notes = (
        NoteLMD.objects
        .filter(
            ecue=ecue,
            semestre=semestre,
            session="1",
        )
    )

    # =========================================================
    # 6. DICTIONNAIRE DES NOTES
    # =========================================================

    notes_dict = {
        note.etudiant_id: note
        for note in notes
    }

    # =========================================================
    # 7. ÉTUDIANTS + NOTES
    # =========================================================

    etudiants_notes = []

    for etudiant in etudiants:

        etudiants_notes.append({
            "etudiant": etudiant,
            "note": notes_dict.get(etudiant.id),
        })

    # =========================================================
    # 8. CONTEXTE
    # =========================================================

    context = {

        "ecue": ecue,

        "ue": ue,

        "filiere": filiere,

        "niveau": niveau,

        # IMPORTANT :
        # C'est cette variable qu'il faut afficher
        # dans le template.
        "semestre": semestre,

        "etudiants_notes": etudiants_notes,

    }

    # =========================================================
    # 9. AFFICHAGE
    # =========================================================

    return render(
        request,
        "lmd/l3/gestion/notes/saisie.html",
        context,
    )

def l3_gestion_ue_editAAAA(request, pk):

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Sciences de Gestion"
    )

    ue = get_object_or_404(
        UE,
        pk=pk,
        filiere=filiere
    )

    niveau = request.GET.get("niveau") or ue.niveau
    semestre = request.GET.get("semestre") or ue.semestre

    grandes_unites = GrandeUnite.objects.filter(
        filiere=filiere,
        niveau=niveau,
        semestre=semestre
    ).order_by("ordre")

    if request.method == "POST":

        niveau = request.POST.get("niveau")
        semestre = request.POST.get("semestre")
        grande_unite_id = request.POST.get("grande_unite")

        grande_unite = get_object_or_404(
            GrandeUnite,
            id=grande_unite_id,
            filiere=filiere
        )

        ue.code = request.POST.get("code")
        ue.libelle = request.POST.get("libelle")
        ue.credit = request.POST.get("credit")
        ue.ordre = request.POST.get("ordre") or 1
        ue.niveau = niveau
        ue.semestre = semestre
        ue.grande_unite = grande_unite

        ue.save()

        messages.success(
            request,
            "UE modifiée avec succès."
        )

        return redirect(
            f"{reverse('l3_gestion_ue_list')}"
            f"?niveau={niveau}&semestre={semestre}"
        )

    return render(
        request,
         "lmd/l3/gestion/ue/form.html",
        {
            "titre": "Modifier une UE - Sciences de Gestion",
            "filiere": filiere,
            "ue": ue,
            "niveau": niveau,
            "semestre": semestre,
            "grandes_unites": grandes_unites,
        }
    )

def l3_gestion_ue_edit(request, pk):

    # =========================================================
    # FILIÈRE
    # =========================================================

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Sciences de Gestion"
    )

    # =========================================================
    # RÉCUPÉRER L'UE
    # =========================================================

    ue = get_object_or_404(
        UE,
        pk=pk,
        filiere=filiere
    )

    # =========================================================
    # VALEURS PAR DÉFAUT
    # =========================================================

    niveau = request.GET.get("niveau") or ue.niveau
    semestre = request.GET.get("semestre") or ue.semestre
    session = request.GET.get("session") or ue.session

    # =========================================================
    # GRANDES UNITÉS
    # =========================================================

    grandes_unites = GrandeUnite.objects.filter(
        filiere=filiere,
        niveau=niveau,
        semestre=semestre
    ).order_by("ordre")

    # =========================================================
    # TRAITEMENT DU FORMULAIRE
    # =========================================================

    if request.method == "POST":

        # -----------------------------------------------------
        # RÉCUPÉRATION DES DONNÉES
        # -----------------------------------------------------

        niveau = request.POST.get("niveau")
        semestre = request.POST.get("semestre")
        session = request.POST.get("session", "1")

        code = request.POST.get("code")
        libelle = request.POST.get("libelle")
        credit = request.POST.get("credit")
        ordre = request.POST.get("ordre") or 1

        grande_unite_id = request.POST.get(
            "grande_unite"
        )

        # -----------------------------------------------------
        # VÉRIFICATION GRANDE UNITÉ
        # -----------------------------------------------------

        grande_unite = get_object_or_404(
            GrandeUnite,
            id=grande_unite_id,
            filiere=filiere,
            niveau=niveau,
            semestre=semestre
        )

        # -----------------------------------------------------
        # MODIFICATION DE L'UE
        # -----------------------------------------------------

        ue.code = code
        ue.libelle = libelle
        ue.credit = credit
        ue.ordre = ordre
        ue.niveau = niveau
        ue.semestre = semestre
        ue.session = session
        ue.grande_unite = grande_unite

        ue.save()

        # -----------------------------------------------------
        # MESSAGE
        # -----------------------------------------------------

        messages.success(
            request,
            "UE modifiée avec succès."
        )

        # -----------------------------------------------------
        # REDIRECTION
        # -----------------------------------------------------

        return redirect(
            f"{reverse('l3_gestion_ue_list')}"
            f"?niveau={niveau}"
            f"&semestre={semestre}"
            f"&session={session}"
        )

    # =========================================================
    # AFFICHAGE DU FORMULAIRE
    # =========================================================

    return render(
        request,
        "lmd/l3/gestion/ue/form.html",
        {
            "titre": "Modifier une UE - Sciences de Gestion",
            "filiere": filiere,
            "ue": ue,
            "niveau": niveau,
            "semestre": semestre,
            "session": session,
            "grandes_unites": grandes_unites,
        }
    )

def l3_gestion_ue_delete(request, pk):

    ue = get_object_or_404(
        UE,
        id=pk
    )

    if request.method == "POST":

        ue.delete()

        return redirect(
            "l3_gestion_ue"
        )


    return render(
        request,
        "lmd/l3/gestion/ue/delete.html",
        {
            "ue": ue
        }
    )




def l3_gestion_ecue_edit(request, pk):

    ecue = get_object_or_404(
        ECUE,
        id=pk
    )


    if request.method == "POST":

        ecue.code = request.POST.get("code")
        ecue.libelle = request.POST.get("libelle")
        ecue.coefficient = request.POST.get("coefficient")
        ecue.credit = request.POST.get("credit")

        ecue.save()

        return redirect(
            "l3_gestion_ecue_list",
            ecue.ue.id
        )


    return render(
        request,
        "lmd/l3/gestion/ecue/edit.html",
        {
            "ecue": ecue
        }
    )

def l3_gestion_ecue_delete(request, pk):

    ecue = get_object_or_404(
        ECUE,
        id=pk
    )

    ue_id = ecue.ue.id


    if request.method == "POST":

        ecue.delete()

        return redirect(
            "l3_gestion_ecue_list",
            ue_id
        )


    return render(
        request,
        "lmd/l3/gestion/ecue/delete.html",
        {
            "ecue": ecue
        }
    )
@login_required(login_url="login")
@role_required("ADMIN")
def liste_bulletins_gestion(request):

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Sciences de Gestion"
    )

    etudiants = EtudiantLMD.objects.filter(
        filiere=filiere
    ).order_by("nom", "prenoms")

    print("========== TEST ==========")
    print("Filière :", filiere.libelle)
    print("Nombre d'étudiants :", etudiants.count())

    for e in etudiants:
        print(e.matricule, e.nom, e.prenoms)

    return render(
        request,
        "lmd/l3/gestion/bulletins.html",
        {
            "etudiants": etudiants,
            "filiere": filiere
        }
    )
@login_required(login_url="login")
@role_required("ADMIN")
def bulletin_gestion_lmd_pdf(request, id, semestre):

    etudiant = get_object_or_404(
        EtudiantLMD,
        id=id
    )


    pdf_dir = os.path.join(
        settings.MEDIA_ROOT,
        "bulletins"
    )

    os.makedirs(
        pdf_dir,
        exist_ok=True
    )


    file_path = os.path.join(
        pdf_dir,
        f"gestion_{etudiant.matricule}_{semestre}.pdf"
    )


    generer_bulletin_gestion_pdf(
        etudiant,
        semestre,
        file_path
    )


    return FileResponse(
        open(file_path, "rb"),
        content_type="application/pdf"
    )
@login_required(login_url="login")
@role_required("ADMIN")
def liste_bulletins_tronc_commun(request):

    etudiants = EtudiantLMD.objects.filter(
        filiere__libelle="Gestion et Droit",
        niveau__in=["L1", "L2"]
    ).order_by(
        "niveau",
        "nom",
        "prenoms"
    )

    return render(
        request,
        "lmd/trom_commun/bulletins.html",
        {
            "etudiants": etudiants,
            "titre": "Bulletins Tronc Commun L1 - L2"
        }
    )

@login_required(login_url="login")
@role_required("ADMIN")
def imprimer_bulletin_tronc_commun(request, pk):

    etudiant = get_object_or_404(
        EtudiantLMD,
        pk=pk
    )


    pdf_dir = os.path.join(
        settings.MEDIA_ROOT,
        "bulletins"
    )


    os.makedirs(
        pdf_dir,
        exist_ok=True
    )


    file_path = os.path.join(
        pdf_dir,
        f"tronc_commun_{etudiant.matricule}.pdf"
    )


    generer_bulletin_tronc_commun_pdf(
        etudiant,
        semestre,
        file_path
    )


    return FileResponse(
        open(file_path,"rb"),
        content_type="application/pdf"
    )


def ajouter_etudiant_tronc_commun(request):

    if request.method == "POST":

        matricule = request.POST.get("matricule")
        nom = request.POST.get("nom")
        prenoms = request.POST.get("prenoms")
        niveau_id = request.POST.get("niveau")


        filiere = FiliereLMD.objects.get(
            code="TC-DG"
        )


        EtudiantLMD.objects.create(
            matricule=matricule,
            nom=nom,
            prenoms=prenoms,
            filiere=filiere,
            niveau_id=niveau_id
        )


        return redirect(
            "liste_etudiants_tronc_commun"
        )


    niveaux = NiveauLMD.objects.filter(
        code__in=["L1","L2"]
    )


    return render(
        request,
        "lmd/add_etudiant_tc.html",
        {
            "niveaux": niveaux
        }
    )

def tronc_commun_etudiants(request):

    etudiants = EtudiantLMD.objects.filter(
        niveau__in=["L1", "L2"],
        filiere__libelle__in=[
            "Droit",
            "Gestion"
        ]
    ).select_related(
        "filiere"
    ).order_by(
        "filiere__libelle",
        "niveau",
        "nom",
        "prenoms"
    )


    filiere = request.GET.get("filiere")
    niveau = request.GET.get("niveau")


    if filiere:
        etudiants = etudiants.filter(
            filiere_id=filiere
        )


    if niveau:
        etudiants = etudiants.filter(
            niveau=niveau
        )


    context = {

        "etudiants": etudiants,


        "filieres": [
            {
                "id": f.id,
                "libelle": f.libelle
            }
            for f in FiliereLMD.objects.filter(
                libelle__in=[
                    "Droit",
                    "Gestion"
                ]
            )
        ],


        "niveaux":[
            "L1",
            "L2"
        ]

    }


    return render(
        request,
        "lmd/tronc_commun_etudiants.html",
        context
    )

def tronc_commun_ue(request):

    # ============================
    # SEMESTRE SELECTIONNE
    # ============================

    semestre = request.GET.get("semestre", "S1")

    # Sécurité
    if semestre not in ["S1", "S2"]:
        semestre = "S1"


    # ============================
    # RECUPERATION DES UE
    # ============================

    ues = UE.objects.filter(
        filiere__libelle="Gestion et Droit",
        semestre=semestre
    ).prefetch_related(
        "ecues"
    ).order_by(
        "code"
    )


    # ============================
    # AFFICHAGE
    # ============================

    return render(
        request,
        "lmd/trom_commun/ue.html",
        {
            "ues": ues,
            "semestre": semestre,
            "titre": "UE / ECUE Tronc Commun L1-L2 Droit & Gestion"
        }
    )
    
def tronc_commun_notes(request):

    # ==============================
    # SEMESTRE SELECTIONNE
    # ==============================

    semestre = request.GET.get(
        "semestre",
        "S1"
    )

    if semestre not in ["S1", "S2"]:
        semestre = "S1"


    # ==============================
    # FILIERE TRONC COMMUN
    # ==============================

    filiere = get_object_or_404(
        FiliereLMD,
        Q(libelle="Gestion et Droit") |
        Q(libelle="Droit et Gestion")
    )


    # ==============================
    # ENREGISTREMENT DES NOTES
    # ==============================

    if request.method == "POST":

        etudiants_post = EtudiantLMD.objects.filter(
            filiere=filiere,
            niveau__in=["L1", "L2"]
        )

        ecues_post = ECUE.objects.filter(
            ue__filiere=filiere,
            ue__semestre=semestre
        )


        for etudiant in etudiants_post:

            for ecue in ecues_post:


                cc = request.POST.get(
                    f"cc_{etudiant.id}_{ecue.id}"
                )

                examen = request.POST.get(
                    f"examen_{etudiant.id}_{ecue.id}"
                )


                # éviter de créer des notes vides

                if cc == "" and examen == "":
                    continue


                cc = float(cc) if cc else 0
                examen = float(examen) if examen else 0


                moyenne = (
                    (cc * 0.4) +
                    (examen * 0.6)
                )


                NoteLMD.objects.update_or_create(

                    etudiant=etudiant,

                    ecue=ecue,

                    semestre=semestre,

                    session="1",

                    defaults={

                        "cc": cc,

                        "examen": examen,

                        "moyenne": round(
                            moyenne,
                            2
                        )

                    }
                )


        messages.success(
            request,
            "Les notes du tronc commun ont été enregistrées avec succès."
        )


        return redirect(
            f"{request.path}?semestre={semestre}"
        )


    # ==============================
    # ETUDIANTS L1 / L2
    # ==============================

    etudiants = EtudiantLMD.objects.filter(
        filiere=filiere,
        niveau__in=["L1", "L2"]
    ).order_by(
        "niveau",
        "nom",
        "prenoms"
    )


    # ==============================
    # ECUE DU SEMESTRE
    # ==============================

    ecues = ECUE.objects.filter(
        ue__filiere=filiere,
        ue__semestre=semestre
    ).select_related(
        "ue"
    ).order_by(
        "code"
    )


    # ==============================
    # NOTES EXISTANTES
    # ==============================

    notes = NoteLMD.objects.filter(
        etudiant__in=etudiants,
        ecue__in=ecues,
        semestre=semestre,
        session="1"
    )


    notes_existantes = {}

    for note in notes:

        notes_existantes[
            (
                note.etudiant_id,
                note.ecue_id
            )
        ] = note



    # ==============================
    # PREPARATION AFFICHAGE
    # ==============================

    for etudiant in etudiants:

        etudiant.notes_affichage = []


        for ecue in ecues:

            note = notes_existantes.get(
                (
                    etudiant.id,
                    ecue.id
                )
            )


            etudiant.notes_affichage.append({

                "ecue_id": ecue.id,

                "cc": note.cc if note else "",

                "examen": note.examen if note else "",

                "moyenne": note.moyenne if note else "",

            })


    return render(
        request,
        "lmd/trom_commun/notes.html",
        {
            "etudiants": etudiants,
            "ecues": ecues,
            "semestre": semestre,
            "filiere": filiere,
        }
    )
# =====================================================
# TRONC COMMUN L1/L2 - DROIT + GESTION
# =====================================================

def liste_etudiants_tronc_commun(request):

    niveau_filtre = request.GET.get(
        "niveau",
        ""
    )


    etudiants = EtudiantLMD.objects.filter(
        Q(filiere__libelle__icontains="Droit")
        |
        Q(filiere__libelle__icontains="Gestion"),
        niveau__in=[
            "L1",
            "L2",
            "Licence 1",
            "Licence 2"
        ]
    )


    # Filtre niveau
    if niveau_filtre:

        etudiants = etudiants.filter(
            niveau=niveau_filtre
        )


    etudiants = etudiants.order_by(
        "niveau",
        "nom"
    )

    return render(
        request,
        "lmd/trom_commun/etudiants.html",
        {
            "etudiants": etudiants,

            "titre": "Tronc Commun Droit & Gestion L1-L2",

            "niveau_filtre": niveau_filtre,
        }
    )
# Vue spéciale Droit
def tronc_commun_droit(request):

    etudiants = EtudiantLMD.objects.filter(
        niveau__in=["L1","L2"],
        filiere__libelle__icontains="Droit"
    )

    return render(
        request,
        "lmd/trom_commun/etudiants.html",
        {
            "etudiants":etudiants,
            "titre":"Tronc Commun Droit L1-L2"
        }
    )



# Vue spéciale Gestion
def tronc_commun_gestion(request):

    etudiants = EtudiantLMD.objects.filter(
        niveau__in=["L1","L2"],
        filiere__libelle__icontains="Gestion"
    )


    return render(
        request,
        "lmd/trom_commun/etudiants.html",
        {
            "etudiants":etudiants,
            "titre":"Tronc Commun Gestion L1-L2"
        }
    )



# =====================================================
# UE TRONC COMMUN
# =====================================================

# =====================================================
# LISTE BULLETINS
# =====================================================
@login_required(login_url="login")
@role_required("ADMIN")
def bulletin_tronc_commun_list(request):

    etudiants = EtudiantLMD.objects.filter(
        niveau__in=["L1","L2"]
    )


    return render(
        request,
        "lmd/trom_commun/bulletins.html",
        {
            "etudiants":etudiants
        }
    ) 
       
@login_required(login_url="login")
@role_required("ADMIN")
def bulletin_tronc_commun_pdf(request, id, semestre):

    etudiant = get_object_or_404(
        EtudiantLMD,
        id=id
    )

    pdf_dir = os.path.join(
        settings.MEDIA_ROOT,
        "bulletins"
    )

    os.makedirs(pdf_dir, exist_ok=True)

    file_path = os.path.join(
        pdf_dir,
        f"tronc_commun_{etudiant.matricule}_{semestre}.pdf"
    )

    generer_bulletin_tronc_commun_pdf(
        etudiant,
        semestre,
        file_path
        
        
    )
  

    return FileResponse(

        open(file_path, "rb"),
        content_type="application/pdf"
    )
 
from .forms import TroncCommunEtudiantForm



def tronc_commun_add(request):

    if request.method == "POST":

        form = TroncCommunEtudiantForm(request.POST)


        if form.is_valid():

            etudiant = form.save()

            messages.success(
                request,
                "Étudiant ajouté avec succès"
            )

            return redirect(
                "liste_bulletins_tronc_commun"
            )


    else:

        form = TroncCommunEtudiantForm()



    return render(
        request,
        "lmd/trom_commun/form.html",
        {
            "form":form,
            "titre":"Ajouter étudiant tronc commun"
        }
    )




def tronc_commun_update(request, pk):

    etudiant = get_object_or_404(
        EtudiantLMD,
        pk=pk
    )

    if request.method == "POST":

        form = TroncCommunEtudiantForm(
            request.POST,
            request.FILES,
            instance=etudiant
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Modification de l'étudiant effectuée avec succès."
            )

            return redirect(
                "liste_etudiants_tronc_commun"
            )

    else:

        form = TroncCommunEtudiantForm(
            instance=etudiant
        )


    context = {
        "form": form,
        "titre": "Modifier étudiant"
    }


    return render(
        request,
        "lmd/trom_commun/form.html",
        context
    )





# =====================================================
# DASHBOARD QHSE L3
# =====================================================

def l3_qhse_dashboard(request):

    filiere = FiliereLMD.objects.filter(
        code="QHSE"
    ).first()


    etudiants = EtudiantLMD.objects.filter(
        filiere=filiere,
        niveau="L3"
    ).count()


    ues = UE.objects.filter(
        filiere=filiere
    ).count()



    context = {

        "filiere": filiere,

        "nb_etudiants": etudiants,

        "nb_ues": ues,

    }


    return render(
        request,
        "lmd/l3_qhse/dashboard.html",
        context
    )



# =====================================================
# ETUDIANTS QHSE
# =====================================================

def l3_qhse_etudiants1(request):

    filiere = FiliereLMD.objects.get(code="QHSE-L3")

    etudiants = EtudiantLMD.objects.filter(
        filiere=filiere,
        niveau="L3"
    )

    print("Nombre =", etudiants.count())

    return render(
        request,
        "lmd/l3_qhse/etudiants.html",
        {
            "etudiants": etudiants,
            "filiere": filiere,
        }
    )

def l3_qhse_etudiantsrrr(request):

    filiere = FiliereLMD.objects.get(code="QHSE-L3")

    etudiants = EtudiantLMD.objects.filter(
         filiere=filiere,
          niveau="L3"
    )

    print("Nombre =", etudiants.count())

    for e in etudiants:
        print(e.id, e.matricule, e.nom, e.prenoms)

    return render(
        request,
        "lmd/l3_qhse/etudiants.html",
        {
            "etudiants": etudiants,
            "filiere": filiere,
        }
    )

def l3_qhse_etudiantseeeeeeee(request):

    filiere = get_filiere_qhse()

    # "niveau" vient maintenant d'un paramètre GET (ex: ?niveau=L2)
    # plutôt que d'être codé en dur sur "L3" — vide = tous les niveaux.
    niveau = request.GET.get("niveau", "")

    etudiants = EtudiantLMD.objects.filter(filiere=filiere)

    if niveau:
        etudiants = etudiants.filter(niveau=niveau)

    etudiants = etudiants.order_by("niveau", "nom", "prenoms")

    return render(
        request,
        "lmd/l3_qhse/etudiants.html",
        {
            "etudiants": etudiants,
            "filiere": filiere,
            "niveau": niveau,
        }
    )

def l3_qhse_etudiantsAAA(request):

    filiere = get_filiere_qhse()

    niveau = request.GET.get("niveau", "").strip()

    etudiants = EtudiantLMD.objects.filter(
        filiere=filiere
    )

    if niveau in ["L1", "L2", "L3"]:
        etudiants = etudiants.filter(
            niveau=niveau
        )

    etudiants = etudiants.order_by(
        "niveau",
        "nom",
        "prenoms"
    )

    return render(
        request,
        "lmd/l3_qhse/etudiants.html",
        {
            "etudiants": etudiants,
            "filiere": filiere,
            "niveau": niveau,
        }
    )

def l3_qhse_etudiant_add(request):

    filiere = FiliereLMD.objects.get(
        code="QHSE-L3"
    )

    if request.method == "POST":

        form = QHSEEtudiantForm(request.POST)

        if form.is_valid():

            etudiant = form.save(commit=False)

            etudiant.filiere = filiere
            etudiant.niveau = "L3"

            etudiant.save()

            return redirect(
                "l3_qhse_etudiants"
            )

    else:

        form = QHSEEtudiantForm()


    return render(
        request,
        "lmd/l3_qhse/etudiant_form.html",
        {
            "form": form,
            "filiere": filiere
        }
    )

def l3_qhse_etudiants(request):

    filiere = get_filiere_qhse()

    niveau = request.GET.get("niveau", "").strip()

    etudiants = EtudiantLMD.objects.filter(
        filiere=filiere
    )

    if niveau in ["L1", "L2", "L3"]:
        etudiants = etudiants.filter(
            niveau=niveau
        )

    etudiants = etudiants.order_by(
        "niveau",
        "nom",
        "prenoms"
    )

    # =========================
    # STATISTIQUES
    # =========================

    total_etudiants = etudiants.count()

    total_l1 = etudiants.filter(
        niveau="L1"
    ).count()

    total_l2 = etudiants.filter(
        niveau="L2"
    ).count()

    total_l3 = etudiants.filter(
        niveau="L3"
    ).count()

    # =========================
    # PAGINATION
    # =========================

    paginator = Paginator(
        etudiants,
        10
    )

    page_number = request.GET.get("page")

    page_obj = paginator.get_page(
        page_number
    )

    return render(
        request,
        "lmd/l3_qhse/etudiants.html",
        {
            "etudiants": page_obj,
            "page_obj": page_obj,
            "paginator": paginator,

            "filiere": filiere,
            "niveau": niveau,

            "total_etudiants": total_etudiants,
            "total_l1": total_l1,
            "total_l2": total_l2,
            "total_l3": total_l3,
        }
    )
    
def l3_qhse_etudiant_update(request,pk):

    etudiant=get_object_or_404(
        EtudiantLMD,
        pk=pk
    )


    if request.method=="POST":

        form=QHSEEtudiantForm(
            request.POST,
            instance=etudiant
        )


        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Modification effectuée"
            )

            return redirect(
                "l3_qhse_etudiants"
            )


    else:

        form=QHSEEtudiantForm(
            instance=etudiant
        )


    return render(
        request,
        "lmd/l3_qhse/etudiant_form.html",
        {
            "form":form,
            "titre":"Modifier étudiant QHSE"
        }
    )

def l3_qhse_etudiant_delete(request,pk):

    etudiant=get_object_or_404(
        EtudiantLMD,
        pk=pk
    )


    if request.method=="POST":

        etudiant.delete()

        messages.success(
            request,
            "Étudiant supprimé"
        )


        return redirect(
            "l3_qhse_etudiants"
        )


    return render(
        request,
        "lmd/l3_qhse/etudiant_delete.html",
        {
            "etudiant":etudiant
        }
    )
# =====================================================
# UE / ECUE QHSE
# =====================================================

def l3_qhse_ueAAA(request):
    semestre = request.GET.get("semestre", "S1")

    ues = (
        UE.objects.filter(
            filiere__code="QHSE-L3",
            semestre=semestre
        )
        .prefetch_related("ecues")
        .order_by("code")
    )

    context = {
        "ues": ues,
        "semestre": semestre,
    }

    return render(
        request,
        "lmd/l3_qhse/ue.html",
        context,
    )

def l3_qhse_ue(request):

    # =========================================================
    # 1. RÉCUPÉRER LES FILTRES
    # =========================================================

    niveau = request.GET.get("niveau", "L1")
    semestre = request.GET.get("semestre", "S1")

    # =========================================================
    # 2. SÉCURISER LE NIVEAU
    # =========================================================

    niveaux_valides = [
        "L1",
        "L2",
        "L3",
    ]

    if niveau not in niveaux_valides:
        niveau = "L3"

    # =========================================================
    # 3. SÉCURISER LE SEMESTRE
    # =========================================================

    semestres_valides = [
        "S1",
        "S2",
        "S3",
        "S4",
        "S5",
        "S6",
    ]

    if semestre not in semestres_valides:
        semestre = "S1"

    # =========================================================
    # 4. RÉCUPÉRER LES UE QHSE
    # =========================================================

    ues = (
        UE.objects
        .filter(
            filiere__code="QHSE-L3",
            niveau=niveau,
            semestre=semestre
        )
        .prefetch_related("ecues")
        .order_by("ordre", "code")
    )

    # =========================================================
    # 5. CONTEXTE
    # =========================================================

    context = {
        "ues": ues,
        "niveau": niveau,
        "semestre": semestre,
    }

    # =========================================================
    # 6. AFFICHAGE
    # =========================================================

    return render(
        request,
        "lmd/l3_qhse/ue.html",
        context,
    )
# =====================================================
# SAISIE NOTES QHSE
# =====================================================

def l3_qhse_notes(request):

    filiere = FiliereLMD.objects.filter(
        code="QHSE"
    ).first()


    notes = NoteLMD.objects.filter(
        etudiant__filiere=filiere
    ).select_related(
        "etudiant",
        "ecue"
    )


    return render(
        request,
        "lmd/l3_qhse/notes.html",
        {
            "notes": notes,
            "filiere": filiere
        }
    )



# =====================================================
# BULLETINS QHSE
# =====================================================
@login_required(login_url="login")
@role_required("ADMIN")
def l3_qhse_bulletins(request):

    filiere = get_filiere_qhse()

    # =========================
    # RECHERCHE
    # =========================

    recherche = request.GET.get("q", "").strip()

    niveau = request.GET.get("niveau", "").strip()

    etudiants = EtudiantLMD.objects.filter(
        filiere=filiere
    ).select_related(
        "filiere"
    )

    # =========================
    # RECHERCHE MATRICULE / NOM / PRÉNOMS
    # =========================

    if recherche:

        etudiants = etudiants.filter(
            Q(matricule__icontains=recherche) |
            Q(nom__icontains=recherche) |
            Q(prenoms__icontains=recherche)
        )

    # =========================
    # FILTRE NIVEAU
    # =========================

    if niveau in ["L1", "L2", "L3"]:

        etudiants = etudiants.filter(
            niveau=niveau
        )

    # =========================
    # TRI
    # =========================

    etudiants = etudiants.order_by(
        "niveau",
        "nom",
        "prenoms"
    )

    # =========================
    # PAGINATION
    # =========================

    paginator = Paginator(
        etudiants,
        10
    )

    page_number = request.GET.get("page")

    page_obj = paginator.get_page(
        page_number
    )

    # =========================
    # STATISTIQUES
    # =========================

    total_etudiants = etudiants.count()

    return render(
        request,
        "lmd/l3_qhse/bulletins.html",
        {
            "etudiants": page_obj,
            "page_obj": page_obj,
            "paginator": paginator,

            "filiere": filiere,

            "recherche": recherche,
            "niveau": niveau,

            "total_etudiants": total_etudiants,
        }
    )

def l3_qhse_ecue_add(request, ue_id):

    ue = get_object_or_404(
        UE,
        id=ue_id
    )


    if request.method == "POST":

        form = QHSEECUEForm(request.POST)


        if form.is_valid():

            ecue = form.save(commit=False)

            # rattachement automatique
            ecue.ue = ue

            ecue.save()


            messages.success(
                request,
                "ECUE ajouté avec succès"
            )


            return redirect(
                "l3_qhse_ue"
            )


    else:

        form = QHSEECUEForm()



    return render(
        request,
        "lmd/l3_qhse/ecue_form.html",
        {
            "form":form,
            "ue":ue,
            "titre":"Ajouter un ECUE"
        }
    )


def l3_qhse_ecue_update(request, pk):

    ecue = get_object_or_404(ECUE, pk=pk)


    if request.method == "POST":

        ecue.code = request.POST.get("code")
        ecue.libelle = request.POST.get("libelle")
        ecue.coefficient = request.POST.get("coefficient")
        ecue.credit = request.POST.get("credit")

        ecue.save()

        return redirect("l3_qhse_ue")


    return render(
        request,
        "lmd/l3_qhse/ecue_update.html",
        {
            "ecue": ecue
        }
    )



def l3_qhse_ecue_delete(request, pk):

    ecue = get_object_or_404(
        ECUE,
        pk=pk
    )


    if request.method == "POST":

        ue = ecue.ue

        ecue.delete()

        return redirect(
            "l3_qhse_ue"
        )


    return redirect(
        "l3_qhse_ue"
    )


def l3_qhse_ue_add(request):

    # =========================================================
    # FILIÈRE
    # =========================================================

    filiere = get_filiere_qhse()

    # =========================================================
    # RÉCUPÉRATION DES FILTRES
    # =========================================================

    niveau = request.GET.get("niveau", "")
    semestre = request.GET.get("semestre", "S1")
    session = request.GET.get("session", "1")

    # =========================================================
    # GRANDES UNITÉS
    # =========================================================

    grandes_unites = GrandeUnite.objects.filter(
        filiere=filiere
    )

    if niveau:
        grandes_unites = grandes_unites.filter(
            niveau=niveau
        )

    if semestre:
        grandes_unites = grandes_unites.filter(
            semestre=semestre
        )

    grandes_unites = grandes_unites.order_by(
        "ordre"
    )

    # =========================================================
    # TRAITEMENT DU FORMULAIRE
    # =========================================================

    if request.method == "POST":

        # -----------------------------------------------------
        # RÉCUPÉRATION DES DONNÉES
        # -----------------------------------------------------

        niveau = request.POST.get("niveau")
        semestre = request.POST.get("semestre")
        session = request.POST.get("session", "1")

        code = request.POST.get("code")
        libelle = request.POST.get("libelle")
        credit = request.POST.get("credit")
        ordre = request.POST.get("ordre") or 1

        grande_unite_id = request.POST.get(
            "grande_unite"
        )

        # -----------------------------------------------------
        # VÉRIFICATION GRANDE UNITÉ
        # -----------------------------------------------------

        grande_unite = get_object_or_404(
            GrandeUnite,
            id=grande_unite_id,
            filiere=filiere,
            niveau=niveau,
            semestre=semestre
        )

        # -----------------------------------------------------
        # CRÉATION DE L'UE
        # -----------------------------------------------------

        UE.objects.create(
            code=code,
            libelle=libelle,
            credit=credit,
            ordre=ordre,
            niveau=niveau,
            semestre=semestre,
            session=session,
            filiere=filiere,
            grande_unite=grande_unite,
        )

        # -----------------------------------------------------
        # MESSAGE
        # -----------------------------------------------------

        messages.success(
            request,
            "UE QHSE ajoutée avec succès."
        )

        # -----------------------------------------------------
        # REDIRECTION
        # -----------------------------------------------------

        return redirect(
            f"{reverse('l3_qhse_ue')}"
            f"?niveau={niveau}"
            f"&semestre={semestre}"
            f"&session={session}"
        )

    # =========================================================
    # AFFICHAGE DU FORMULAIRE
    # =========================================================

    return render(
        request,
        "lmd/l3_qhse/ue_add.html",
        {
            "titre": "Nouvelle UE - Management QHSE",
            "filiere": filiere,
            "niveau": niveau,
            "semestre": semestre,
            "session": session,
            "grandes_unites": grandes_unites,
        }
    )

def l3_qhse_ue_update(request, pk):

    ue = get_object_or_404(
        UE,
        pk=pk
    )


    if request.method == "POST":

        ue.code = request.POST.get("code")
        ue.libelle = request.POST.get("libelle")
        ue.credit = request.POST.get("credit")

        ue.save()

        return redirect(
            "l3_qhse_ue"
        )


    return render(
        request,
        "lmd/l3_qhse/ue_update.html",
        {
            "ue": ue
        }
    )

def l3_qhse_ue_delete(request, pk):

    ue = get_object_or_404(
        UE,
        pk=pk
    )


    if request.method == "POST":

        ue.delete()

        return redirect(
            "l3_qhse_ue"
        )


    return redirect(
        "l3_qhse_ue"
    )

def master_uePAS(request, id):

    programme = get_object_or_404(
        MasterProgramme,
        id=id
    )

    ues = MasterUE.objects.filter(
        programme=programme
    )

    return render(
        request,
        "lmd/master/ue_list.html",
        {
            "programme": programme,
            "ues": ues,
        }
    )

def master_ue(request, id):

    programme = get_object_or_404(
        MasterProgramme,
        id=id
    )


    ues = MasterUE.objects.filter(
        programme=programme
    ).prefetch_related(
        "ecues"
    )



    return render(
        request,
        "lmd/master/ue_list.html",
        {
            "programme": programme,
            "ues": ues
        }
    )
 
def master_ue_add(request,id):

    programme = get_object_or_404(
        MasterProgramme,
        id=id
    )


    if request.method=="POST":


        MasterUE.objects.create(

            programme=programme,

            code=request.POST["code"],

            libelle=request.POST["libelle"],

            credit=request.POST["credit"],

            semestre=request.POST["semestre"]

        )


        messages.success(
            request,
            "UE ajoutée avec succès"
        )


        return redirect(
            "master_ue",
            id=id
        )



    return render(
        request,
        "lmd/master/ue_form.html",
        {
            "programme":programme
        }
    )


def master_ue_edit(request,id):

    ue = get_object_or_404(
        MasterUE,
        id=id
    )


    if request.method=="POST":

        form = MasterUEForm(
            request.POST,
            instance=ue
        )


        if form.is_valid():

            form.save()

            return redirect(
                "master_ue",
                ue.programme.id
            )


    else:

        form = MasterUEForm(
            instance=ue
        )


    return render(
        request,
        "lmd/master/ue_form.html",
        {
            "form":form,
            "titre":"Modifier UE Master"
        }
    )

def master_ue_delete(request,id):

    ue = get_object_or_404(
        MasterUE,
        id=id
    )

    programme_id = ue.programme.id


    ue.delete()


    return redirect(
        "master_ue",
        programme_id
    )

def master_ecue(request, id):

    ue = get_object_or_404(
        MasterUE,
        id=id
    )


    ecues = ue.ecues.all()


    return render(
        request,
        "lmd/master/ecue_list.html",
        {
            "ue": ue,
            "ecues": ecues
        }
    )
    
def master_ecue_add(request, id):

    ue = get_object_or_404(
        MasterUE,
        id=id
    )


    if request.method == "POST":


        form = MasterECUEForm(
            request.POST
        )


        if form.is_valid():


            ecue = form.save(
                commit=False
            )


            ecue.ue = ue


            ecue.save()


            return redirect(
                "master_ecue",
                ue.id
            )


    else:


        form = MasterECUEForm()



    return render(
        request,
        "lmd/master/ecue_form.html",
        {
            "form":form,
            "titre":"Ajouter ECUE",
            "ue":ue
        }
    ) 
    
def master_programme_list(request):
    programmes = MasterProgramme.objects.all()

    return render(
        request,
        "lmd/master/programmes/master_programme_list.html",
        {
            "programmes": programmes
        }
    )
    

@login_required(login_url="login")
@role_required("ADMIN")
def imprimer_bulletin_licence_qhse(request, pk, semestre):

    print("====================================")
    print("IMPRESSION QHSE")
    print("PK :", pk)
    print("SEMESTRE :", semestre)

    try:

        etudiant = get_object_or_404(
            EtudiantLMD,
            pk=pk
        )

        print("Étudiant :", etudiant)
        print("Matricule :", etudiant.matricule)
        print("Filière :", etudiant.filiere)
        print("Niveau :", etudiant.niveau)

        pdf_dir = os.path.join(
            settings.MEDIA_ROOT,
            "bulletins_licence_qhse"
        )

        os.makedirs(
            pdf_dir,
            exist_ok=True
        )

        fichier = os.path.join(
            pdf_dir,
            f"bulletin_{etudiant.matricule}_{semestre}.pdf"
        )

        print("Fichier :", fichier)

        resultat = generer_bulletin_qhse_pdf(
            etudiant,
            semestre,
            fichier
        )

        print("Résultat génération :", resultat)

        print("Fichier existe :", os.path.exists(fichier))

        if not os.path.exists(fichier):
            raise Exception(
                f"Le fichier PDF n'a pas été créé : {fichier}"
            )

        print("Taille :", os.path.getsize(fichier))

        return FileResponse(
            open(fichier, "rb"),
            content_type="application/pdf",
            filename=f"Bulletin_{etudiant.matricule}_{semestre}.pdf"
        )

    except Exception as ex:

        print("====================================")
        print("ERREUR PDF QHSE")
        print(type(ex).__name__)
        print(str(ex))
        print("====================================")

        raise
    
def l3_tc_ue_list(request):

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Droit et Gestion"
    )


    ues = (
        UE.objects
        .filter(
            filiere=filiere
        )
        .prefetch_related(
            "ecues"
        )
    )


    return render(
        request,
        "lmd/trom_commun/ue.html",
        {
            "ues": ues,
        },
    )
    
def l3_tc_ue_add(request):
    filiere = (
        FiliereLMD.objects
        .filter(
            Q(libelle="Gestion et Droit") |
            Q(libelle="Droit et Gestion")
        )
        .first()
    )

    if not filiere:
        messages.warning(
            request,
            "La filière Tronc Commun n'existe pas. "
            "Veuillez d'abord créer la filière 'Gestion et Droit' "
            "ou 'Droit et Gestion'."
        )
        return redirect("filiere_lmd_list")

    if request.method == "POST":
        form = UEForm(request.POST)

        if form.is_valid():
            ue = form.save(commit=False)
            ue.filiere = filiere
            ue.save()

            messages.success(
                request,
                "UE ajoutée avec succès."
            )

            return redirect("l3_tc_ue_list")

    else:
        form = UEForm()

    return render(
        request,
        "lmd/trom_commun/eu_form.html",
        {
            "form": form,
            "titre": "Ajouter une UE Tronc Commun",
            "filiere": filiere,
        }
    )


def l3_tc_ue_update(request, pk):
    ue = get_object_or_404(
        UE,
        pk=pk
    )

    if request.method == "POST":
        form = UEForm(
            request.POST,
            instance=ue
        )

        if form.is_valid():
            form.save()

            messages.success(
                request,
                "UE modifiée avec succès."
            )

            return redirect("l3_tc_ue_list")

    else:
        form = UEForm(
            instance=ue
        )

    return render(
        request,
        "lmd/trom_commun/eu_form.html",
        {
            "form": form,
            "titre": "Modifier une UE",
            "filiere": ue.filiere,
        }
    )
    

def l3_tc_ue_delete(request, pk):

    ue = get_object_or_404(
        UE,
        pk=pk
    )


    if request.method == "POST":

        ue.delete()

        return redirect(
            "l3_tc_ue_list"
        )


    return render(
        request,
        "lmd/trom_commun/ue_delete.html",
        {
            "objet": ue
        }
    )

def l3_tc_ecue_add(request, ue_id):

    ue = get_object_or_404(
        UE,
        id=ue_id
    )


    if request.method == "POST":

        form = ECUEForm(request.POST)


        if form.is_valid():

            ecue = form.save(commit=False)

            ecue.ue = ue

            ecue.save()


            return redirect(
                "l3_tc_ue_list"
            )


    else:

        form = ECUEForm()


    return render(
        request,
        "lmd/trom_commun/ecue_form.html",
        {
            "form": form,
            "ue": ue,
            "titre": "Ajouter un ECUE"
        }
    )

def l3_tc_ecue_update(request, pk):

    ecue = get_object_or_404(
        ECUE,
        id=pk
    )


    if request.method == "POST":

        form = ECUEForm(
            request.POST,
            instance=ecue
        )


        if form.is_valid():

            form.save()

            return redirect(
                "l3_tc_ue_list"
            )


    else:

        form = ECUEForm(
            instance=ecue
        )


    return render(
        request,
        "lmd/trom_commun/ecue_form.html",
        {
            "form": form,
            "titre": "Modifier ECUE"
        }
    )

def l3_tc_ecue_delete(request, pk):

    ecue = get_object_or_404(
        ECUE,
        id=pk
    )


    if request.method == "POST":

        ecue.delete()

        return redirect(
            "l3_tc_ue_list"
        )


    return render(
        request,
        "lmd/trom_commun/ecue_delete.html",
        {
            "objet": ecue
        }
    )

@login_required(login_url="login")
@role_required("ADMIN")    
def bulletin_rattrapage_pdf(request, id, semestre):

    etudiant = get_object_or_404(
        EtudiantLMD,
        id=id
    )


    filiere = etudiant.filiere.libelle.lower()


    if "gestion" in filiere and "droit" in filiere:

        return bulletin_tronc_commun_pdf(
            request,
            id,
            semestre
        )


    elif "droit privé" in filiere:

        return pdf_droit_prive_service_pdf(
            request,
            id,
            semestre
        )


    elif "qhse" in filiere:

        return pdf_licence_qhse(
            request,
            id,
            semestre
        )

    else:

        return HttpResponse(
            "Aucun service PDF trouvé pour cette filière"
        )    
        
def import_tronc_commun_excel(request):
    if request.method == "POST":

        fichier = request.FILES.get("excel_file")

        if not fichier:
            messages.error(request, "Veuillez sélectionner un fichier Excel.")
            return redirect("import_tronc_commun_excel")

        if not fichier.name.endswith(".xlsx"):
            messages.error(request, "Le fichier doit être au format .xlsx")
            return redirect("import_tronc_commun_excel")

        wb = load_workbook(fichier)
        ws = wb.active

        importes = 0
        erreurs = []

        for numero, ligne in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

            matricule = str(ligne[0]).strip() if ligne[0] else ""
            nom = str(ligne[1]).strip().upper() if ligne[1] else ""
            prenoms = str(ligne[2]).strip().title() if ligne[2] else ""
            sexe = str(ligne[3]).strip().upper() if ligne[3] else ""
            niveau = str(ligne[4]).strip().upper() if ligne[4] else ""
            filiere_nom = str(ligne[5]).strip() if ligne[5] else ""

            if not matricule:
                erreurs.append(f"Ligne {numero}: matricule vide.")
                continue

            if EtudiantLMD.objects.filter(matricule=matricule).exists():
                erreurs.append(f"Ligne {numero}: {matricule} existe déjà.")
                continue

            filiere = FiliereLMD.objects.filter(libelle__iexact=filiere_nom).first()

            if not filiere:
                erreurs.append(
                    f"Ligne {numero}: filière '{filiere_nom}' introuvable."
                )
                continue

            EtudiantLMD.objects.create(
                matricule=matricule,
                nom=nom,
                prenoms=prenoms,
                sexe=sexe,
                niveau=niveau,
                filiere=filiere,
            )

            importes += 1

        if importes:
            messages.success(
                request,
                f"{importes} étudiant(s) importé(s) avec succès."
            )

        for err in erreurs:
            messages.warning(request, err)

        return redirect("liste_etudiants_tronc_commun")

    return render(
        request,
        "lmd/trom_commun/import_excel.html",
    )
    


def tronc_commun_delete(request, id):

    etudiant = get_object_or_404(
        EtudiantLMD,
        id=id
    )

    etudiant.delete()

    messages.success(
        request,
        "Étudiant supprimé avec succès."
    )

    return redirect(
        "liste_etudiants_tronc_commun"
    )



from datetime import datetime
from datetime import datetime, date


def convertir_date(date_value):

    if not date_value:
        return None


    # Si Excel retourne déjà une date
    if isinstance(date_value, datetime):

        return date_value.date()


    if isinstance(date_value, date):

        return date_value



    if isinstance(date_value, str):

        formats = [

            "%d/%m/%Y",   # 15/02/2000
            "%d-%m-%Y",   # 15-02-2000
            "%Y-%m-%d",   # 2000-02-15
            "%Y/%m/%d",   # 2000/02/15
            "%d.%m.%Y",   # 15.02.2000
            "%m/%d/%Y",   # 02/15/2000

        ]


        for fmt in formats:

            try:

                return datetime.strptime(
                    date_value.strip(),
                    fmt
                ).date()


            except ValueError:

                continue



    return None

def l3_droit_etudiant_import(request):

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Droit Privé"
    )


    colonnes_attendues = [
        "Matricule",
        "Nom",
        "Prénoms",
        "Lieu naissance",
        "Date naissance",
        "Email",
        "Annee academique",
        "Niveau",
        "Sexe",
        "Téléphone",
    ]


    if request.method == "POST":


        fichier = request.FILES.get("fichier")


        if not fichier:

            messages.error(
                request,
                "❌ Aucun fichier sélectionné."
            )

            return redirect(
                "droit_prive_etudiants"
            )

        # Vérification extension

        if not fichier.name.endswith(".xlsx"):

            messages.error(
                request,
                "❌ Format incorrect. Veuillez importer un fichier Excel (.xlsx)."
            )

            return redirect(
                "droit_prive_etudiants"
            )

        try:
            workbook = load_workbook(
                fichier
            )

            sheet = workbook.active

        except Exception:

            messages.error(
                request,
                "❌ Impossible de lire le fichier Excel."
            )

            return redirect(
                "droit_prive_etudiants"
            )



        # Vérification des entêtes

        headers = [
            str(cell.value).strip()
            for cell in sheet[1]
            if cell.value
        ]



        if headers != colonnes_attendues:

            messages.error(
                request,
                "❌ Format du fichier incorrect. "
                "Veuillez télécharger et utiliser le modèle Excel fourni."
            )

            return redirect(
                "droit_prive_etudiants"
            )



        total = 0
        doublons = 0
        erreurs = 0



        for index, row in enumerate(
            sheet.iter_rows(
                min_row=2,
                values_only=True
            ),
            start=2
        ):


            # supprimer les cellules vides à la fin

            row = list(row)

            while row and row[-1] is None:
                row.pop()



            # ignorer ligne vide

            if not row:
                continue



            # vérifier nombre de colonnes

            if len(row) != 10:

                messages.warning(
                    request,
                    f"⚠️ Ligne {index} ignorée : "
                    f"{len(row)} colonnes trouvées au lieu de 10."
                )

                erreurs += 1

                continue



            (
                matricule,
                nom,
                prenoms,
                lieu_naissance,
                date_naissance,
                email,
                annee_academique,
                niveau,
                sexe,
                telephone

            ) = row




            # champs obligatoires

            if not matricule or not nom or not prenoms:


                messages.warning(
                    request,
                    f"⚠️ Ligne {index} ignorée : "
                    "Matricule, nom ou prénom manquant."
                )

                erreurs += 1

                continue




            # doublon matricule

            if EtudiantLMD.objects.filter(
                matricule=matricule
            ).exists():


                doublons += 1

                continue

            # conversion date Excel

            # if isinstance(
            #     date_naissance,
            #     datetime
            # ):
            date_naissance = convertir_date(
               date_naissance
               )

            EtudiantLMD.objects.create(

                matricule=matricule,

                nom=nom,

                prenoms=prenoms,

                lieu_naissance=lieu_naissance,

                date_naissance=date_naissance,

                email=email,

                annee_academique=annee_academique,

                niveau=niveau,

                sexe=sexe,

                telephone=telephone,

                filiere=filiere

            )



            total += 1




        messages.success(
            request,
            f"✅ {total} étudiant(s) importé(s) avec succès."
        )



        if doublons:

            messages.warning(
                request,
                f"⚠️ {doublons} étudiant(s) déjà existant(s) ignoré(s)."
            )



        if erreurs:

            messages.warning(
                request,
                f"⚠️ {erreurs} ligne(s) non importée(s)."
            )



        return redirect(
            "droit_prive_etudiants"
        )



    return render(
        request,
        "lmd/l3/droit/import_etudiants.html"
    )
    
from openpyxl import Workbook



def l3_droit_etudiant_modele_excel(request):

    wb = Workbook()

    ws = wb.active

    ws.title = "Etudiants Droit Privé"


    headers = [

        "Matricule",
        "Nom",
        "Prénoms",
        "Lieu naissance",
        "Date naissance",
        "Email",
        "Annee academique",
        "Niveau",
        "Sexe",
        "Téléphone",

    ]


    ws.append(headers)



    # exemple
    ws.append([

        "DP001",
        "KONE",
        "Lacina",
        "Abidjan",
        "15/02/2000",
        "lacina@gmail.com",
        "2025-2026",
        "L3",
        "M",
        "07000000",
        "S1"

    ])



    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


    response["Content-Disposition"] = (
        'attachment; filename="modele_import_etudiants_droit_prive.xlsx"'
    )


    wb.save(response)


    return response

@login_required(login_url="login")
@role_required("ADMIN")
def imprimer_bulletin_droit_prive(request, id, semestre):

    etudiant = get_object_or_404(
        EtudiantLMD,
        id=id,
        filiere__libelle="Droit Privé"
    )


    pdf_dir = os.path.join(
        settings.MEDIA_ROOT,
        "bulletins"
    )

    os.makedirs(
        pdf_dir,
        exist_ok=True
    )


    file_path = os.path.join(
        pdf_dir,
         f"bulletin_{etudiant.matricule}_{etudiant.niveau}_{semestre}.pdf"
    )


    generer_bulletin_droit_prive_pdf(
        etudiant,
        semestre,
        file_path
    )


    return FileResponse(
        open(file_path,"rb"),
        content_type="application/pdf"
    )
@login_required(login_url="login")
@role_required("ADMIN")    
def imprimer_bulletin_lmd(request, id, semestre):

    etudiant = get_object_or_404(
        EtudiantLMD,
        id=id
    )


    pdf_dir = os.path.join(
        settings.MEDIA_ROOT,
        "bulletins"
    )

    os.makedirs(
        pdf_dir,
        exist_ok=True
    )


    file_path = os.path.join(
        pdf_dir,
        f"bulletin_{etudiant.matricule}_{semestre}.pdf"
    )


    # ==============================
    # CHOIX DU BULLETIN
    # ==============================

    if (
        etudiant.filiere.libelle == "Droit Privé"
        and etudiant.niveau == "L2"
    ):

        generer_bulletin_licence2_droit_prive_pdf(
            etudiant,
            semestre,
            file_path
        )


    elif etudiant.filiere.libelle == "Droit Privé":

        generer_bulletin_droit_prive_pdf(
            etudiant,
            semestre,
            file_path
        )


    else:

        # futur générateur pour les autres filières
        generer_bulletin_droit_prive_pdf(
            etudiant,
            semestre,
            file_path
        )


    return FileResponse(
        open(file_path, "rb"),
        content_type="application/pdf"
    )
    
def imprimer_bulletin_lmd(request, id, semestre):

    etudiant = get_object_or_404(
        EtudiantLMD,
        id=id
    )


    pdf_dir = os.path.join(
        settings.MEDIA_ROOT,
        "bulletins"
    )

    os.makedirs(
        pdf_dir,
        exist_ok=True
    )


    file_path = os.path.join(
        pdf_dir,
        f"bulletin_{etudiant.matricule}_{semestre}.pdf"
    )


    generer_bulletin_lmd_pdf(
        etudiant=etudiant,
        semestre=semestre,
        file_path=file_path
    )


    return FileResponse(
        open(file_path, "rb"),
        content_type="application/pdf"
    )
    
def gestion_etudiants_sciences_gestion(request, niveau):
    """
    Liste les étudiants de Sciences de Gestion
    pour le niveau demandé : L1, L2 ou L3.
    """

    niveaux_valides = ["L1", "L2", "L3"]

    if niveau not in niveaux_valides:
        return render(
            request,
            "lmd/erreur.html",
            {
                "message": "Niveau invalide."
            }
        )

    etudiants = (
        EtudiantLMD.objects
        .filter(
            filiere__libelle__icontains="Sciences de Gestion",
            niveau=niveau
        )
        .order_by("nom", "prenoms")
    )

    return render(
        request,
        "lmd/gestion/etudiants/etudiants.html",
        {
            "etudiants": etudiants,
            "niveau": niveau,
        }
    )


    

def gestion_etudiant_list(request):
    """
    Liste des étudiants de Sciences de Gestion
    pour les niveaux L1, L2 et L3.
    """

    niveau = request.GET.get("niveau", "").strip()
    recherche = request.GET.get("q", "").strip()

    etudiants = (
        EtudiantLMD.objects
        .filter(
            filiere__libelle__icontains="Sciences de Gestion",
            niveau__in=["L1", "L2", "L3"],
        )
        .order_by("niveau", "nom", "prenoms")
    )

    # Filtre par niveau
    if niveau in ["L1", "L2", "L3"]:
        etudiants = etudiants.filter(niveau=niveau)

    # Recherche
    if recherche:
        etudiants = etudiants.filter(
            matricule__icontains=recherche
        ) | etudiants.filter(
            nom__icontains=recherche
        ) | etudiants.filter(
            prenoms__icontains=recherche
        )

    return render(
        request,
        "lmd/gestion/etudiants/list.html",
        {
            "etudiants": etudiants,
            "niveau_selectionne": niveau,
            "recherche": recherche,
        },
    )


def gestion_etudiant_add(request):

    filiere = get_object_or_404(
        FiliereLMD,
        libelle__icontains="Sciences de Gestion",
    )

    if request.method == "POST":

        matricule = request.POST.get("matricule", "").strip()
        nom = request.POST.get("nom", "").strip()
        prenoms = request.POST.get("prenoms", "").strip()
        sexe = request.POST.get("sexe", "").strip()
        date_naissance = request.POST.get("date_naissance") or None
        lieu_naissance = request.POST.get("lieu_naissance", "").strip()
        statut = request.POST.get("statut", "").strip()
        niveau = request.POST.get("niveau", "").strip()
        annee_academique = request.POST.get(
            "annee_academique",
            "2025-2026",
        ).strip()

        # =========================
        # VALIDATIONS
        # =========================

        if not matricule or not nom or not prenoms:
            messages.error(
                request,
                "Le matricule, le nom et les prénoms sont obligatoires.",
            )

            return render(
                request,
                "lmd/gestion/etudiants/form.html",
                {
                    "etudiant": None,
                    "filiere": filiere,
                },
            )

        if niveau not in ["L1", "L2", "L3"]:
            messages.error(
                request,
                "Veuillez sélectionner un niveau valide.",
            )

            return render(
                request,
                "lmd/gestion/etudiants/form.html",
                {
                    "etudiant": None,
                    "filiere": filiere,
                },
            )

        if EtudiantLMD.objects.filter(
            matricule=matricule
        ).exists():

            messages.error(
                request,
                f"Le matricule {matricule} existe déjà.",
            )

            return render(
                request,
                "lmd/gestion/etudiants/form.html",
                {
                    "etudiant": None,
                    "filiere": filiere,
                },
            )

        # =========================
        # CREATION
        # =========================

        EtudiantLMD.objects.create(
            matricule=matricule,
            nom=nom,
            prenoms=prenoms,
            sexe=sexe,
            date_naissance=date_naissance,
            lieu_naissance=lieu_naissance,
            statut=statut,
            niveau=niveau,
            filiere=filiere,
            annee_academique=annee_academique,
        )

        messages.success(
            request,
            f"L'étudiant {nom} {prenoms} a été ajouté avec succès.",
        )

        return redirect("gestion_etudiant_list")

    return render(
        request,
        "lmd/gestion/etudiants/form.html",
        {
            "etudiant": None,
            "filiere": filiere,
        },
    )


def gestion_etudiant_edit(request, id):

    etudiant = get_object_or_404(
        EtudiantLMD,
        id=id,
        filiere__libelle__icontains="Sciences de Gestion",
        niveau__in=["L1", "L2", "L3"],
    )

    if request.method == "POST":

        matricule = request.POST.get("matricule", "").strip()
        nom = request.POST.get("nom", "").strip()
        prenoms = request.POST.get("prenoms", "").strip()
        sexe = request.POST.get("sexe", "").strip()
        date_naissance = request.POST.get("date_naissance") or None
        lieu_naissance = request.POST.get("lieu_naissance", "").strip()
        statut = request.POST.get("statut", "").strip()
        niveau = request.POST.get("niveau", "").strip()
        annee_academique = request.POST.get(
            "annee_academique",
            etudiant.annee_academique,
        ).strip()

        if not matricule or not nom or not prenoms:
            messages.error(
                request,
                "Le matricule, le nom et les prénoms sont obligatoires.",
            )

            return render(
                request,
                "lmd/gestion/etudiants/form.html",
                {
                    "etudiant": etudiant,
                    "filiere": etudiant.filiere,
                },
            )

        if niveau not in ["L1", "L2", "L3"]:
            messages.error(
                request,
                "Veuillez sélectionner un niveau valide.",
            )

            return render(
                request,
                "lmd/gestion/etudiants/form.html",
                {
                    "etudiant": etudiant,
                    "filiere": etudiant.filiere,
                },
            )

        matricule_existe = (
            EtudiantLMD.objects
            .filter(matricule=matricule)
            .exclude(id=etudiant.id)
            .exists()
        )

        if matricule_existe:

            messages.error(
                request,
                f"Le matricule {matricule} est déjà utilisé.",
            )

            return render(
                request,
                "lmd/gestion/etudiants/form.html",
                {
                    "etudiant": etudiant,
                    "filiere": etudiant.filiere,
                },
            )

        # =========================
        # MODIFICATION
        # =========================

        etudiant.matricule = matricule
        etudiant.nom = nom
        etudiant.prenoms = prenoms
        etudiant.sexe = sexe
        etudiant.date_naissance = date_naissance
        etudiant.lieu_naissance = lieu_naissance
        etudiant.statut = statut
        etudiant.niveau = niveau
        etudiant.annee_academique = annee_academique

        # La filière reste Sciences de Gestion
        etudiant.save()

        messages.success(
            request,
            f"L'étudiant {nom} {prenoms} a été modifié avec succès.",
        )

        return redirect("gestion_etudiant_list")

    return render(
        request,
        "lmd/gestion/etudiants/form.html",
        {
            "etudiant": etudiant,
            "filiere": etudiant.filiere,
        },
    )


def gestion_etudiant_delete(request, id):

    etudiant = get_object_or_404(
        EtudiantLMD,
        id=id,
        filiere__libelle__icontains="Sciences de Gestion",
        niveau__in=["L1", "L2", "L3"],
    )

    if request.method == "POST":

        nom_complet = f"{etudiant.nom} {etudiant.prenoms}"

        etudiant.delete()

        messages.success(
            request,
            f"L'étudiant {nom_complet} a été supprimé avec succès.",
        )

        return redirect("gestion_etudiant_list")

    return render(
        request,
        "lmd/gestion/etudiants/delete.html",
        {
            "etudiant": etudiant,
        },
    )



def l3_gestion_etudiant_add(request):
    filiere = get_object_or_404(
        FiliereLMD,
        libelle__icontains="Sciences de Gestion",
    )

    if request.method == "POST":
        matricule = request.POST.get("matricule", "").strip()
        nom = request.POST.get("nom", "").strip()
        prenoms = request.POST.get("prenoms", "").strip()
        sexe = request.POST.get("sexe", "").strip()
        date_naissance = request.POST.get("date_naissance") or None
        lieu_naissance = request.POST.get("lieu_naissance", "").strip()
        statut = request.POST.get("statut", "").strip()
        annee_academique = request.POST.get(
            "annee_academique",
            "2025-2026",
        ).strip()

        if not matricule or not nom or not prenoms:
            messages.error(
                request,
                "Le matricule, le nom et les prénoms sont obligatoires.",
            )
            return render(
                request,
                "lmd/l3/gestion/form.html",
                {
                    "etudiant": None,
                    "filiere": filiere,
                },
            )

        if EtudiantLMD.objects.filter(matricule=matricule).exists():
            messages.error(
                request,
                f"Le matricule {matricule} existe déjà.",
            )
            return render(
                request,
                "lmd/l3/gestion/form.html",
                {
                    "etudiant": None,
                    "filiere": filiere,
                },
            )

        EtudiantLMD.objects.create(
            matricule=matricule,
            nom=nom,
            prenoms=prenoms,
            sexe=sexe,
            date_naissance=date_naissance,
            lieu_naissance=lieu_naissance,
            statut=statut,
            niveau="L3",
            filiere=filiere,
            annee_academique=annee_academique,
        )

        messages.success(
            request,
            f"L'étudiant {nom} {prenoms} a été ajouté avec succès.",
        )

        return redirect("l3_gestion_etudiant_list")

    return render(
        request,
        "lmd/l3/gestion/form.html",
        {
            "etudiant": None,
            "filiere": filiere,
        },
    )


def l3_gestion_etudiant_edit(request, id):
    etudiant = get_object_or_404(
        EtudiantLMD,
        id=id,
        niveau="L3",
        filiere__libelle__icontains="Sciences de Gestion",
    )

    if request.method == "POST":
        matricule = request.POST.get("matricule", "").strip()
        nom = request.POST.get("nom", "").strip()
        prenoms = request.POST.get("prenoms", "").strip()
        sexe = request.POST.get("sexe", "").strip()
        date_naissance = request.POST.get("date_naissance") or None
        lieu_naissance = request.POST.get("lieu_naissance", "").strip()
        statut = request.POST.get("statut", "").strip()
        annee_academique = request.POST.get(
            "annee_academique",
            etudiant.annee_academique,
        ).strip()

        if not matricule or not nom or not prenoms:
            messages.error(
                request,
                "Le matricule, le nom et les prénoms sont obligatoires.",
            )
            return render(
                request,
                "lmd/l3/gestion/form.html",
                {
                    "etudiant": etudiant,
                    "filiere": etudiant.filiere,
                },
            )

        matricule_existe = (
            EtudiantLMD.objects
            .filter(matricule=matricule)
            .exclude(id=etudiant.id)
            .exists()
        )

        if matricule_existe:
            messages.error(
                request,
                f"Le matricule {matricule} est déjà utilisé.",
            )
            return render(
                request,
                "lmd/l3/gestion/form.html",
                {
                    "etudiant": etudiant,
                    "filiere": etudiant.filiere,
                },
            )

        etudiant.matricule = matricule
        etudiant.nom = nom
        etudiant.prenoms = prenoms
        etudiant.sexe = sexe
        etudiant.date_naissance = date_naissance
        etudiant.lieu_naissance = lieu_naissance
        etudiant.statut = statut
        etudiant.niveau = "L3"
        etudiant.annee_academique = annee_academique

        # On conserve volontairement la filière Sciences de Gestion.
        etudiant.save()

        messages.success(
            request,
            f"L'étudiant {nom} {prenoms} a été modifié avec succès.",
        )

        return redirect("l3_gestion_etudiant_list")

    return render(
        request,
        "lmd/l3/gestion/form.html",
        {
            "etudiant": etudiant,
            "filiere": etudiant.filiere,
        },
    )


def l3_gestion_etudiant_delete(request, id):
    etudiant = get_object_or_404(
        EtudiantLMD,
        id=id,
        niveau="L3",
        filiere__libelle__icontains="Sciences de Gestion",
    )

    if request.method == "POST":
        nom_complet = f"{etudiant.nom} {etudiant.prenoms}"

        etudiant.delete()

        messages.success(
            request,
            f"L'étudiant {nom_complet} a été supprimé avec succès.",
        )

        return redirect("l3_gestion_etudiant_list")

    return render(
        request,
        "lmd/l3/gestion/delete.html",
        {
            "etudiant": etudiant,
        },
    )
    

# ============================================================
# SCIENCES DE GESTION - ÉTUDIANTS L1 / L2 / L3
# ============================================================

def sciences_gestion_etudiant_addrrr(request):

  if request.method == "POST":

    matricule = request.POST.get("matricule")
    nom = request.POST.get("nom")
    prenoms = request.POST.get("prenoms")
    niveau = request.POST.get("niveau")

    sexe = request.POST.get("sexe")
    date_naissance = request.POST.get("date_naissance")
    lieu_naissance = request.POST.get("lieu_naissance")
    statut = request.POST.get("statut")
    annee_academique = request.POST.get(
        "annee_academique",
        "2025-2026"
    )

    # Vérification du niveau
    if niveau not in ["L1", "L2", "L3"]:
        messages.error(
            request,
            "Veuillez sélectionner un niveau valide."
        )

        return render(
            request,
            "lmd/l3/gestion/form.html",
        )

    # Récupération de la filière
    filiere = FiliereLMD.objects.get(
        libelle__icontains="Sciences de Gestion"
    )

    # Création de l'étudiant
    EtudiantLMD.objects.create(
        matricule=matricule,
        nom=nom,
        prenoms=prenoms,
        niveau=niveau,
        filiere=filiere,
        sexe=sexe or None,
        date_naissance=date_naissance or None,
        lieu_naissance=lieu_naissance,
        statut=statut,
        annee_academique=annee_academique,
    )

    messages.success(
        request,
        f"Étudiant ajouté avec succès en {niveau} Sciences de Gestion."
    )

    return redirect("sciences_gestion_etudiants")

  return render(
     request,
    "lmd/l3/gestion/form.html"
)

def sciences_gestion_etudiant_add(request):

    # Récupération de la filière
    filiere = FiliereLMD.objects.filter(
        libelle__icontains="Sciences de Gestion"
    ).first()

    if not filiere:
        messages.error(
            request,
            "La filière Sciences de Gestion est introuvable."
        )
        return redirect("sciences_gestion_etudiants")

    if request.method == "POST":

        matricule = request.POST.get("matricule", "").strip()
        nom = request.POST.get("nom", "").strip()
        prenoms = request.POST.get("prenoms", "").strip()
        niveau = request.POST.get("niveau", "").strip()

        sexe = request.POST.get("sexe", "").strip()
        date_naissance = request.POST.get("date_naissance") or None
        lieu_naissance = request.POST.get("lieu_naissance", "").strip()
        statut = request.POST.get("statut", "").strip()

        annee_academique = request.POST.get(
            "annee_academique",
            "2025-2026"
        ).strip()

        # Vérification du niveau
        if niveau not in ["L1", "L2", "L3"]:

            messages.error(
                request,
                "Veuillez sélectionner un niveau valide."
            )

            return render(
                request,
                "lmd/l3/gestion/form.html",
                {
                    "niveau": niveau,
                }
            )

        # Création de l'étudiant
        EtudiantLMD.objects.create(
            matricule=matricule,
            nom=nom,
            prenoms=prenoms,
            niveau=niveau,
            filiere=filiere,
            sexe=sexe or None,
            date_naissance=date_naissance,
            lieu_naissance=lieu_naissance,
            statut=statut,
            annee_academique=annee_academique,
        )

        messages.success(
            request,
            f"Étudiant ajouté avec succès en {niveau} Sciences de Gestion."
        )

        return redirect("sciences_gestion_etudiants")

    return render(
        request,
        "lmd/l3/gestion/form.html",
        {
            "niveau": "",
        }
    )

def sciences_gestion_etudiant_edit(request, id):

    etudiant = get_object_or_404(
        EtudiantLMD,
        id=id,
        filiere__libelle__icontains="Sciences de Gestion"
    )

    if request.method == "POST":

        matricule = request.POST.get("matricule")
        nom = request.POST.get("nom")
        prenoms = request.POST.get("prenoms")
        niveau = request.POST.get("niveau")

        sexe = request.POST.get("sexe")
        if not sexe:
            # Si aucun sexe n'est sélectionné,
            # on conserve l'ancienne valeur
            sexe = etudiant.sexe

        
        date_naissance = request.POST.get("date_naissance")
        lieu_naissance = request.POST.get("lieu_naissance")
        statut = request.POST.get("statut")
        annee_academique = request.POST.get(
            "annee_academique",
            "2025-2026"
        )

        # Vérification du niveau
        if niveau not in ["L1", "L2", "L3"]:

            messages.error(
                request,
                "Veuillez sélectionner un niveau valide."
            )

            return render(
                request,
                "lmd/l3/gestion/form.html",
                {
                    "etudiant": etudiant,
                }
            )

        # Mise à jour
        etudiant.matricule = matricule
        etudiant.nom = nom
        etudiant.prenoms = prenoms
        etudiant.niveau = niveau
        # etudiant.sexe = sexe or None
        etudiant.sexe = sexe
        etudiant.date_naissance = date_naissance or None
        etudiant.lieu_naissance = lieu_naissance
        etudiant.statut = statut
        etudiant.annee_academique = annee_academique

        etudiant.save()

        messages.success(
            request,
            f"Étudiant {nom} {prenoms} modifié avec succès."
        )

        return redirect(
            "l3_sciences_gestion_etudiants"
        )

    return render(
        request,
        "lmd/l3/gestion/form.html",
        {
            "etudiant": etudiant,
        }
    )


def sciences_gestion_etudiant_deleteH(request, niveau, id):

    if niveau not in ["L1", "L2", "L3"]:
        raise Http404("Niveau invalide")

    etudiant = get_object_or_404(
        EtudiantLMD,
        id=id,
        niveau=niveau,
        filiere__libelle__iexact="Sciences de Gestion",
    )

    if request.method == "POST":

        etudiant.delete()

        return redirect(
            "l1_gestion_etudiants"
            if niveau == "L1"
            else "l2_gestion_etudiants"
            if niveau == "L2"
            else "l3_gestion_etudiants"
        )

    return render(
        request,
        "lmd/l3/gestion/confirm_delete.html",
        {
            "etudiant": etudiant,
            "niveau": niveau,
        }
    )
    
@login_required(login_url="login")
@role_required("ADMIN")
def sciences_gestion_bulletins(request):
  etudiants = EtudiantLMD.objects.filter(
     filiere__libelle__icontains="Sciences de Gestion"
  ).order_by(
    "niveau",
    "nom",
    "prenoms"
   )

   # Recherche
  q = request.GET.get("q", "").strip()
  if q:
    etudiants = etudiants.filter(
        Q(matricule__icontains=q)
        | Q(nom__icontains=q)
        | Q(prenoms__icontains=q)
    )

  # Filtre niveau
  niveau = request.GET.get("niveau", "").strip()

  if niveau:
    etudiants = etudiants.filter(
        niveau=niveau
    )

  return render(
    request,
    "lmd/l3/gestion/bulletins.html",
    {
        "etudiants": etudiants,
    }
)


def sciences_gestion_etudiant_delete(request, id):

    etudiant = get_object_or_404(
        EtudiantLMD,
        id=id,
        filiere__libelle__icontains="Sciences de Gestion"
    )

    nom_complet = f"{etudiant.nom} {etudiant.prenoms}"

    etudiant.delete()

    messages.success(
        request,
        f"L'étudiant {nom_complet} a été supprimé avec succès."
    )

    return redirect(
        "l3_sciences_gestion_etudiants"
    )
    
def l3_droit_grande_unite(request):

    filiere = get_object_or_404(FiliereLMD, libelle="Droit Privé")

    niveau = request.GET.get("niveau")
    semestre = request.GET.get("semestre")

    grandes_unites = GrandeUnite.objects.filter(filiere=filiere)

    if niveau:
        grandes_unites = grandes_unites.filter(niveau=niveau)
    if semestre:
        grandes_unites = grandes_unites.filter(semestre=semestre)

    return render(
        request,
        "lmd/l3/droit/grande_unite_list.html",
        {
            "filiere": filiere,
            "grandes_unites": grandes_unites,
            "niveau": niveau,
            "semestre": semestre,
        }
    )


def l3_droit_grande_unite_add(request):

    filiere = get_object_or_404(FiliereLMD, libelle="Droit Privé")

    if request.method == "POST":

        niveau = request.POST.get("niveau")
        semestre = request.POST.get("semestre")

        GrandeUnite.objects.create(
            code=request.POST.get("code"),
            nom=request.POST.get("nom"),
            ordre=request.POST.get("ordre") or 1,
            niveau=niveau,
            semestre=semestre,
            filiere=filiere,
        )

        messages.success(request, "Grande unité ajoutée avec succès.")

        return redirect(
            f"{reverse('l3_droit_grande_unite')}?niveau={niveau}&semestre={semestre}"
        )

    return render(
        request,
        "lmd/l3/droit/grande_unite_form.html",
        {
            "titre": "Ajouter une grande unité - Droit Privé",
            "filiere": filiere,
            "niveau": request.GET.get("niveau"),
            "semestre": request.GET.get("semestre"),
        }
    )
    
    
def l3_droit_grande_unite_edit(request, pk):

    filiere = get_object_or_404(FiliereLMD, libelle="Droit Privé")
    grande_unite = get_object_or_404(GrandeUnite, pk=pk, filiere=filiere)

    if request.method == "POST":

        niveau = request.POST.get("niveau")
        semestre = request.POST.get("semestre")

        grande_unite.code = request.POST.get("code")
        grande_unite.nom = request.POST.get("nom")
        grande_unite.ordre = request.POST.get("ordre") or 1
        grande_unite.niveau = niveau
        grande_unite.semestre = semestre
        grande_unite.save()

        messages.success(request, "Grande unité modifiée avec succès.")

        return redirect(
            f"{reverse('l3_droit_grande_unite')}?niveau={niveau}&semestre={semestre}"
        )

    return render(
        request,
        "lmd/l3/droit/grande_unite_form.html",
        {
            "titre": "Modifier une grande unité - Droit Privé",
            "filiere": filiere,
            "grande_unite": grande_unite,
            "niveau": request.GET.get("niveau"),
            "semestre": request.GET.get("semestre"),
        }
    )


def l3_droit_grande_unite_delete(request, pk):

    filiere = get_object_or_404(FiliereLMD, libelle="Droit Privé")
    grande_unite = get_object_or_404(GrandeUnite, pk=pk, filiere=filiere)

    niveau = request.GET.get("niveau")
    semestre = request.GET.get("semestre")

    if grande_unite.ues.exists():
        messages.error(
            request,
            "Impossible de supprimer : des UE sont encore rattachées à cette grande unité."
        )
    else:
        grande_unite.delete()
        messages.success(request, "Grande unité supprimée avec succès.")

    return redirect(
        f"{reverse('l3_droit_grande_unite')}?niveau={niveau}&semestre={semestre}"
    )
    
    
def l3_gestion_grande_unite(request):

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Sciences de Gestion"
    )

    niveau = request.GET.get("niveau")
    semestre = request.GET.get("semestre")

    grandes_unites = GrandeUnite.objects.filter(
        filiere=filiere
    ).order_by("niveau", "semestre", "ordre")

    if niveau:
        grandes_unites = grandes_unites.filter(
            niveau=niveau
        )

    if semestre:
        grandes_unites = grandes_unites.filter(
            semestre=semestre
        )

    return render(
        request,
        "lmd/l3/gestion/grande_unite_list.html",
        {
            "filiere": filiere,
            "grandes_unites": grandes_unites,
            "niveau": niveau,
            "semestre": semestre,
        }
    )
    
def l3_gestion_grande_unite_add(request):

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Sciences de Gestion"
    )

    if request.method == "POST":

        niveau = request.POST.get("niveau")
        semestre = request.POST.get("semestre")

        GrandeUnite.objects.create(
            code=request.POST.get("code"),
            nom=request.POST.get("nom"),
            ordre=request.POST.get("ordre") or 1,
            niveau=niveau,
            semestre=semestre,
            filiere=filiere,
        )

        messages.success(
            request,
            "Grande unité ajoutée avec succès."
        )

        return redirect(
            f"{reverse('l3_gestion_grande_unite')}"
            f"?niveau={niveau}&semestre={semestre}"
        )

    return render(
        request,
        "lmd/l3/gestion/grande_unite_form.html",
        {
            "titre": "Ajouter une grande unité - Sciences de Gestion",
            "filiere": filiere,
            "niveau": request.GET.get("niveau"),
            "semestre": request.GET.get("semestre"),
        }
    )

def l3_gestion_grande_unite_edit(request, pk):

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Sciences de Gestion"
    )

    grande_unite = get_object_or_404(
        GrandeUnite,
        pk=pk,
        filiere=filiere
    )

    if request.method == "POST":

        niveau = request.POST.get("niveau")
        semestre = request.POST.get("semestre")

        grande_unite.code = request.POST.get("code")
        grande_unite.nom = request.POST.get("nom")
        grande_unite.ordre = request.POST.get("ordre") or 1
        grande_unite.niveau = niveau
        grande_unite.semestre = semestre

        grande_unite.save()

        messages.success(
            request,
            "Grande unité modifiée avec succès."
        )

        return redirect(
            f"{reverse('l3_gestion_grande_unite')}"
            f"?niveau={niveau}&semestre={semestre}"
        )

    return render(
        request,
        "lmd/l3/gestion/grande_unite_form.html",
        {
            "titre": "Modifier une grande unité - Sciences de Gestion",
            "filiere": filiere,
            "grande_unite": grande_unite,
            "niveau": request.GET.get("niveau"),
            "semestre": request.GET.get("semestre"),
        }
    )
    
def l3_gestion_grande_unite_delete(request, pk):

    filiere = get_object_or_404(
        FiliereLMD,
        libelle="Sciences de Gestion"
    )

    grande_unite = get_object_or_404(
        GrandeUnite,
        pk=pk,
        filiere=filiere
    )

    niveau = request.GET.get("niveau")
    semestre = request.GET.get("semestre")

    if grande_unite.ues.exists():

        messages.error(
            request,
            "Impossible de supprimer : des UE sont encore "
            "rattachées à cette grande unité."
        )

    else:

        grande_unite.delete()

        messages.success(
            request,
            "Grande unité supprimée avec succès."
        )

    return redirect(
        f"{reverse('l3_gestion_grande_unite')}"
        f"?niveau={niveau}&semestre={semestre}"
    )
    
    
def sciences_gestion_etudiant_import(request):

    if request.method == "POST":

        fichier = request.FILES.get("fichier")

        if not fichier:
            messages.error(
                request,
                "Veuillez sélectionner un fichier Excel."
            )

            return redirect("l3_sciences_gestion_etudiants")


        # Vérification extension
        if not fichier.name.lower().endswith((".xlsx", ".xls")):

            messages.error(
                request,
                "Veuillez sélectionner un fichier Excel valide (.xlsx ou .xls)."
            )

            return redirect("l3_sciences_gestion_etudiants")


        try:

            # Lecture Excel
            df = pd.read_excel(fichier)

            # Nettoyage des noms de colonnes
            df.columns = (
                df.columns
                .astype(str)
                .str.strip()
                .str.lower()
            )


            # Colonnes obligatoires
            colonnes_obligatoires = [
                "matricule",
                "nom",
                "prenoms",
                "niveau",
            ]


            colonnes_manquantes = [
                colonne
                for colonne in colonnes_obligatoires
                if colonne not in df.columns
            ]


            if colonnes_manquantes:

                messages.error(
                    request,
                    "Colonnes manquantes : "
                    + ", ".join(colonnes_manquantes)
                )

                return redirect(
                    "sciences_gestion_etudiant_import"
                )


            # Recherche de la filière
            filiere = FiliereLMD.objects.filter(
                libelle__iexact="Sciences de Gestion"
            ).first()


            if not filiere:

                messages.error(
                    request,
                    "La filière Sciences de Gestion "
                    "n'existe pas dans la base de données."
                )

                return redirect(
                    "sciences_gestion_etudiant_import"
                )


            importes = 0
            ignores = 0


            for _, ligne in df.iterrows():

                matricule = str(
                    ligne.get("matricule", "")
                ).strip()

                nom = str(
                    ligne.get("nom", "")
                ).strip()

                prenoms = str(
                    ligne.get("prenoms", "")
                ).strip()

                niveau = str(
                    ligne.get("niveau", "")
                ).strip().upper()


                # Ignorer les lignes vides
                if not matricule or not nom or not prenoms:

                    ignores += 1
                    continue


                # Vérification niveau
                if niveau not in ["L1", "L2", "L3"]:

                    ignores += 1
                    continue


                # Vérifier si l'étudiant existe déjà
                existe = EtudiantLMD.objects.filter(
                    matricule=matricule
                ).exists()


                if existe:

                    ignores += 1
                    continue


                # Création étudiant
                EtudiantLMD.objects.create(

                    matricule=matricule,

                    nom=nom,

                    prenoms=prenoms,

                    niveau=niveau,

                    filiere=filiere,

                )

                importes += 1


            messages.success(
                request,
                f"{importes} étudiant(s) importé(s) avec succès."
            )


            if ignores:

                messages.warning(
                    request,
                    f"{ignores} ligne(s) ignorée(s)."
                )


            return redirect(
                "sciences_gestion_etudiants"
            )


        except Exception as e:

            messages.error(
                request,
                f"Erreur lors de l'importation : {str(e)}"
            )

            return redirect(
                "sciences_gestion_etudiant_import"
            )


    return render(
        request,
            "lmd/l3/gestion/etudiant_import.html"
    )

def l3_droit_dashboard(request):
    """
    Tableau de bord Droit Privé
    Tous les niveaux : L1, L2 et L3
    """

    total_etudiants = Etudiant.objects.filter(
        filiere__nom__iexact="Droit Privé"
    ).count()

    total_ue = UE.objects.filter(
        filiere__nom__iexact="Droit Privé"
    ).count()

    total_notes = NoteLMD.objects.filter(
        etudiant__filiere__nom__iexact="Droit Privé"
    ).count()

    total_bulletins = total_etudiants

    context = {
        "total_etudiants": total_etudiants,
        "total_ue": total_ue,
        "total_notes": total_notes,
        "total_bulletins": total_bulletins,
    }

    return render(
        request,
        "lmd/l3/droit/dashboard.html",
        context
    )


def get_filiere_qhse():
    """Récupère la filière QHSE de façon fiable.

    Centralise la correction du bug répété dans plusieurs vues :
    - le libellé exact en base n'est ni "Management QHSE" ni
      "Qualité, Hygiène, Sécurité" seuls, mais "Management de la
      Qualité, Hygiène, Sécurité et Environnement" ;
    - il existe actuellement DEUX lignes FiliereLMD avec ce même
      libellé (doublon en base à nettoyer). `get_object_or_404`
      planterait avec une erreur 500 (MultipleObjectsReturned) tant
      que ce doublon existe, donc on utilise filter().first().

    TODO : une fois le doublon nettoyé en base (voir discussion),
    cette fonction pourra revenir à un simple
    get_object_or_404(FiliereLMD, libelle=...).
    """
    filiere = (
        FiliereLMD.objects
        .filter(libelle__icontains="Qualité, Hygiène, Sécurité")
        .order_by("id")
        .first()
    )
    if filiere is None:
        raise Http404("Filière QHSE introuvable")
    return filiere

def l3_qhse_grande_unite(request):

    filiere = get_filiere_qhse()

    niveau = request.GET.get("niveau")
    semestre = request.GET.get("semestre")

    grandes_unites = GrandeUnite.objects.filter(
        filiere=filiere
    )

    if niveau:
        grandes_unites = grandes_unites.filter(
            niveau=niveau
        )

    if semestre:
        grandes_unites = grandes_unites.filter(
            semestre=semestre
        )

    return render(
        request,
        "lmd/l3_qhse/grande_unite_list.html",
        {
            "filiere": filiere,
            "grandes_unites": grandes_unites,
            "niveau": niveau,
            "semestre": semestre,
        }
    )


def l3_qhse_grande_unite_add(request):

    filiere = get_filiere_qhse()

    if request.method == "POST":
        nom = request.POST.get("nom")
        code = request.POST.get("code")
        niveau = request.POST.get("niveau")
        semestre = request.POST.get("semestre")

        GrandeUnite.objects.create(
            nom=nom,
            code=code,
            ordre=request.POST.get("ordre") or 1,
            filiere=filiere,
            niveau=niveau,
            semestre=semestre,
        )

        messages.success(
            request,
            "Grande unité QHSE ajoutée avec succès."
        )

        return redirect(
            f"{reverse('l3_qhse_grande_unite')}"
            f"?niveau={niveau}&semestre={semestre}"
        )

    return render(
        request,
        "lmd/l3_qhse/grande_unite_form.html",
        {
            "titre": "Ajouter une grande unité - Management QHSE",
            "filiere": filiere,
            "niveau": request.GET.get("niveau"),
            "semestre": request.GET.get("semestre"),
        }
    )


def l3_qhse_grande_unite_edit(request, pk):

    filiere = get_filiere_qhse()

    grande_unite = get_object_or_404(
        GrandeUnite,
        pk=pk,
        filiere=filiere
    )

    if request.method == "POST":

        niveau = request.POST.get("niveau")
        semestre = request.POST.get("semestre")

        grande_unite.code = request.POST.get("code")
        grande_unite.nom = request.POST.get("nom")
        grande_unite.ordre = request.POST.get("ordre") or 1
        grande_unite.niveau = niveau
        grande_unite.semestre = semestre
        grande_unite.save()

        messages.success(
            request,
            "Grande unité QHSE modifiée avec succès."
        )

        return redirect(
            f"{reverse('l3_qhse_grande_unite')}?niveau={niveau}&semestre={semestre}"
        )

    return render(
        request,
        "lmd/l3_qhse/grande_unite_form.html",
        {
            "titre": "Modifier une grande unité - Management QHSE",
            "filiere": filiere,
            "grande_unite": grande_unite,
            "niveau": request.GET.get("niveau"),
            "semestre": request.GET.get("semestre"),
        }
    )


def l3_qhse_grande_unite_delete(request, pk):

    filiere = get_filiere_qhse()

    grande_unite = get_object_or_404(
        GrandeUnite,
        pk=pk,
        filiere=filiere
    )

    niveau = request.GET.get("niveau")
    semestre = request.GET.get("semestre")

    if grande_unite.ues.exists():
        messages.error(
            request,
            "Impossible de supprimer : des UE sont encore rattachées à cette grande unité."
        )
    else:
        grande_unite.delete()
        messages.success(
            request,
            "Grande unité QHSE supprimée avec succès."
        )

    return redirect(
        f"{reverse('l3_qhse_grande_unite')}?niveau={niveau}&semestre={semestre}"
    )


def l3_qhse_etudiant_import(request):

    filiere = get_filiere_qhse()

    colonnes_attendues = [
        "Matricule",
        "Nom",
        "Prénoms",
        "Lieu naissance",
        "Date naissance",
        "Email",
        "Annee academique",
        "Niveau",
        "Sexe",
        "Téléphone",
    ]

    if request.method == "POST":

        fichier = request.FILES.get("fichier")

        if not fichier:
            messages.error(
                request,
                "❌ Aucun fichier sélectionné."
            )

            return redirect(
                "l3_qhse_etudiants"
            )

        # Vérification extension
        if not fichier.name.lower().endswith(".xlsx"):

            messages.error(
                request,
                "❌ Format incorrect. Veuillez importer un fichier Excel (.xlsx)."
            )

            return redirect(
                "l3_qhse_etudiants"
            )

        try:

            workbook = load_workbook(
                fichier
            )

            sheet = workbook.active

        except Exception:

            messages.error(
                request,
                "❌ Impossible de lire le fichier Excel."
            )

            return redirect(
                "l3_qhse_etudiants"
            )

        # Vérification des entêtes
        headers = [
            str(cell.value).strip()
            for cell in sheet[1]
            if cell.value is not None
        ]

        if headers != colonnes_attendues:

            messages.error(
                request,
                "❌ Format du fichier incorrect. "
                "Veuillez télécharger et utiliser le modèle Excel fourni."
            )

            return redirect(
                "l3_qhse_etudiants"
            )

        total = 0
        doublons = 0
        erreurs = 0

        for index, row in enumerate(
            sheet.iter_rows(
                min_row=2,
                values_only=True
            ),
            start=2
        ):

            # Supprimer les cellules vides à la fin
            row = list(row)

            while row and row[-1] is None:
                row.pop()

            # Ignorer les lignes complètement vides
            if not row:
                continue

            # Vérifier le nombre de colonnes
            if len(row) != 10:

                messages.warning(
                    request,
                    f"⚠️ Ligne {index} ignorée : "
                    f"{len(row)} colonnes trouvées au lieu de 10."
                )

                erreurs += 1
                continue

            (
                matricule,
                nom,
                prenoms,
                lieu_naissance,
                date_naissance,
                email,
                annee_academique,
                niveau,
                sexe,
                telephone
            ) = row

            # Champs obligatoires
            if not matricule or not nom or not prenoms:

                messages.warning(
                    request,
                    f"⚠️ Ligne {index} ignorée : "
                    "Matricule, nom ou prénom manquant."
                )

                erreurs += 1
                continue

            # Vérification doublon matricule
            if EtudiantLMD.objects.filter(
                matricule=matricule
            ).exists():

                doublons += 1
                continue

            # Conversion de la date Excel
            date_naissance = convertir_date(
                date_naissance
            )

            # Création de l'étudiant QHSE
            EtudiantLMD.objects.create(

                matricule=matricule,

                nom=nom,

                prenoms=prenoms,

                lieu_naissance=lieu_naissance,

                date_naissance=date_naissance,

                email=email,

                annee_academique=annee_academique,

                niveau=niveau,

                sexe=sexe,

                telephone=telephone,

                filiere=filiere
            )

            total += 1

        # Messages de résultat
        messages.success(
            request,
            f"✅ {total} étudiant(s) QHSE importé(s) avec succès."
        )

        if doublons:

            messages.warning(
                request,
                f"⚠️ {doublons} étudiant(s) déjà existant(s) ignoré(s)."
            )

        if erreurs:

            messages.warning(
                request,
                f"⚠️ {erreurs} ligne(s) non importée(s)."
            )

        return redirect(
            "l3_qhse_etudiants"
        )

    return render(
        request,
        "lmd/l3_qhse/import_etudiants.html"
    )




def l3_qhse_etudiant_modele_excel(request):

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Etudiants QHSE"

    # ==============================
    # EN-TÊTES
    # ==============================

    colonnes = [
        "Matricule",
        "Nom",
        "Prénoms",
        "Lieu naissance",
        "Date naissance",
        "Email",
        "Annee academique",
        "Niveau",
        "Sexe",
        "Téléphone",
    ]

    sheet.append(colonnes)

    # ==============================
    # LARGEUR DES COLONNES
    # ==============================

    largeurs = {
        "A": 18,
        "B": 25,
        "C": 30,
        "D": 25,
        "E": 18,
        "F": 35,
        "G": 20,
        "H": 12,
        "I": 12,
        "J": 20,
    }

    for colonne, largeur in largeurs.items():
        sheet.column_dimensions[colonne].width = largeur

    # ==============================
    # STYLE DES EN-TÊTES
    # ==============================

    from openpyxl.styles import Font, PatternFill, Alignment

    for cell in sheet[1]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="198754"
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # ==============================
    # RÉPONSE EXCEL
    # ==============================

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="modele_etudiants_qhse.xlsx"'
    )

    workbook.save(response)

    return response