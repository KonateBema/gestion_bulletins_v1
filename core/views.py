# core/views.py

from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, FileResponse
from django.views.decorators.csrf import csrf_protect
from .forms import EtudiantForm,ClasseForm,MatiereForm,AffectationForm,NoteForm
from django.shortcuts import get_object_or_404
from django.core.paginator import Paginator
from django.db.models import Q
from .services import calcul_moyenne_etudiant
from openpyxl import load_workbook
from .models import Filierebts
from django.contrib import messages
from .models import Salle
from .models import SaisieNotesBTS
from .models import ( Classe,Niveau)
from lmd.models import EtudiantLMD,FiliereLMD
# from .models import NoteBTS
from .models import (Etudiant, Professeur, Matiere, Note,AffectationMatiere, Inscription, Profile)
from .forms import UserRegisterForm
from .utils import generate_matricule
from .services import (mention,)
from .pdf_service import generate_bulletin_pdf
from datetime import datetime
from core.decorators import role_required
from django.http import JsonResponse

from django.views.decorators.http import require_http_methods
from django.template.loader import render_to_string
from .models import GrandeUnite
from .forms import GrandeUniteForm
from django.http import FileResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
# =========================
# 🔐 LOGIN
# =========================
@csrf_protect
def login_viewAAAA(request):

    if request.method == "POST":

        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)

            profile = Profile.objects.filter(user=user).first()

            if not profile:
                return render(request, "login.html", {
                    "error": "Profil utilisateur introuvable"
                })

            if profile.role == "ADMIN":
                return redirect("dashboard_admin")

            elif profile.role == "PROF":
                return redirect("dashboard_prof")

            else:
                return redirect("dashboard_etudiant")

        return render(request, "login.html", {
            "error": "Identifiants incorrects"
        })

    return render(request, "login.html")

from django.views.decorators.csrf import csrf_protect


# ==========================================================
# 🔐 LOGIN
# ==========================================================

@csrf_protect
def login_view(request):

    # ==========================================================
    # UTILISATEUR DÉJÀ CONNECTÉ
    # ==========================================================

    if request.user.is_authenticated:

        profile = Profile.objects.filter(
            user=request.user
        ).first()

        if not profile:

            logout(request)

            return render(
                request,
                "login.html",
                {
                    "error": "Profil utilisateur introuvable."
                }
            )

        # Redirection selon le rôle
        if profile.role == "ADMIN":

            return redirect("dashboard_admin")

        elif profile.role == "GESTIONNAIRE":
            print("ROLE GESTIONNAIRE DETECTE")
            print("Redirection vers dashboard_gestionnaire")

            return redirect("dashboard_gestionnaire")

        elif profile.role == "PROF":

            return redirect("dashboard_prof")

        elif profile.role == "ETUDIANT":

            return redirect("dashboard_etudiant")

        else:

            logout(request)

            return render(
                request,
                "login.html",
                {
                    "error": "Rôle utilisateur non reconnu."
                }
            )

    # ==========================================================
    # FORMULAIRE DE CONNEXION
    # ==========================================================

    if request.method == "POST":

        username = request.POST.get(
            "username",
            ""
        ).strip()

        password = request.POST.get(
            "password",
            ""
        )

        # Vérification des champs
        if not username or not password:

            return render(
                request,
                "login.html",
                {
                    "error":
                    "Veuillez renseigner votre identifiant "
                    "et votre mot de passe."
                }
            )

        # ======================================================
        # AUTHENTIFICATION DJANGO
        # ======================================================

        user = authenticate(
            request,
            username=username,
            password=password
        )

        if user is None:

            return render(
                request,
                "login.html",
                {
                    "error":
                    "Identifiant ou mot de passe incorrect."
                }
            )

        # ======================================================
        # CONNEXION
        # ======================================================

        login(
            request,
            user
        )

        # ======================================================
        # RÉCUPÉRATION DU PROFIL
        # ======================================================

        profile = Profile.objects.filter(
            user=user
        ).first()

        # Aucun profil
        if not profile:

            logout(request)

            return render(
                request,
                "login.html",
                {
                    "error":
                    "Profil utilisateur introuvable."
                }
            )

        # ======================================================
        # REDIRECTION SELON LE RÔLE
        # ======================================================

        if profile.role == "ADMIN":

            return redirect(
                "dashboard_admin"
            )

        elif profile.role == "GESTIONNAIRE":

            return redirect(
                "dashboard_gestionnaire"
            )

        elif profile.role == "PROF":

            return redirect(
                "dashboard_prof"
            )

        elif profile.role == "ETUDIANT":

            return redirect(
                "dashboard_etudiant"
            )

        # ======================================================
        # RÔLE INCONNU
        # ======================================================

        logout(request)

        return render(
            request,
            "login.html",
            {
                "error":
                "Rôle utilisateur non reconnu."
            }
        )

    # ==========================================================
    # AFFICHAGE PAGE LOGIN
    # ==========================================================

    return render(
        request,
        "login.html"
    )


# =========================
# 🚪 LOGOUT
# =========================
def logout_view(request):
    logout(request)
    return redirect('login')

@login_required
def dashboard(request):

    filieres_l3 = FiliereLMD.objects.filter(
        niveau_formation="L3"
    )

    filieres_master = FiliereLMD.objects.filter(
        niveau_formation="M1-M2"
    )

    context = {

        # =====================
        # BTS
        # =====================
        "etudiants_count": Etudiant.objects.count(),
        "professeurs_count": Professeur.objects.count(),
        "classes_count": Classe.objects.count(),
        "matieres_count": Matiere.objects.count(),
        "notes_count": Note.objects.count(),


        # =====================
        # LMD
        # =====================
        "l1_count": EtudiantLMD.objects.filter(
            niveau="L1"
        ).count(),

        "l2_count": EtudiantLMD.objects.filter(
            niveau="L2"
        ).count(),

        "l3_count": EtudiantLMD.objects.filter(
            niveau="L3"
        ).count(),

        "master_count": EtudiantLMD.objects.filter(
            niveau__in=["M1", "M2"]
        ).count(),


        # =====================
        # MENU DYNAMIQUE
        # =====================
        "filieres_l3": filieres_l3,

        "filieres_master": filieres_master,

    }


    return render(
        request,
        "dashboard.html",
        context
    )

# =========================
# 🧑‍💼 ADMIN
# =========================
@login_required
def dashboard_admin(request):

    return render(request, "admin_dashboard.html", {
        "etudiants": Etudiant.objects.count(),
        "professeurs": Professeur.objects.count(),
        "classes": Classe.objects.count(),
    })


# =========================
# 👨‍🏫 PROF
# =========================
@login_required
def dashboard_prof(request):

    prof = Professeur.objects.filter(user=request.user).first()

    if not prof:
        return HttpResponse("❌ Profil professeur introuvable")

    matieres = AffectationMatiere.objects.filter(professeur=prof)

    return render(request, "prof_dashboard.html", {
        "matieres": matieres,
    })


# =========================
# 🎓 ETUDIANT
# =========================
@login_required
def dashboard_etudiant(request):

    etudiant = Etudiant.objects.filter(user=request.user).first()

    if not etudiant:
        return HttpResponse("❌ Aucun profil étudiant trouvé")

    notes = Note.objects.filter(etudiant=etudiant)

    return render(request, "etudiant_dashboard.html", {
        "etudiant": etudiant,
        "notes": notes,
    })


# ==========================================================
# 👨‍💼 DASHBOARD GESTIONNAIRE
# ==========================================================

@login_required(login_url="login")
@role_required("GESTIONNAIRE")
def dashboard_gestionnaire(request):
    print("=== DASHBOARD GESTIONNAIRE ===")
    print("Etudiants :", Etudiant.objects.count())
    print("Classes :", Classe.objects.count())
    print("Filieres :", Filierebts.objects.count())
    print("Matieres :", Matiere.objects.count())

    return render(
        request,
        "gestionnaire_dashboard.html",
        {
            "etudiants": Etudiant.objects.count(),
            "classes": Classe.objects.count(),
            "filieres": Filierebts.objects.count(),
            "matieres": Matiere.objects.count(),
        }
    )



# =========================
# 📝 INSCRIPTION UTILISATEUR
# =========================
def register_user(request):

    if request.method == "POST":

        form = UserRegisterForm(request.POST)

        if form.is_valid():

            user = form.save()

            # 🎓 ETUDIANT
            if user.role == "ETUD":

                etudiant = Etudiant.objects.create(
                    user=user,
                    matricule=generate_matricule("ETU"),
                    date_naissance="2000-01-01",
                    sexe="M",
                    telephone="00000000",
                    classe=Classe.objects.first()
                )

                Inscription.objects.create(
                    etudiant=etudiant,
                    classe=etudiant.classe,
                    annee="2025-2026"
                )

            # 👨‍🏫 PROF
            elif user.role == "PROF":

                Professeur.objects.create(
                    user=user,
                    matricule=generate_matricule("PROF"),
                    specialite="Non définie",
                    telephone="00000000"
                )

            return redirect('login')

    else:
        form = UserRegisterForm()

    return render(request, 'register.html', {'form': form})

# =========================
# 📊 BULLETIN ETUDIANT
# =========================
@role_required("ADMIN")
@login_required
def bulletin_etudiant(request):

    etudiant = Etudiant.objects.first()  # ou filtre propre

    if not etudiant:
        return HttpResponse("Aucun étudiant trouvé")

    moyenne = calcul_moyenne_etudiant(etudiant)

    return render(request, "bulletin.html", {
        "etudiant": etudiant,
        "moyenne": moyenne,
        "mention": mention(moyenne),
    })


# =========================
# 📄 PDF BULLETIN
# =========================
@role_required("ADMIN", "GESTIONNAIRE")
@login_required(login_url="login")
def etudiant_listPRO(request):

    query = request.GET.get("q", "")
    classe_id = request.GET.get("classe", "")
    filiere_bts_id = request.GET.get("filiere_bts", "")
    niveau = request.GET.get("niveau", "")


    # =========================
    # LISTE ETUDIANTS
    # =========================

    etudiants = Etudiant.objects.select_related(
        "classe",
        "filiere_bts"
    ).order_by("nom", "prenoms")



    # =========================
    # RECHERCHE
    # =========================

    if query:

        etudiants = etudiants.filter(

            Q(matricule__icontains=query) |

            Q(nom__icontains=query) |

            Q(prenoms__icontains=query)

        )



    # =========================
    # FILTRE CLASSE
    # =========================

    if classe_id:

        etudiants = etudiants.filter(
            classe_id=classe_id
        )



    # =========================
    # FILTRE NIVEAU BTS
    # =========================

    if niveau:

        etudiants = etudiants.filter(

            filiere_bts__niveaux__nom=niveau

        ).distinct()



    # =========================
    # FILTRE FILIERE BTS
    # =========================

    if filiere_bts_id:

        etudiants = etudiants.filter(

            filiere_bts_id=filiere_bts_id

        )



    # =========================
    # PAGINATION
    # =========================

    paginator = Paginator(
        etudiants,
        10
    )


    page_number = request.GET.get("page")


    page_obj = paginator.get_page(
        page_number
    )



    return render(
        request,
        "etudiants/list.html",
        {

            "page_obj": page_obj,


            # données filtres

            "classes": Classe.objects.all().order_by("nom"),


            "filieres_bts": Filierebts.objects.all().order_by("nom"),


            "niveau": niveau,


            "classe_selected": classe_id,


            "filiere_selected": filiere_bts_id,


            "query": query,

        }
    )

def etudiant_list(request):

    etudiants = Etudiant.objects.select_related(
        "filiere_bts",
        "classe"
    ).all()

    q = request.GET.get("q", "").strip()
    niveau = request.GET.get("niveau", "").strip()
    filiere_bts = request.GET.get("filiere_bts", "").strip()

    # Recherche
    if q:
        etudiants = etudiants.filter(
            Q(nom__icontains=q) |
            Q(prenoms__icontains=q) |
            Q(matricule__icontains=q)
        )

    # Filtre BTS 1 / BTS 2
    if niveau:
        etudiants = etudiants.filter(
            classe__niveau__nom=niveau
        )

    # Filtre filière
    if filiere_bts:
        etudiants = etudiants.filter(
            filiere_bts_id=filiere_bts
        )

    filieres_bts = Filierebts.objects.all().order_by("nom")

    paginator = Paginator(etudiants, 20)

    page_number = request.GET.get("page")

    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "filieres_bts": filieres_bts,
    }

    return render(
        request,
        "etudiants/list.html",
        context
    )


def etudiant_create(request):

    if request.method == "POST":

        form = EtudiantForm(request.POST)

        if form.is_valid():
            form.save()
            return redirect('etudiant_list')

    else:
        form = EtudiantForm()

    filieres_bts = Filierebts.objects.all()

    return render(request, 'etudiants/form.html', {
        'form': form,
        'filieres_bts': filieres_bts
    })


def etudiants_par_salle(request):

    salles = Salle.objects.prefetch_related(
        'classe_set__etudiants'
    )

    return render(
        request,
        'etudiants/par_salle.html',
        {
            'salles': salles
        }
    )

def etudiant_update(request, id):

    etudiant = get_object_or_404(Etudiant, id=id)

    if request.method == "POST":

        form = EtudiantForm(request.POST, instance=etudiant)

        if form.is_valid():
            form.save()
            return redirect('etudiant_list')

    else:

        form = EtudiantForm(instance=etudiant)

    return render(request, 'etudiants/form.html', {
        'form': form
    })


def etudiant_delete(request, id):
    Etudiant.objects.get(id=id).delete()
    return redirect("etudiant_list")

def classe_list(request):

    # ==========================
    # CREATION CLASSE
    # ==========================
    if request.method == "POST":

        form = ClasseForm(request.POST)

        if form.is_valid():
            form.save()
            return redirect("classe_list")

        else:
            print(form.errors)

    else:
        form = ClasseForm()



    # ==========================
    # FILTRES
    # ==========================
    query = request.GET.get("q", "")
    filiere_bts = request.GET.get("filiere_bts", "")
    niveau = request.GET.get("niveau", "")



    # ==========================
    # LISTE DES CLASSES
    # ==========================
    classes = Classe.objects.select_related(
        "filiere_bts",
        "niveau",
        "salle"
    ).order_by("-id")



    # Recherche par nom
    if query:
        classes = classes.filter(
            nom__icontains=query
        )


    # Filtre filière BTS
    if filiere_bts:
        classes = classes.filter(
            filiere_bts_id=filiere_bts
        )


    # Filtre niveau
    if niveau:
        classes = classes.filter(
            niveau_id=niveau
        )



    # ==========================
    # PAGINATION
    # ==========================
    paginator = Paginator(
        classes,
        10
    )

    page_number = request.GET.get(
        "page"
    )

    page_obj = paginator.get_page(
        page_number
    )



    # ==========================
    # DONNEES POUR SELECTS
    # ==========================
    filieres = Filierebts.objects.all()

    niveaux = Niveau.objects.all()  # noqa: F821



    return render(
        request,
        "classes/list.html",
        {
            "page_obj": page_obj,
            "form": form,
            "filieres": filieres,
            "niveaux": niveaux,
            "query": query,
            "filiere_selected": filiere_bts,
            "niveau_selected": niveau,
        }
    )


def classe_create(request):
    if request.method == "POST":
        form = ClasseForm(request.POST)

        if form.is_valid():
            form.save()
            return redirect('classe_list')
    else:
        form = ClasseForm()

    return render(request, 'classes/form.html', {
        'form': form
    })


def matiere_listAAA(request):

    filieres = Filierebts.objects.prefetch_related(
        "matiere_set"
    ).order_by("nom")


    q = request.GET.get("q")


    if q:

        filieres = filieres.filter(
            matiere__libelle__icontains=q
        ).distinct()



    context = {

        "filieres": filieres,

        "total_matieres": Matiere.objects.count(),

    }


    return render(
        request,
        "matieres/list.html",
        context
    )

from django.core.paginator import Paginator
from django.db.models import Q

def matiere_list(request):

    q = request.GET.get("q", "")
    filiere_id = request.GET.get("filiere_bts", "")

    matieres = Matiere.objects.select_related(
        "filiere_bts"
    ).order_by(
        "filiere_bts__nom",
        "libelle"
    )

    if q:
        matieres = matieres.filter(
            Q(libelle__icontains=q) |
            Q(code__icontains=q)
        )

    if filiere_id:
        matieres = matieres.filter(
            filiere_bts_id=filiere_id
        )

    paginator = Paginator(matieres, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "total_matieres": matieres.count(),
        "filiere_list": Filierebts.objects.order_by("nom"),
        "filiere_selectionnee": filiere_id,
    }

    return render(
        request,
        "matieres/list.html",
        context, 
    )


def matiere_createAAA(request):
    form = MatiereForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect("matiere_list")
    return render(request, "matieres/form.html", {"form": form})

def matiere_create(request):
    form = MatiereForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Matière ajoutée avec succès.")
        return redirect('matiere_list')

    return render(request, 'matieres/form.html', {
        'form': form,
        'is_edit': False,
    })

def affectation_list(request):
    return render(request, "affectations/list.html", {
        "affectations": Affectation.objects.select_related(  # noqa: F821
            "professeur", "matiere", "classe"
        )
    })


def affectation_create(request):
    form = AffectationForm(request.POST or None)

    if form.is_valid():
        form.save()
        return redirect("affectation_list")

    return render(request, "affectations/form.html", {
        "form": form,
        "title": "Affecter professeur"
    })


def affectation_delete(request, id):
    Affectation.objects.get(id=id).delete()  # noqa: F821
    return redirect("affectation_list")



def note_list(request):

    notes = Note.objects.select_related(
        "etudiant",
        "matiere"
    ).all().order_by("-id")

    etudiant_id = request.GET.get("etudiant")
    matiere_id = request.GET.get("matiere")
    semestre = request.GET.get("semestre")

    if etudiant_id:
        notes = notes.filter(etudiant_id=etudiant_id)

    if matiere_id:
        notes = notes.filter(matiere_id=matiere_id)

    if semestre:
        notes = notes.filter(semestre=semestre)

    paginator = Paginator(notes, 10)  # 10 lignes par page

    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "notes/list.html", {
        "notes": page_obj,
        "page_obj": page_obj,
        "etudiants": Etudiant.objects.all(),
        "matieres": Matiere.objects.all(),
    })

def note_create(request):

    form = NoteForm(request.POST or None)

    if request.method == "POST":

        if form.is_valid():
            note = form.save(commit=False)

            # DEBUG OPTIONNEL
            print("✔ Note enregistrée")

            note.save()
            return redirect("note_list")

        else:
            print(form.errors)

    return render(request, "notes/form.html", {
        "form": form,
        "title": "Ajouter note"
    })
    
def note_update(request, id):
    note = Note.objects.get(id=id)
    form = NoteForm(request.POST or None, instance=note)

    if form.is_valid():
        form.save()
        return redirect("note_list")

    return render(request, "notes/form.html", {
        "form": form,
        "title": "Modifier note"
    })


def note_delete(request, id):
    Note.objects.get(id=id).delete()
    return redirect("note_list")

def moyenne_etudiant(etudiant):
    notes = Note.objects.filter(etudiant=etudiant)

    if not notes:
        return 0

    total = sum(n.moyenne for n in notes)
    return total / notes.count()


def classe_edit(request, pk):
    classe = Classe.objects.get(pk=pk)
    form = ClasseForm(request.POST or None, instance=classe)

    if request.method == "POST":
        if form.is_valid():
            form.save()
            return redirect('classe_list')

    return render(request, 'classes/form.html', {
        'form': form
    })

def classe_delete(request, pk):
    classe = get_object_or_404(Classe, pk=pk)
    classe.delete()
    return redirect('classe_list')

def matiere_updateAAA(request, id):
    matiere = get_object_or_404(Matiere, id=id)
    form = MatiereForm(request.POST or None, instance=matiere)

    if form.is_valid():
        form.save()
        return redirect('matiere_list')

    return render(request, 'matieres/form.html', {'form': form})

def matiere_delete(request, id):
    matiere = get_object_or_404(Matiere, id=id)
    matiere.delete()
    return redirect('matiere_list')

def matiere_update(request, id):
    matiere = get_object_or_404(Matiere, id=id)
    form = MatiereForm(request.POST or None, instance=matiere)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Matière modifiée avec succès.")
        return redirect('matiere_list')

    return render(request, 'matieres/form.html', {
        'form': form,
        'matiere': matiere,
        'is_edit': True,
    })

def download_bulletin_pdf(request, etudiant_id, classe_id, semestre):

    etudiant = get_object_or_404(
        Etudiant,
        id=etudiant_id
    )

    classe = get_object_or_404(
        Classe,
        id=classe_id
    )

    file_path = generate_bulletin_pdf(
        etudiant=etudiant,
        classe=classe,
        semestre=semestre
    )

    return FileResponse(
        open(file_path, "rb"),
        as_attachment=True,
        filename=f"bulletin_S{semestre}_{etudiant.matricule}.pdf"
    )

def bulletin_classe(request, classe_id):

    classe = Classe.objects.get(id=classe_id)

    data = classement(classe)  # noqa: F821

    return render(request, "bulletin_classe.html", {
        "classe": classe,
        "data": data
    })

@login_required(login_url="login")
def bulletin_list(request):

    etudiants = Etudiant.objects.select_related("classe").all()

    # 🔎 Filtres GET
    matricule = request.GET.get("matricule")
    telephone = request.GET.get("telephone")
    filiere = request.GET.get("filiere")
    classe = request.GET.get("classe")

    # 🔽 Filtrage dynamique
    if matricule:
        etudiants = etudiants.filter(matricule__icontains=matricule)

    if telephone:
        etudiants = etudiants.filter(telephone__icontains=telephone)

    if filiere:
        etudiants = etudiants.filter(filiere__icontains=filiere)

    if classe:
        etudiants = etudiants.filter(classe_id=classe)

    # 📄 PAGINATION
    paginator = Paginator(etudiants, 10)  # 10 étudiants par page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "bulletins/list.html", {
        "etudiants": page_obj,
        "page_obj": page_obj,
        "classes": Classe.objects.all(),
    })
 


def liste_filieres_bts(request):
    filieres = Filierebts.objects.all().order_by('nom')

    return render(request, 'bts/liste_filieres_bts.html', {
        'filieres': filieres
    })

def ajouter_filiere_btsGGG(request):
    if request.method == "POST":
        nom = request.POST.get("nom")

        Filierebts.objects.create(
            nom=nom
        )

        messages.success(request, "Filière BTS ajoutée avec succès.")
        return redirect('liste_filieres_bts')

    return render(request, 'bts/ajouter_filiere_bts.html')


def ajouter_filiere_bts(request):

    if request.method == "POST":

        nom = request.POST.get("nom")
        niveaux_ids = request.POST.getlist("niveaux")

        # Création de la filière
        filiere = Filierebts.objects.create(
            nom=nom
        )

        # Association des niveaux sélectionnés
        filiere.niveaux.set(niveaux_ids)

        messages.success(
            request,
            "Filière BTS ajoutée avec succès."
        )

        return redirect("liste_filieres_bts")

    niveaux = Niveau.objects.all()

    return render(
        request,
        "bts/ajouter_filiere_bts.html",
        {
            "niveaux": niveaux
        }
    )

def modifier_filiere_bts(request, pk):

    filiere = get_object_or_404(Filierebts, pk=pk)
    niveau = filiere.niveaux.first()

    if request.method == "POST":

        filiere.nom = request.POST.get("nom")
        filiere.save()

        nom_niveau = request.POST.get("niveau")

        niveau, _ = Niveau.objects.get_or_create(
            nom=nom_niveau
        )

        filiere.niveaux.set([niveau])

        messages.success(request, "Filière modifiée avec succès.")
        return redirect("liste_filieres_bts")

    return render(
        request,
        "bts/modifier_filiere_bts.html",
        {
            "filiere": filiere,
            "niveau": niveau,
        }
    )


def supprimer_filiere_bts(request, pk):
    filiere = get_object_or_404(Filierebts, pk=pk)

    filiere.delete()

    messages.success(request, "Filière BTS supprimée.")
    return redirect('liste_filieres_bts')

def salle_list(request):
    salles = Salle.objects.all()
    return render(request, 'salles/salle_list.html', {
        'salles': salles
    })

def salle_create(request):
    if request.method == "POST":
        code = request.POST.get("code")
        nom = request.POST.get("nom")
        capacite = request.POST.get("capacite")

        Salle.objects.create(
            code=code,
            nom=nom,
            capacite=capacite
        )

        messages.success(request, "Salle ajoutée avec succès")
        return redirect('salle_list')

    return render(request, 'salles/salle_form.html')

def salle_edit(request, pk):
    salle = get_object_or_404(Salle, pk=pk)

    if request.method == "POST":
        salle.code = request.POST.get("code")
        salle.nom = request.POST.get("nom")
        salle.capacite = request.POST.get("capacite")
        salle.save()

        messages.success(request, "Salle modifiée avec succès")
        return redirect('salle_list')

    return render(request, 'salles/salle_form.html', {
        'salle': salle
    })

def salle_delete(request, pk):
    salle = get_object_or_404(Salle, pk=pk)
    salle.delete()

    messages.success(request, "Salle supprimée")
    return redirect('salle_list')


def saisie_note_groupee(request):

    classes = Classe.objects.select_related(
        "filiere_bts",
        "niveau",
        "salle"
    )

    matieres = Matiere.objects.all()

    etudiants = []
    notes_existantes = {}

    classe_id = request.GET.get("classe")
    matiere_id = request.GET.get("matiere")
    semestre = request.GET.get("semestre")


    # ==========================
    # ENREGISTREMENT DES NOTES
    # ==========================
    if request.method == "POST":

        classe_id = request.POST.get("classe")
        matiere_id = request.POST.get("matiere")
        semestre = request.POST.get("semestre")


        classe = Classe.objects.get(id=classe_id)

        etudiants = Etudiant.objects.filter(
            classe=classe
        )


        # créer ou récupérer une saisie
        saisie, created = SaisieNotesBTS.objects.get_or_create(
            classe=classe,
            matiere_id=matiere_id,
            semestre=semestre
        )


        for etudiant in etudiants:

            cc = float(
                request.POST.get(f"cc_{etudiant.id}") or 0
            )
            
            devoir = float(
               request.POST.get(f"devoir_{etudiant.id}") or 0
            )

            examen = float(
                request.POST.get(f"examen_{etudiant.id}") or 0
            )


            Note.objects.update_or_create(
                etudiant=etudiant,
                matiere_id=matiere_id,
                semestre=semestre,
                defaults={
                    "saisie": saisie,
                    "cc": cc,
                    "devoir": devoir,
                    "examen": examen,
                }
            )


        messages.success(
            request,
            "Les notes ont été enregistrées avec succès."
        )


        return redirect(
            f"{request.path}?classe={classe_id}&matiere={matiere_id}&semestre={semestre}"
        )



    # ==========================
    # AFFICHAGE
    # ==========================

    if classe_id and matiere_id and semestre:


        classe = Classe.objects.get(
            id=classe_id
        )


        etudiants = Etudiant.objects.filter(
            classe=classe
        )


        notes = Note.objects.filter(
            etudiant__in=etudiants,
            matiere_id=matiere_id,
            semestre=semestre
        )


        for note in notes:

            notes_existantes[note.etudiant_id] = note



        # envoyer la note directement dans le template
        for etudiant in etudiants:

            etudiant.note_existante = notes_existantes.get(
                etudiant.id
            )



    context = {

        "classes": classes,

        "matieres": matieres,

        "etudiants": etudiants,

        "notes_existantes": notes_existantes,

    }


    return render(
        request,
        "notes/saisie_groupee.html",
        context
    )


import unicodedata
    

import re



# ==============================================================
# NORMALISATION GÉNÉRALE
# ==============================================================

def normaliser_texte(texte):
    """
    Normalise un texte pour les comparaisons :

    - minuscules
    - suppression des accents
    - apostrophes uniformisées
    - espaces multiples supprimés
    """

    if texte is None:
        return ""

    texte = str(texte).strip().lower()

    # Uniformiser les apostrophes
    texte = texte.replace("’", "'")
    texte = texte.replace("`", "'")
    texte = texte.replace("´", "'")

    # Supprimer les accents
    texte = unicodedata.normalize(
        "NFD",
        texte
    )

    texte = "".join(
        caractere
        for caractere in texte
        if unicodedata.category(caractere) != "Mn"
    )

    # Remplacer les espaces multiples par un seul espace
    texte = " ".join(texte.split())

    return texte


# ==============================================================
# NORMALISATION FILIÈRE
# ==============================================================

def normaliser_filiere(texte):
    """
    Normalise une filière pour permettre les comparaisons
    même si le fichier Excel ne contient pas le code (IDA), (AD), etc.
    """

    texte = normaliser_texte(texte)

    # Supprimer les codes entre parenthèses
    # Exemple :
    # INFORMATIQUE ET DEVELOPPEMENT D'APPLICATIONS (IDA)
    # devient :
    # INFORMATIQUE ET DEVELOPPEMENT D'APPLICATIONS
    texte = re.sub(
        r"\s*\([^)]*\)",
        "",
        texte
    )

    # Uniformiser les apostrophes restantes
    texte = texte.replace("'", "")

    # Supprimer les espaces multiples
    texte = " ".join(texte.split())

    return texte


# ==============================================================
# IMPORT DES ÉTUDIANTS EXCEL
# ==============================================================

from datetime import datetime, date

# ----------------------------------------------------------
# Conversion flexible des dates
# ----------------------------------------------------------

def convertir_date(valeur):

    if valeur is None:
        return None

    # Déjà une date Python
    if isinstance(valeur, date) and not isinstance(valeur, datetime):
        return valeur

    # Date Excel (datetime)
    if isinstance(valeur, datetime):
        return valeur.date()

    # Nombre Excel (parfois 45000, etc.)
    if isinstance(valeur, (int, float)):
        try:
            return datetime.fromordinal(
                datetime(1899, 12, 30).toordinal() + int(valeur)
            ).date()
        except Exception:
            return None

    # Texte
    if isinstance(valeur, str):

        texte = valeur.strip()

        formats = [
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%d.%m.%Y",
            "%d %m %Y",
        ]

        for fmt in formats:
            try:
                return datetime.strptime(texte, fmt).date()
            except ValueError:
                continue

    return None


# ----------------------------------------------------------
# Importation des étudiants
# ----------------------------------------------------------

def import_etudiants_excel(request):

    if request.method != "POST":
        return render(request, "import_etudiants_excel.html")

    fichier = request.FILES.get("excel_file")

    if not fichier:
        messages.error(
            request,
            "Veuillez sélectionner un fichier Excel."
        )
        return redirect("import_etudiants_excel")

    if not fichier.name.lower().endswith(".xlsx"):
        messages.error(
            request,
            "Format incorrect. Veuillez importer uniquement un fichier .xlsx."
        )
        return redirect("import_etudiants_excel")

    try:
        wb = load_workbook(fichier, data_only=True)
    except Exception as e:
        messages.error(
            request,
            f"Impossible de lire le fichier Excel : {e}"
        )
        return redirect("import_etudiants_excel")

    ws = wb.active

    compteur_creation = 0
    compteur_modification = 0
    erreurs = []

    for ligne, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True),
        start=2
    ):

        try:

            if not row or all(
                value is None or str(value).strip() == ""
                for value in row
            ):
                continue

            matricule = str(row[0]).strip() if row[0] else ""
            nom = str(row[1]).strip() if row[1] else ""
            prenoms = str(row[2]).strip() if row[2] else ""

            # -----------------------------
            # Date flexible
            # -----------------------------
            date_naissance = convertir_date(row[3])

            if not date_naissance:

                erreurs.append(
                    f"Ligne {ligne}: date de naissance invalide."
                )

                continue

            lieu_naissance = str(row[4]).strip() if row[4] else ""
            sexe = str(row[5]).strip().upper() if row[5] else ""
            telephone = str(row[6]).strip() if row[6] else ""
            email = str(row[7]).strip().lower() if row[7] else ""
            classe_nom = str(row[8]).strip() if row[8] else ""
            filiere_nom = str(row[9]).strip() if row[9] else ""

            # -----------------------------
            # Recherche classe
            # -----------------------------
            classe = Classe.objects.filter(
                nom__iexact=classe_nom
            ).first()

            if not classe:

                erreurs.append(
                    f"Ligne {ligne}: classe '{classe_nom}' introuvable."
                )

                continue

            # -----------------------------
            # Recherche filière
            # -----------------------------
            filiere = Filierebts.objects.filter(
                nom__iexact=filiere_nom
            ).first()

            if not filiere:

                erreurs.append(
                    f"Ligne {ligne}: filière '{filiere_nom}' introuvable."
                )

                continue

            # -----------------------------
            # Étudiant existant ?
            # -----------------------------
            etudiant = Etudiant.objects.filter(
                matricule__iexact=matricule
            ).first()

            if etudiant:

                etudiant.nom = nom
                etudiant.prenoms = prenoms
                etudiant.date_naissance = date_naissance
                etudiant.lieu_naissance = lieu_naissance
                etudiant.sexe = sexe
                etudiant.telephone = telephone
                etudiant.email = email
                etudiant.classe = classe
                etudiant.filiere_bts = filiere
                etudiant.save()

                compteur_modification += 1

            else:

                Etudiant.objects.create(
                    matricule=matricule,
                    nom=nom,
                    prenoms=prenoms,
                    date_naissance=date_naissance,
                    lieu_naissance=lieu_naissance,
                    sexe=sexe,
                    telephone=telephone,
                    email=email,
                    classe=classe,
                    filiere_bts=filiere,
                )

                compteur_creation += 1

        except Exception as e:

            erreurs.append(
                f"Ligne {ligne}: erreur inattendue : {e}"
            )

    # ----------------------------------------------------------
    # Messages de résultat
    # ----------------------------------------------------------

    if compteur_creation:
        messages.success(
            request,
            f"{compteur_creation} étudiant(s) créé(s) avec succès."
        )

    if compteur_modification:
        messages.success(
            request,
            f"{compteur_modification} étudiant(s) mis à jour avec succès."
        )

    if erreurs:

        messages.warning(
            request,
            f"{len(erreurs)} ligne(s) n'ont pas été importée(s)."
        )

        for erreur in erreurs:
            messages.warning(request, erreur)

    return redirect("etudiant_list")



def matieres_par_classeAAA(request):
    classe_id = request.GET.get("classe")

    if not classe_id:
        return JsonResponse([], safe=False)

    try:
        classe = Classe.objects.select_related("filiere_bts").get(pk=classe_id)
    except Classe.DoesNotExist:
        return JsonResponse([], safe=False)

    matieres = Matiere.objects.filter(
        filiere_bts=classe.filiere_bts
    ).order_by("libelle")

    data = [
        {
            "id": m.id,
            "code": m.code,
            "libelle": m.libelle,
        }
        for m in matieres
    ]

    return JsonResponse(data, safe=False)

def grande_unite_list(request):
    """HTML du tableau des grandes unités (rechargé dans le modal)."""
    grandes_unites = GrandeUnite.objects.select_related("filiere_bts").prefetch_related("matieres")
    html = render_to_string(
        "matieres/grande_unite_table.html",
        {"grandes_unites": grandes_unites},
        request=request,
    )
    return JsonResponse({"html": html})


@require_http_methods(["GET", "POST"])
def grande_unite_create(request):
    if request.method == "POST":
        form = GrandeUniteForm(request.POST)
        if form.is_valid():
            form.save()
            return JsonResponse({"success": True})
        return JsonResponse({"success": False, "errors": form.errors}, status=400)

    form = GrandeUniteForm()
    html = render_to_string(
        "matieres/grande_unite_form.html",
        {"form": form, "action": "create"},
        request=request,
    )
    return JsonResponse({"html": html})


@require_http_methods(["GET", "POST"])
def grande_unite_update(request, pk):
    grande_unite = get_object_or_404(GrandeUnite, pk=pk)

    if request.method == "POST":
        form = GrandeUniteForm(request.POST, instance=grande_unite)
        if form.is_valid():
            form.save()
            return JsonResponse({"success": True})
        return JsonResponse({"success": False, "errors": form.errors}, status=400)

    form = GrandeUniteForm(instance=grande_unite)
    html = render_to_string(
         "matieres/grande_unite_form.html",
        {"form": form, "action": "update", "grande_unite": grande_unite},
        request=request,
    )
    return JsonResponse({"html": html})


@require_http_methods(["POST"])
def grande_unite_delete(request, pk):
    grande_unite = get_object_or_404(GrandeUnite, pk=pk)
    nb_matieres = grande_unite.matieres.count()

    if nb_matieres > 0:
        return JsonResponse({
            "success": False,
            "error": f"Impossible de supprimer : {nb_matieres} matière(s) rattachée(s)."
        }, status=400)

    grande_unite.delete()
    return JsonResponse({"success": True})


def matieres_par_classe(request):
    """
    ⚠️ Adapte le filtre ci-dessous à ta vraie relation Classe -> Filiere/Niveau.
    Je n'ai pas le modèle Classe donc je pars sur classe.filiere_bts_id.
    """
    classe_id = request.GET.get("classe")

    matieres = Matiere.objects.filter(
        filiere_bts__classe__id=classe_id  # <-- à adapter selon ton modèle Classe
    ).select_related("grande_unite").order_by("grande_unite__ordre", "libelle")

    data = [
        {
            "id": m.id,
            "code": m.code,
            "libelle": m.libelle,
            "grande_unite": m.grande_unite.libelle if m.grande_unite else "Sans grande unité",
            "grande_unite_ordre": m.grande_unite.ordre if m.grande_unite else 999,
        }
        for m in matieres
    ]
    return JsonResponse(data, safe=False)

def grandes_unites_par_filiere(request):
    filiere_id = request.GET.get("filiere_bts")

    grandes_unites = GrandeUnite.objects.filter(
        filiere_bts_id=filiere_id
    ).order_by("ordre")

    data = [
        {"id": gu.id, "code": gu.code, "libelle": gu.libelle}
        for gu in grandes_unites
    ]
    return JsonResponse(data, safe=False)


import io

from django.http import FileResponse
from django.db.models import Q

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)


def export_etudiants_pdf(request):

    filiere_id = request.GET.get("filiere_bts")
    niveau = request.GET.get("niveau")
    recherche = request.GET.get("q", "").strip()

    # =====================================================
    # RÉCUPÉRATION DES ÉTUDIANTS
    # =====================================================

    etudiants = Etudiant.objects.select_related(
        "classe",
        "filiere_bts"
    ).all()

    # Filtre filière
    if filiere_id:
        etudiants = etudiants.filter(
            filiere_bts_id=filiere_id
        )

    # Filtre classe / niveau
    if niveau:
        etudiants = etudiants.filter(
            classe__niveau__nom=niveau
        )

    # Recherche
    if recherche:
        etudiants = etudiants.filter(
            Q(nom__icontains=recherche)
            | Q(prenoms__icontains=recherche)
            | Q(matricule__icontains=recherche)
        )

    etudiants = etudiants.order_by(
        "filiere_bts__nom",
        "classe__nom",
        "nom",
        "prenoms"
    )

    # =====================================================
    # PDF
    # =====================================================

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=0.8 * cm,
        leftMargin=0.8 * cm,
        topMargin=0.8 * cm,
        bottomMargin=0.8 * cm,
    )

    styles = getSampleStyleSheet()

    titre_style = ParagraphStyle(
        "Titre",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=14,
        alignment=1,
        spaceAfter=8,
    )

    sous_titre_style = ParagraphStyle(
        "SousTitre",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        alignment=1,
        spaceAfter=10,
    )

    cell_style = ParagraphStyle(
        "Cell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7,
        leading=8,
    )

    cell_center = ParagraphStyle(
        "CellCenter",
        parent=cell_style,
        alignment=1,
    )

    elements = []

    # =====================================================
    # TITRE
    # =====================================================

    elements.append(
        Paragraph(
            "LISTE DES ÉTUDIANTS BTS",
            titre_style
        )
    )

    # =====================================================
    # FILTRES AFFICHÉS
    # =====================================================

    filiere_nom = "Toutes les filières"

    if filiere_id:
        etudiant_filiere = (
            etudiants.first()
        )

        if etudiant_filiere and etudiant_filiere.filiere_bts:
            filiere_nom = etudiant_filiere.filiere_bts.nom

    niveau_affiche = niveau or "Toutes les années"

    elements.append(
        Paragraph(
            f"<b>Filière :</b> {filiere_nom} "
            f"&nbsp;&nbsp;&nbsp; "
            f"<b>Année :</b> {niveau_affiche} "
            f"&nbsp;&nbsp;&nbsp; "
            f"<b>Total :</b> {etudiants.count()} étudiant(s)",
            sous_titre_style
        )
    )
    header_style = ParagraphStyle(
    "HeaderStyle",
    parent=styles["Normal"],
    fontName="Helvetica-Bold",
    fontSize=8,
    leading=9,
    alignment=1,
    textColor=colors.white,  # 👈 écriture blanche
)

    # =====================================================
    # TABLEAU
    # =====================================================

    data = [
        [
            Paragraph("<b>N°</b>", header_style),
            Paragraph("<b>NOM</b>", header_style),
            Paragraph("<b>PRÉNOMS</b>", header_style),
            Paragraph("<b>MATRICULE</b>", header_style),
            Paragraph("<b>SEXE</b>", header_style),
            Paragraph("<b>CLASSE</b>", header_style),
            Paragraph("<b>FILIÈRE</b>", header_style),
        ]
    ]

    for index, etudiant in enumerate(etudiants, start=1):

        sexe = (
            "Homme"
            if etudiant.sexe == "M"
            else "Femme"
            if etudiant.sexe == "F"
            else "-"
        )

        classe = (
            etudiant.classe.nom
            if etudiant.classe
            else "-"
        )

        filiere = (
            etudiant.filiere_bts.nom
            if etudiant.filiere_bts
            else "-"
        )

        data.append(
            [
                Paragraph(str(index), cell_center),

                Paragraph(
                    etudiant.nom or "",
                    cell_style
                ),

                Paragraph(
                    etudiant.prenoms or "",
                    cell_style
                ),

                Paragraph(
                    etudiant.matricule or "",
                    cell_center
                ),

                Paragraph(
                    sexe,
                    cell_center
                ),

                Paragraph(
                    classe,
                    cell_style
                ),

                Paragraph(
                    filiere,
                    cell_style
                ),
            ]
        )

    table = Table(
        data,
        colWidths=[
            1.0 * cm,
            4.0 * cm,
            6.0 * cm,
            4.0 * cm,
            2.5 * cm,
            4.5 * cm,
            7.0 * cm,
        ],
        repeatRows=1,
    )

    table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#1f3a5f"),
                ),

                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white,
                ),

                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold",
                ),

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.4,
                    colors.grey,
                ),

                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE",
                ),

                (
                    "ALIGN",
                    (0, 0),
                    (0, -1),
                    "CENTER",
                ),

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    4,
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    4,
                ),

                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    4,
                ),

                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    4,
                ),
            ]
        )
    )

    elements.append(table)

    # =====================================================
    # GÉNÉRATION
    # =====================================================

    doc.build(elements)

    buffer.seek(0)

    return FileResponse(
        buffer,
        as_attachment=False,
        filename="liste_etudiants_bts.pdf",
        content_type="application/pdf",
    )

   
    